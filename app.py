from flask import Flask, render_template, request, jsonify, send_file, redirect, url_for, flash, session
import pandas as pd
import os
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash
import json
from datetime import datetime
from psycopg2 import IntegrityError as PgIntegrityError
from db_config import get_db_connection as get_pg_connection, release_db_connection, init_connection_pool
from dotenv import load_dotenv
from flask_login import LoginManager, login_user, logout_user, login_required, current_user
from auth import User, admin_required, usuario_plus_required, module_permission_required, usados_section_required, get_all_users, create_user, update_user, delete_user
from time import time
import uva_utils

# Cargar variables de entorno
load_dotenv()

# Caché global para datos de patentamientos
patentamientos_cache = {
    'data': None,
    'timestamp': 0,
    'ttl': 1800  # 30 minutos de caché
}

app = Flask(__name__)
app.config['UPLOAD_FOLDER'] = 'uploads'
app.config['MAX_CONTENT_LENGTH'] = 16 * 1024 * 1024  # 16MB max file size
app.config['SECRET_KEY'] = os.environ.get('SECRET_KEY', 'clave-secreta-por-defecto-cambiar')
app.config['SEND_FILE_MAX_AGE_DEFAULT'] = 0  # Deshabilitar caché

# Inicializar Flask-Login
login_manager = LoginManager()
login_manager.init_app(app)
login_manager.login_view = 'login'
login_manager.login_message = 'Debes iniciar sesión para acceder a esta página'
login_manager.login_message_category = 'warning'

@login_manager.user_loader
def load_user(user_id):
    return User.get_by_id(user_id)

@app.after_request
def add_header(response):
    """Agregar headers para deshabilitar caché completamente"""
    response.headers['Cache-Control'] = 'no-store, no-cache, must-revalidate, post-check=0, pre-check=0, max-age=0'
    response.headers['Pragma'] = 'no-cache'
    response.headers['Expires'] = '-1'
    return response

# Inicializar pool de conexiones al arrancar
try:
    init_connection_pool()
    print("✅ Conexión a PostgreSQL establecida")
except Exception as e:
    print(f"❌ Error al conectar con PostgreSQL: {e}")

# Asegurar que existe la carpeta de uploads
os.makedirs(app.config['UPLOAD_FOLDER'], exist_ok=True)

ALLOWED_EXTENSIONS = {'xlsx', 'xls'}

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS


# ==================== RUTAS DE AUTENTICACIÓN ====================

@app.route('/login', methods=['GET', 'POST'])
def login():
    """Página de login"""
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    
    if request.method == 'POST':
        username = request.form.get('username')
        password = request.form.get('password')
        
        user = User.get_by_username(username)
        
        if user and user.active and user.check_password(password):
            login_user(user)
            user.update_last_login()
            flash(f'Bienvenido, {user.full_name or user.username}!', 'success')
            
            # Redirigir a la página solicitada o al home
            next_page = request.args.get('next')
            return redirect(next_page if next_page else url_for('home'))
        else:
            flash('Usuario o contraseña incorrectos', 'error')
    
    return render_template('login.html')


@app.route('/logout')
@login_required
def logout():
    """Cerrar sesión"""
    logout_user()
    flash('Has cerrado sesión correctamente', 'info')
    return redirect(url_for('login'))


# ==================== PANEL DE ADMINISTRACIÓN ====================

@app.route('/admin/users')
@login_required
@admin_required
def admin_users():
    """Panel de gestión de usuarios (solo admin)"""
    return render_template('admin_users.html')


# API para gestión de usuarios
@app.route('/api/users', methods=['GET'])
@login_required
@admin_required
def api_get_users():
    """Obtener lista de usuarios"""
    users = get_all_users()
    return jsonify({'success': True, 'users': users})


@app.route('/api/users/<int:user_id>', methods=['GET'])
@login_required
@admin_required
def api_get_user(user_id):
    """Obtener un usuario específico"""
    user = User.get_by_id(user_id)
    if user:
        return jsonify({
            'success': True,
            'user': {
                'id': user.id,
                'username': user.username,
                'role': user.role,
                'full_name': user.full_name,
                'email': user.email,
                'active': user.active
            }
        })
    return jsonify({'success': False, 'error': 'Usuario no encontrado'}), 404


@app.route('/api/users', methods=['POST'])
@login_required
@admin_required
def api_create_user():
    """Crear un nuevo usuario"""
    data = request.json
    result = create_user(
        username=data.get('username'),
        password=data.get('password'),
        role=data.get('role', 'user'),
        full_name=data.get('full_name'),
        email=data.get('email'),
        permiso_planeamiento=data.get('permiso_planeamiento', False),
        permiso_ventas=data.get('permiso_ventas', False),
        permiso_gestoria=data.get('permiso_gestoria', False),
        permiso_entregas=data.get('permiso_entregas', False),
        permiso_bi=data.get('permiso_bi', False),
        permiso_rrhh=data.get('permiso_rrhh', False),
        permiso_usados=data.get('permiso_usados', False),
        permiso_postventa=data.get('permiso_postventa', False)
    )
    return jsonify(result)


@app.route('/api/users/<int:user_id>', methods=['PUT'])
@login_required
@admin_required
def api_update_user(user_id):
    """Actualizar un usuario"""
    data = request.json
    result = update_user(
        user_id=user_id,
        username=data.get('username'),
        password=data.get('password') if data.get('password') else None,
        role=data.get('role'),
        full_name=data.get('full_name'),
        email=data.get('email'),
        active=data.get('active'),
        permiso_planeamiento=data.get('permiso_planeamiento'),
        permiso_ventas=data.get('permiso_ventas'),
        permiso_gestoria=data.get('permiso_gestoria'),
        permiso_entregas=data.get('permiso_entregas'),
        permiso_bi=data.get('permiso_bi'),
        permiso_rrhh=data.get('permiso_rrhh'),
        permiso_usados=data.get('permiso_usados'),
        permiso_postventa=data.get('permiso_postventa')
    )
    return jsonify(result)


@app.route('/api/users/<int:user_id>', methods=['DELETE'])
@login_required
@admin_required
def api_delete_user(user_id):
    """Eliminar un usuario"""
    if user_id == current_user.id:
        return jsonify({'success': False, 'error': 'No puedes eliminar tu propio usuario'}), 400
    
    result = delete_user(user_id)
    return jsonify(result)


# ==================== RUTAS PRINCIPALES ====================

# Ruta raíz: redirigir a login si no está autenticado, o a home si ya lo está
@app.route('/')
def index():
    """Ruta raíz - redirige según estado de autenticación"""
    if current_user.is_authenticated:
        return redirect(url_for('home'))
    return redirect(url_for('login'))

# Página de inicio (selector de aplicaciones)
@app.route('/home')
@login_required
def home():
    return render_template('home.html')

# Aplicación PLANEAMIENTO (requiere permiso)
@app.route('/planeamiento')
@login_required
@module_permission_required('planeamiento')
def planeamiento():
    return render_template('planeamiento.html')

@app.route('/planeamiento/creditos')
@login_required
@module_permission_required('planeamiento')
def planeamiento_creditos():
    return render_template('planeamiento_creditos.html')

# Aplicación VENTAS
@app.route('/ventas')
@login_required
@module_permission_required('ventas')
def ventas():
    return render_template('ventas.html')

@app.route('/ventas/disponible')
@login_required
@module_permission_required('ventas')
def ventas_disponible():
    return render_template('ventas_disponible.html')

@app.route('/ventas/descuentos')
@login_required
@module_permission_required('ventas')
def ventas_descuentos():
    return render_template('ventas_descuentos.html')

@app.route('/ventas/cotizador')
@login_required
@module_permission_required('ventas')
def ventas_cotizador():
    return render_template('ventas_cotizador.html')

@app.route('/ventas/credito_detalle')
@login_required
@module_permission_required('ventas')
def ventas_credito_detalle():
    return render_template('ventas_credito_detalle.html')

@app.route('/diagnostico/uva')
@login_required
def diagnostico_uva():
    """Página de diagnóstico para visualizar datos de CER y UVA"""
    return render_template('diagnostico_uva.html')

# API para obtener datos de UVA
@app.route('/api/uva/actual')
@login_required
def get_uva_actual():
    """Obtener valor actual de UVA"""
    try:
        datos = uva_utils.get_valor_uva_actual()
        return jsonify(datos)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

@app.route('/api/uva/proyeccion/<int:meses>')
@login_required
def get_proyeccion_uva(meses):
    """Obtener proyección de UVA para N meses"""
    try:
        # Tasa en porcentaje (4.0) se convierte a decimal (0.04)
        tasa_porcentaje = request.args.get('tasa', 4.0, type=float)
        tasa_decimal = tasa_porcentaje / 100.0
        proyeccion = uva_utils.proyectar_uva(meses, tasa_decimal)
        return jsonify(proyeccion)
    except Exception as e:
        return jsonify({'error': str(e)}), 500

# Aplicación ENTREGAS
@app.route('/entregas')
@login_required
@module_permission_required('entregas')
def entregas():
    return render_template('entregas.html')

@app.route('/entregas/reportes')
@login_required
@module_permission_required('entregas')
def entregas_reportes():
    return render_template('entregas_reportes.html')


# ==================== MÓDULO USADOS ====================

@app.route('/usados')
@login_required
@module_permission_required('usados')
def usados():
    """Página principal del módulo Usados"""
    return render_template('usados.html', user_role=current_user.role)

@app.route('/usados/stock')
@login_required
@module_permission_required('usados')
@usados_section_required('stock')
def usados_stock():
    """Módulo de Gestión de Stock (KINTO, TEST DRIVE, USADOS)"""
    return render_template('usados_stock.html')

@app.route('/usados/ingresos')
@login_required
@module_permission_required('usados')
@usados_section_required('ingresos')
def usados_ingresos():
    """Módulo de Ingresos de Vehículos Usados"""
    return render_template('usados_ingresos.html')

@app.route('/usados/planificacion')
@login_required
@module_permission_required('usados')
@usados_section_required('planificacion')
def usados_planificacion():
    """Módulo de Planificación de Operaciones (Gantt)"""
    return render_template('usados_planificacion.html')

@app.route('/usados/lavadero')
@login_required
@module_permission_required('usados')
@usados_section_required('lavadero')
def usados_lavadero():
    """Módulo del Lavadero (Panel operarios)"""
    return render_template('usados_lavadero.html')

@app.route('/usados/registro-lavados')
@login_required
@module_permission_required('usados')
def usados_registro_lavados():
    """Módulo de Registro de Lavados (Histórico de operaciones)"""
    return render_template('usados_registro_lavados.html')

@app.route('/usados/reportes')
@login_required
@module_permission_required('usados')
@usados_section_required('reportes')
def usados_reportes():
    """Módulo de Reportes"""
    return render_template('usados_reportes.html')

@app.route('/usados/proveedores')
@login_required
@module_permission_required('usados')
@usados_section_required('proveedores')
def usados_proveedores():
    """Módulo de Proveedores"""
    return render_template('usados_proveedores.html')

# DEPRECADO: Funcionalidad migrada a /usados/stock
# @app.route('/kinto')
# @login_required
# @module_permission_required('usados')
# def kinto():
#     """Módulo de reservas Kinto (autos de alquiler)"""
#     return render_template('kinto.html')

# ==================== JORNADAS LABORALES ====================

@app.route('/api/usados/jornadas', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_jornadas():
    """Obtener configuración de jornadas laborales"""
    conn = None
    try:
        mes = request.args.get('mes')  # formato: YYYY-MM
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        if mes:
            cursor.execute("""
                SELECT id, fecha::TEXT, tipo_jornada, observaciones
                FROM config_jornadas_lavadero
                WHERE TO_CHAR(fecha, 'YYYY-MM') = %s
                ORDER BY fecha ASC
            """, (mes,))
        else:
            cursor.execute("""
                SELECT id, fecha::TEXT, tipo_jornada, observaciones
                FROM config_jornadas_lavadero
                ORDER BY fecha ASC
            """)
        
        jornadas = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(j) for j in jornadas])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/jornadas', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_jornada():
    """Crear o actualizar configuración de jornada"""
    conn = None
    try:
        data = request.get_json()
        fecha = data['fecha']
        tipo_jornada = data['tipo_jornada']  # 'corrida' o 'cortada'
        observaciones = data.get('observaciones', '')
        usuario = current_user.username if current_user.is_authenticated else 'admin'
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Upsert: insertar o actualizar si ya existe
        cursor.execute("""
            INSERT INTO config_jornadas_lavadero 
            (fecha, tipo_jornada, observaciones, usuario_creacion)
            VALUES (%s, %s, %s, %s)
            ON CONFLICT (fecha) 
            DO UPDATE SET 
                tipo_jornada = EXCLUDED.tipo_jornada,
                observaciones = EXCLUDED.observaciones,
                usuario_creacion = EXCLUDED.usuario_creacion
            RETURNING id
        """, (fecha, tipo_jornada, observaciones, usuario))
        
        result = cursor.fetchone()
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': result['id']}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/jornadas/<int:jornada_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
def eliminar_jornada(jornada_id):
    """Eliminar configuración de jornada (vuelve a cortada por defecto)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            DELETE FROM config_jornadas_lavadero
            WHERE id = %s
        """, (jornada_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/jornada/<fecha>', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_jornada_fecha(fecha):
    """Obtener tipo de jornada para una fecha específica"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT tipo_jornada, 
                   hora_inicio_manana::TEXT, hora_fin_manana::TEXT,
                   hora_inicio_tarde::TEXT, hora_fin_tarde::TEXT
            FROM config_jornadas_lavadero
            WHERE fecha = %s
        """, (fecha,))
        
        jornada = cursor.fetchone()
        release_db_connection(conn)
        
        if jornada:
            return jsonify(dict(jornada))
        else:
            # Por defecto: jornada cortada
            return jsonify({
                'tipo_jornada': 'cortada',
                'hora_inicio_manana': '08:30:00',
                'hora_fin_manana': '13:00:00',
                'hora_inicio_tarde': '16:00:00',
                'hora_fin_tarde': '20:00:00'
            })
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# ==================== BUSINESS INTELLIGENCE ANALYSIS ====================

@app.route('/bi_analysis')
@login_required
@module_permission_required('bi')
def bi_analysis():
    """Página principal de Business Intelligence Analysis"""
    return render_template('bi_analysis.html')

@app.route('/bi/bases_datos')
@login_required
@module_permission_required('bi')
def bi_bases_datos():
    """Módulo de carga de bases de datos"""
    return render_template('bi_bases_datos.html')

@app.route('/bi/patentamientos')
@login_required
@module_permission_required('bi')
def bi_patentamientos():
    """Módulo de análisis de patentamientos"""
    return render_template('bi_patentamientos.html')

@app.route('/bi/entregas')
@login_required
@module_permission_required('bi')
def bi_entregas():
    """Módulo de análisis de entregas y plan de negocio"""
    return render_template('bi_entregas.html')

@app.route('/bi/recaudacion')
@login_required
@module_permission_required('bi')
def bi_recaudacion():
    """Módulo de análisis de recaudación (en construcción)"""
    return render_template('bi_recaudacion.html')


# ==================== RECURSOS HUMANOS ====================

@app.route('/rrhh')
@login_required
@module_permission_required('rrhh')
def rrhh():
    """Página principal de Recursos Humanos"""
    # Si es supervisor, redirigir a su perfil
    if current_user.role == 'supervisor':
        return redirect(url_for('rrhh_supervisor_perfil'))
    return render_template('rrhh.html')

@app.route('/rrhh/posiciones')
@login_required
@module_permission_required('rrhh')
def rrhh_posiciones():
    """Módulo de gestión de posiciones"""
    # Supervisores no tienen acceso
    if current_user.role == 'supervisor':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('rrhh'))
    return render_template('rrhh_posiciones.html')

@app.route('/rrhh/empleados')
@login_required
@module_permission_required('rrhh')
def rrhh_empleados():
    """Módulo de gestión de empleados"""
    # Supervisores no tienen acceso
    if current_user.role == 'supervisor':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('rrhh'))
    return render_template('rrhh_empleados.html')

# API: Obtener todas las posiciones
@app.route('/api/rrhh/posiciones', methods=['GET'])
@login_required
@module_permission_required('rrhh')
def get_posiciones():
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT id, nombre, premio_toyota, activo, fecha_creacion, fecha_actualizacion
            FROM posiciones
            ORDER BY nombre
        ''')
        posiciones = cursor.fetchall()
        conn.rollback()
        return jsonify(posiciones)
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Crear nueva posición
@app.route('/api/rrhh/posiciones', methods=['POST'])
@login_required
@module_permission_required('rrhh')
def create_posicion():
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO posiciones (nombre, premio_toyota, activo)
            VALUES (%s, %s, %s)
            RETURNING id, nombre, premio_toyota, activo, fecha_creacion, fecha_actualizacion
        ''', (data['nombre'], data.get('premio_toyota', 0), data.get('activo', True)))
        
        nueva_posicion = cursor.fetchone()
        conn.commit()
        return jsonify(nueva_posicion), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Actualizar posición
@app.route('/api/rrhh/posiciones/<int:id>', methods=['PUT'])
@login_required
@module_permission_required('rrhh')
def update_posicion(id):
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            UPDATE posiciones
            SET nombre = %s, premio_toyota = %s, activo = %s, fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, nombre, premio_toyota, activo, fecha_creacion, fecha_actualizacion
        ''', (data['nombre'], data['premio_toyota'], data['activo'], id))
        
        posicion_actualizada = cursor.fetchone()
        conn.commit()
        
        if posicion_actualizada:
            return jsonify(posicion_actualizada)
        else:
            return jsonify({'error': 'Posición no encontrada'}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Eliminar posición
@app.route('/api/rrhh/posiciones/<int:id>', methods=['DELETE'])
@login_required
@module_permission_required('rrhh')
def delete_posicion(id):
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM posiciones WHERE id = %s', (id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Obtener todos los empleados
@app.route('/api/rrhh/empleados', methods=['GET'])
@login_required
@module_permission_required('rrhh')
def get_empleados():
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT e.id, e.legajo, e.nombre_completo, e.posicion_id, 
                   p.nombre as posicion_nombre, p.premio_toyota,
                   e.fecha_alta, e.fecha_baja, e.activo, e.observaciones,
                   s.nombre_completo as inmediato_superior,
                   e.supervisor_id,
                   e.fecha_creacion, e.fecha_actualizacion
            FROM empleados e
            LEFT JOIN posiciones p ON e.posicion_id = p.id
            LEFT JOIN supervisores s ON e.supervisor_id = s.id
            ORDER BY e.nombre_completo
        ''')
        empleados = cursor.fetchall()
        conn.rollback()
        return jsonify(empleados)
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Crear nuevo empleado
@app.route('/api/rrhh/empleados', methods=['POST'])
@login_required
@module_permission_required('rrhh')
def create_empleado():
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            INSERT INTO empleados (legajo, nombre_completo, posicion_id, fecha_alta, activo, observaciones, inmediato_superior)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, legajo, nombre_completo, posicion_id, fecha_alta, fecha_baja, activo, observaciones, inmediato_superior
        ''', (
            data['legajo'], 
            data['nombre_completo'], 
            data.get('posicion_id') or None,
            data.get('fecha_alta') or None,
            data.get('activo', True),
            data.get('observaciones', ''),
            data.get('inmediato_superior', '')
        ))
        
        nuevo_empleado = cursor.fetchone()
        conn.commit()
        return jsonify(nuevo_empleado), 201
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Actualizar empleado
@app.route('/api/rrhh/empleados/<int:id>', methods=['PUT'])
@login_required
@module_permission_required('rrhh')
def update_empleado(id):
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            UPDATE empleados
            SET legajo = %s, nombre_completo = %s, posicion_id = %s, 
                fecha_alta = %s, fecha_baja = %s, activo = %s, 
                observaciones = %s, inmediato_superior = %s, fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, legajo, nombre_completo, posicion_id, fecha_alta, fecha_baja, activo, observaciones, inmediato_superior
        ''', (
            data['legajo'],
            data['nombre_completo'],
            data.get('posicion_id') or None,
            data.get('fecha_alta') or None,
            data.get('fecha_baja') or None,
            data['activo'],
            data.get('observaciones', ''),
            data.get('inmediato_superior', ''),
            id
        ))
        
        empleado_actualizado = cursor.fetchone()
        conn.commit()
        
        if empleado_actualizado:
            return jsonify(empleado_actualizado)
        else:
            return jsonify({'error': 'Empleado no encontrado'}), 404
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Eliminar empleado
@app.route('/api/rrhh/empleados/<int:id>', methods=['DELETE'])
@login_required
@module_permission_required('rrhh')
def delete_empleado(id):
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM empleados WHERE id = %s', (id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# ====== SUPERVISORES ======

@app.route('/rrhh/supervisor/perfil')
@login_required
@module_permission_required('rrhh')
def rrhh_supervisor_perfil():
    """Perfil del supervisor actual"""
    # Solo supervisores pueden acceder
    if current_user.role != 'supervisor':
        flash('Esta sección es solo para supervisores', 'error')
        return redirect(url_for('rrhh'))
    return render_template('rrhh_supervisor_perfil.html')

@app.route('/rrhh/supervisores')
@login_required
@module_permission_required('rrhh')
def rrhh_supervisores():
    """Módulo de gestión de supervisores"""
    # Supervisores no tienen acceso
    if current_user.role == 'supervisor':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('rrhh'))
    return render_template('rrhh_supervisores.html')

@app.route('/rrhh/liquidaciones')
@login_required
@module_permission_required('rrhh')
def rrhh_liquidaciones():
    """Módulo de liquidaciones - consolidado de evaluaciones"""
    # Supervisores no tienen acceso
    if current_user.role == 'supervisor':
        flash('No tienes permisos para acceder a esta sección', 'error')
        return redirect(url_for('rrhh'))
    return render_template('rrhh_liquidaciones.html')

@app.route('/rrhh/supervisor/<int:supervisor_id>')
@login_required
@module_permission_required('rrhh')
def ver_supervisor(supervisor_id):
    """Perfil del supervisor y su equipo con evaluaciones (para RRHH)"""
    return render_template('rrhh_ver_supervisor.html')

# API: Obtener datos del supervisor y su equipo
@app.route('/api/rrhh/supervisor/<int:supervisor_id>')
@login_required
@module_permission_required('rrhh')
def api_supervisor_perfil(supervisor_id):
    """Obtener información del supervisor y su equipo"""
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        # Obtener datos del supervisor
        cursor.execute('''
            SELECT id, nombre_completo, activo, fecha_creacion
            FROM supervisores
            WHERE id = %s
        ''', (supervisor_id,))
        supervisor = cursor.fetchone()
        
        if not supervisor:
            conn.rollback()
            return jsonify({'error': 'Supervisor no encontrado'}), 404
        
        # Obtener empleados a cargo
        cursor.execute('''
            SELECT 
                e.id,
                e.legajo,
                e.nombre_completo,
                p.nombre as posicion_nombre,
                p.premio_toyota,
                e.fecha_alta,
                e.activo
            FROM empleados e
            LEFT JOIN posiciones p ON e.posicion_id = p.id
            WHERE e.supervisor_id = %s
            ORDER BY e.activo DESC, e.legajo
        ''', (supervisor_id,))
        
        # RealDictRow ya devuelve diccionarios, solo convertir a dict normales
        equipo_rows = cursor.fetchall()
        equipo = [dict(row) for row in equipo_rows]
        
        conn.rollback()
        return jsonify({
            'supervisor': dict(supervisor),
            'equipo': equipo
        })
    except Exception as e:
        conn.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print("=" * 80)
        print("ERROR en /api/rrhh/supervisor:")
        print(error_trace)
        print("=" * 80)
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Obtener datos del supervisor actual (para supervisores)
@app.route('/api/rrhh/mi-perfil-supervisor', methods=['GET'])
@login_required
@module_permission_required('rrhh')
def get_mi_perfil_supervisor():
    """Obtener datos del supervisor vinculado al usuario actual"""
    if current_user.role != 'supervisor':
        return jsonify({'error': 'Solo supervisores pueden acceder a esta información'}), 403
    
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT s.id, s.nombre_completo, s.email, s.telefono, s.area, s.activo, 
                   s.observaciones, s.user_id,
                   u.username, u.active as user_active
            FROM supervisores s
            LEFT JOIN users u ON s.user_id = u.id
            WHERE s.user_id = %s
        ''', (current_user.id,))
        
        supervisor = cursor.fetchone()
        conn.rollback()
        
        if not supervisor:
            return jsonify({'error': 'No se encontró un perfil de supervisor vinculado a tu usuario'}), 404
        
        return jsonify({'supervisor': supervisor})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Obtener todos los supervisores
@app.route('/api/rrhh/supervisores', methods=['GET'])
@login_required
@module_permission_required('rrhh')
def get_supervisores():
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT s.id, s.nombre_completo, s.email, s.telefono, s.area, s.activo, 
                   s.observaciones, s.fecha_creacion, s.fecha_actualizacion, s.user_id,
                   u.username, u.active as user_active
            FROM supervisores s
            LEFT JOIN users u ON s.user_id = u.id
            ORDER BY s.nombre_completo
        ''')
        supervisores = cursor.fetchall()
        conn.rollback()
        return jsonify(supervisores)
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Crear nuevo supervisor
@app.route('/api/rrhh/supervisores', methods=['POST'])
@login_required
@module_permission_required('rrhh')
def create_supervisor():
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        user_id = None
        
        # Si se proporciona información de usuario, crear el usuario
        if data.get('crear_usuario') and data.get('username'):
            cursor.execute('''
                INSERT INTO users (username, password_hash, role, full_name, email, active)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                data['username'],
                generate_password_hash(data.get('password', 'supervisor123')),
                'supervisor',
                data['nombre_completo'],
                data.get('email'),
                True
            ))
            user_result = cursor.fetchone()
            user_id = user_result['id']
        
        cursor.execute('''
            INSERT INTO supervisores (nombre_completo, email, telefono, area, activo, observaciones, user_id)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id, nombre_completo, email, telefono, area, activo, observaciones, user_id
        ''', (
            data['nombre_completo'],
            data.get('email') or None,
            data.get('telefono') or None,
            data.get('area') or None,
            data.get('activo', True),
            data.get('observaciones') or None,
            user_id
        ))
        
        nuevo_supervisor = cursor.fetchone()
        conn.commit()
        
        # Recargar con información del usuario si existe
        if nuevo_supervisor and nuevo_supervisor.get('user_id'):
            cursor.execute('''
                SELECT s.*, u.username, u.active as user_active
                FROM supervisores s
                LEFT JOIN users u ON s.user_id = u.id
                WHERE s.id = %s
            ''', (nuevo_supervisor['id'],))
            nuevo_supervisor = cursor.fetchone()
        
        return jsonify(nuevo_supervisor), 201
    except Exception as e:
        conn.rollback()
        print(f"Error creando supervisor: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Actualizar supervisor
@app.route('/api/rrhh/supervisores/<int:id>', methods=['PUT'])
@login_required
@module_permission_required('rrhh')
def update_supervisor(id):
    data = request.json
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        # Obtener supervisor actual
        cursor.execute('SELECT user_id FROM supervisores WHERE id = %s', (id,))
        supervisor_actual = cursor.fetchone()
        
        if not supervisor_actual:
            return jsonify({'error': 'Supervisor no encontrado'}), 404
        
        user_id = supervisor_actual['user_id']
        
        # Si se solicita crear usuario y no tiene uno
        if data.get('crear_usuario') and data.get('username') and not user_id:
            cursor.execute('''
                INSERT INTO users (username, password_hash, role, full_name, email, active)
                VALUES (%s, %s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                data['username'],
                generate_password_hash(data.get('password', 'supervisor123')),
                'supervisor',
                data['nombre_completo'],
                data.get('email'),
                True
            ))
            user_result = cursor.fetchone()
            user_id = user_result['id']
        
        # Actualizar supervisor
        cursor.execute('''
            UPDATE supervisores
            SET nombre_completo = %s, email = %s, telefono = %s, area = %s, 
                activo = %s, observaciones = %s, user_id = %s, fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = %s
            RETURNING id, nombre_completo, email, telefono, area, activo, observaciones, user_id
        ''', (
            data['nombre_completo'],
            data.get('email') or None,
            data.get('telefono') or None,
            data.get('area') or None,
            data['activo'],
            data.get('observaciones') or None,
            user_id,
            id
        ))
        
        supervisor_actualizado = cursor.fetchone()
        conn.commit()
        
        if supervisor_actualizado:
            # Recargar con información del usuario si existe
            if supervisor_actualizado.get('user_id'):
                cursor.execute('''
                    SELECT s.*, u.username, u.active as user_active
                    FROM supervisores s
                    LEFT JOIN users u ON s.user_id = u.id
                    WHERE s.id = %s
                ''', (supervisor_actualizado['id'],))
                supervisor_actualizado = cursor.fetchone()
            
            return jsonify(supervisor_actualizado)
        else:
            return jsonify({'error': 'Supervisor no encontrado'}), 404
    except Exception as e:
        conn.rollback()
        print(f"Error actualizando supervisor: {e}")
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Eliminar supervisor
@app.route('/api/rrhh/supervisores/<int:id>', methods=['DELETE'])
@login_required
@module_permission_required('rrhh')
def delete_supervisor(id):
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('DELETE FROM supervisores WHERE id = %s', (id,))
        conn.commit()
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# ====== EVALUACIONES MENSUALES ======

# API: Obtener evaluaciones de un período
@app.route('/api/rrhh/evaluaciones', methods=['GET'])
@login_required
@module_permission_required('rrhh')
def get_evaluaciones():
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    supervisor_id = request.args.get('supervisor_id', type=int)
    
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        query = '''
            SELECT e.*, emp.nombre_completo as empleado_nombre, emp.legajo
            FROM evaluaciones_mensuales e
            JOIN empleados emp ON e.empleado_id = emp.id
            WHERE e.mes = %s AND e.anio = %s
        '''
        params = [mes, anio]
        
        if supervisor_id:
            query += ' AND e.supervisor_id = %s'
            params.append(supervisor_id)
        
        cursor.execute(query, params)
        evaluaciones = cursor.fetchall()
        conn.rollback()
        return jsonify(evaluaciones)
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

# API: Guardar evaluaciones (batch)
@app.route('/api/rrhh/evaluaciones/guardar', methods=['POST'])
@login_required
@module_permission_required('rrhh')
def guardar_evaluaciones():
    data = request.json
    evaluaciones = data.get('evaluaciones', [])
    
    if not evaluaciones:
        return jsonify({'error': 'No se proporcionaron evaluaciones'}), 400
    
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        guardadas = 0
        for ev in evaluaciones:
            cursor.execute('''
                INSERT INTO evaluaciones_mensuales 
                (empleado_id, supervisor_id, mes, anio, performance, comisiones, 
                 anticipo_comisiones, horas_extras_50, horas_extras_100)
                VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s)
                ON CONFLICT (empleado_id, mes, anio)
                DO UPDATE SET
                    supervisor_id = EXCLUDED.supervisor_id,
                    performance = EXCLUDED.performance,
                    comisiones = EXCLUDED.comisiones,
                    anticipo_comisiones = EXCLUDED.anticipo_comisiones,
                    horas_extras_50 = EXCLUDED.horas_extras_50,
                    horas_extras_100 = EXCLUDED.horas_extras_100,
                    fecha_actualizacion = CURRENT_TIMESTAMP
            ''', (
                ev['empleado_id'], ev['supervisor_id'], ev['mes'], ev['anio'],
                ev['performance'], ev['comisiones'], ev['anticipo_comisiones'],
                ev['horas_extras_50'], ev['horas_extras_100']
            ))
            guardadas += 1
        
        conn.commit()
        return jsonify({'success': True, 'guardadas': guardadas})
    except Exception as e:
        conn.rollback()
        return jsonify({'error': str(e)}), 500
    finally:
        release_db_connection(conn)

@app.route('/api/rrhh/liquidaciones')
@login_required
@module_permission_required('rrhh')
def api_liquidaciones():
    """Obtener consolidado de empleados con evaluaciones del período"""
    mes = request.args.get('mes', type=int)
    anio = request.args.get('anio', type=int)
    
    if not mes or not anio:
        return jsonify({'error': 'Debe especificar mes y año'}), 400
    
    conn = get_pg_connection()
    cursor = conn.cursor()
    try:
        cursor.execute('''
            SELECT 
                e.id,
                e.legajo,
                e.nombre_completo,
                p.nombre as posicion,
                s.nombre_completo as inmediato_superior,
                e.supervisor_id,
                p.premio_toyota,
                ev.performance,
                ev.comisiones,
                ev.anticipo_comisiones,
                ev.horas_extras_50,
                ev.horas_extras_100,
                CASE WHEN ev.id IS NOT NULL THEN true ELSE false END as tiene_evaluacion
            FROM empleados e
            LEFT JOIN posiciones p ON e.posicion_id = p.id
            LEFT JOIN supervisores s ON e.supervisor_id = s.id
            LEFT JOIN evaluaciones_mensuales ev 
                ON e.id = ev.empleado_id 
                AND ev.mes = %s 
                AND ev.anio = %s
            WHERE e.activo = true
            ORDER BY e.legajo
        ''', (mes, anio))
        
        # RealDictRow ya devuelve diccionarios, solo convertir a dict normales
        rows = cursor.fetchall()
        
        print(f"DEBUG: Total filas: {len(rows)}")
        if rows:
            print(f"DEBUG: Primera fila tipo: {type(rows[0])}")
            print(f"DEBUG: Primera fila: {dict(rows[0])}")
        
        # Convertir RealDictRow a dict normales
        empleados = [dict(row) for row in rows]
        
        print(f"DEBUG: Primer empleado: {empleados[0] if empleados else 'ninguno'}")
        
        conn.rollback()
        return jsonify(empleados)
    except Exception as e:
        conn.rollback()
        import traceback
        error_trace = traceback.format_exc()
        print("=" * 80)
        print("ERROR en /api/rrhh/liquidaciones:")
        print(error_trace)
        print("=" * 80)
        return jsonify({'error': str(e), 'trace': error_trace}), 500
    finally:
        release_db_connection(conn)


# ==================== FUNCIONES DE BASE DE DATOS ====================

def get_db_connection():
    """Crear conexión a la base de datos PostgreSQL"""
    return get_pg_connection()


# Módulo 1: Lista de precios
@app.route('/modulo1')
@login_required
@module_permission_required('planeamiento')
def modulo1():
    return render_template('modulo1_nuevo.html')

# API para obtener precios
@app.route('/api/precios', methods=['GET'])
@login_required
def get_precios():
    # Permitir acceso a usuarios con permiso de planeamiento o ventas (solo lectura)
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        flash('No tienes permisos para acceder a esta función', 'error')
        return jsonify({'error': 'Acceso denegado'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT modelo, precio_ars, precio_usd, cotizacion, descuento, descuento_futuro, visible, dado_baja, familia, id_modelo, id_salesforce FROM precios ORDER BY familia, modelo')
    rows = cursor.fetchall()
    conn.rollback()
    release_db_connection(conn)
    
    modelos = []
    modelos_ocultos = []
    
    for row in rows:
        modelo_data = {
            'nombre': row['modelo'],
            'precio_ars': row['precio_ars'],
            'precio_usd': row['precio_usd'],
            'cotizacion': row['cotizacion'],
            'descuento': row['descuento'],
            'descuento_futuro': row.get('descuento_futuro', 0) or 0,
            'dado_baja': row['dado_baja'],
            'familia': row['familia'] or 'OTROS',
            'id_modelo': row.get('id_modelo', '') or '',
            'id_salesforce': row.get('id_salesforce', '') or ''
        }
        modelos.append(modelo_data)
        
        if row['visible'] == 0:
            modelos_ocultos.append(row['modelo'])
    
    return jsonify({
        'modelos': modelos,
        'modelos_ocultos': modelos_ocultos
    })

# API para guardar precios
@app.route('/api/precios', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_precios():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Actualizar precios de TODOS los modelos (convencionales y SC)
        for modelo in data.get('modelos', []):
            cursor.execute('''
                UPDATE precios 
                SET precio_ars = %s, precio_usd = %s, cotizacion = %s, descuento = %s, descuento_futuro = %s, dado_baja = %s, id_modelo = %s, id_salesforce = %s, fecha_actualizacion = CURRENT_TIMESTAMP
                WHERE modelo = %s
            ''', (modelo['precio_ars'], modelo['precio_usd'], modelo['cotizacion'], modelo['descuento'], modelo.get('descuento_futuro', 0), modelo.get('dado_baja', 0), modelo.get('id_modelo', ''), modelo.get('id_salesforce', ''), modelo['nombre']))
        
        # Actualizar visibilidad
        # Primero poner todos como visibles
        cursor.execute('UPDATE precios SET visible = 1')
        
        # Luego ocultar los que están en la lista
        for modelo_oculto in data.get('modelos_ocultos', []):
            cursor.execute('UPDATE precios SET visible = 0 WHERE modelo = %s', (modelo_oculto,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para aplicar descuento por familia
@app.route('/api/descuento_familia', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def aplicar_descuento_familia():
    data = request.json
    familia = data.get('familia')
    descuento = data.get('descuento', 0)
    
    if not familia:
        return jsonify({'success': False, 'error': 'Familia no especificada'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Contar modelos afectados
        cursor.execute('SELECT COUNT(*) as total FROM precios WHERE familia = %s', (familia,))
        total = cursor.fetchone()['total']
        
        # Aplicar descuento
        cursor.execute('UPDATE precios SET descuento = %s WHERE familia = %s', (descuento, familia))
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'modelos_actualizados': total})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para obtener descuentos adicionales
@app.route('/api/descuentos_adicionales', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_descuentos_adicionales():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT tipo, clave, valor FROM descuentos_adicionales')
    rows = cursor.fetchall()
    release_db_connection(conn)
    
    descuentos = {}
    for row in rows:
        if row['tipo'] not in descuentos:
            descuentos[row['tipo']] = {}
        descuentos[row['tipo']][row['clave']] = row['valor']
    
    return jsonify(descuentos)

# API para guardar descuentos adicionales
@app.route('/api/descuentos_adicionales', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_descuentos_adicionales():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        for tipo, valores in data.items():
            for clave, valor in valores.items():
                cursor.execute('''
                    INSERT INTO descuentos_adicionales (tipo, clave, valor, fecha_actualizacion)
                    VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                    ON CONFLICT (tipo, clave) 
                    DO UPDATE SET valor = EXCLUDED.valor, fecha_actualizacion = CURRENT_TIMESTAMP
                ''', (tipo, clave, valor))
        
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para obtener descuentos matriz (fecha/ubicación)
@app.route('/api/descuentos_matriz', methods=['GET'])
@login_required
def get_descuentos_matriz():
    # Permitir acceso a usuarios con permiso de planeamiento o ventas (solo lectura)
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        flash('No tienes permisos para acceder a esta función', 'error')
        return jsonify({'error': 'Acceso denegado'}), 403
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT modelo, 
               desc_mes_actual_menos_2, desc_mes_actual_menos_1, desc_mes_actual, desc_mes_actual_mas,
               desc_stock, desc_produccion, desc_playa_externa, desc_otro
        FROM descuentos_matriz
        ORDER BY modelo
    ''')
    rows = cursor.fetchall()
    release_db_connection(conn)
    
    descuentos = [dict(row) for row in rows]
    return jsonify(descuentos)

# API para guardar descuentos matriz
@app.route('/api/descuentos_matriz', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_descuentos_matriz():
    data = request.json  # Lista de objetos con modelo y sus descuentos
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        for item in data:
            cursor.execute('''
                UPDATE descuentos_matriz 
                SET desc_mes_actual_menos_2 = %s,
                    desc_mes_actual_menos_1 = %s,
                    desc_mes_actual = %s,
                    desc_mes_actual_mas = %s,
                    desc_stock = %s,
                    desc_produccion = %s,
                    desc_playa_externa = %s,
                    desc_otro = %s,
                    fecha_actualizacion = CURRENT_TIMESTAMP
                WHERE modelo = %s
            ''', (
                item.get('desc_mes_actual_menos_2', 0),
                item.get('desc_mes_actual_menos_1', 0),
                item.get('desc_mes_actual', 0),
                item.get('desc_mes_actual_mas', 0),
                item.get('desc_stock', 0),
                item.get('desc_produccion', 0),
                item.get('desc_playa_externa', 0),
                item.get('desc_otro', 0),
                item['modelo']
            ))
        
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# ==================== API PLANES DE CRÉDITO ====================

@app.route('/api/planes_credito', methods=['GET'])
@login_required
def get_planes_credito():
    """Obtener todos los planes de crédito"""
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        return jsonify({'error': 'No autorizado'}), 403
    
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nombre_plan, tipo_credito, tasa_inflacion_mensual, importe_maximo, plazos, tasas, quebrantos, activo, fecha_creacion
            FROM planes_credito
            ORDER BY activo DESC, nombre_plan ASC
        """)
        planes = cursor.fetchall()
        release_db_connection(conn)
        
        return jsonify([dict(plan) for plan in planes])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/planes_credito', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def crear_plan_credito():
    """Crear un nuevo plan de crédito"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO planes_credito (nombre_plan, tipo_credito, tasa_inflacion_mensual, importe_maximo, plazos, tasas, quebrantos, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s, TRUE)
            RETURNING id
        """, (
            data['nombre_plan'],
            data.get('tipo_credito', 'pesos'),
            data.get('tasa_inflacion'),
            data.get('importe_maximo', 0),
            json.dumps(data['plazos']),
            json.dumps(data['tasas']),
            json.dumps(data['quebrantos'])
        ))
        
        plan_id = cursor.fetchone()['id']
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': plan_id}), 201
    except PgIntegrityError as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': 'Ya existe un plan con ese nombre'}), 400
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/planes_credito/<int:plan_id>', methods=['PUT'])
@login_required
@module_permission_required('planeamiento')
def actualizar_plan_credito(plan_id):
    """Actualizar un plan de crédito existente"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE planes_credito
            SET nombre_plan = %s, tipo_credito = %s, tasa_inflacion_mensual = %s, importe_maximo = %s, plazos = %s, tasas = %s, quebrantos = %s,
                fecha_actualizacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (
            data['nombre_plan'],
            data.get('tipo_credito', 'pesos'),
            data.get('tasa_inflacion'),
            data.get('importe_maximo', 0),
            json.dumps(data['plazos']),
            json.dumps(data['tasas']),
            json.dumps(data['quebrantos']),
            plan_id
        ))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except PgIntegrityError as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': 'Ya existe un plan con ese nombre'}), 400
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/planes_credito/<int:plan_id>', methods=['DELETE'])
@login_required
@module_permission_required('planeamiento')
def eliminar_plan_credito(plan_id):
    """Eliminar un plan de crédito"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute('DELETE FROM planes_credito WHERE id = %s', (plan_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# API para obtener unidades postergadas
@app.route('/api/unidades_postergadas', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_unidades_postergadas():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT numero_fabrica FROM unidades_postergadas ORDER BY fecha_agregado DESC')
    rows = cursor.fetchall()
    release_db_connection(conn)
    
    unidades = [row['numero_fabrica'] for row in rows]
    return jsonify(unidades)

# API para agregar unidad postergada
@app.route('/api/unidades_postergadas', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def add_unidad_postergada():
    data = request.json
    numero_fabrica = data.get('numero_fabrica', '').strip()
    
    if not numero_fabrica:
        return jsonify({'success': False, 'error': 'Número de fábrica vacío'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO unidades_postergadas (numero_fabrica) VALUES (%s)', (numero_fabrica,))
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except PgIntegrityError:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': 'Número de fábrica ya existe'}), 400
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para eliminar unidad postergada
@app.route('/api/unidades_postergadas/<numero_fabrica>', methods=['DELETE'])
@login_required
@module_permission_required('planeamiento')
def delete_unidad_postergada(numero_fabrica):
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM unidades_postergadas WHERE numero_fabrica = %s', (numero_fabrica,))
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


# Módulo 2: Tabla editable (a desarrollar)
@app.route('/modulo2')
@login_required
@module_permission_required('planeamiento')
def modulo2():
    return render_template('modulo2.html')


# ==================== API MÓDULO 2: PREVENTA ====================

# API para obtener preventa
@app.route('/api/preventa', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_preventa():
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('''
        SELECT id, numero_fabrica, modelo_version, operacion, vendedor, color, informado, cancelado, asignado 
        FROM preventa 
        ORDER BY id ASC
    ''')
    rows = cursor.fetchall()
    release_db_connection(conn)
    
    preventa = []
    for row in rows:
        preventa.append({
            'id': row['id'],
            'numero_fabrica': row['numero_fabrica'],
            'modelo_version': row['modelo_version'],
            'operacion': row['operacion'],
            'vendedor': row['vendedor'],
            'color': row['color'],
            'informado': row['informado'],
            'cancelado': row['cancelado'],
            'asignado': row['asignado']
        })
    
    return jsonify(preventa)

# API para guardar/actualizar preventa (reemplaza toda la tabla)
@app.route('/api/preventa', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_preventa():
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Paso 1: Eliminar TODAS las unidades de preventa del módulo disponible
        # Esto asegura que si la bitácora se limpia, el disponible también se limpia
        cursor.execute('''
            DELETE FROM disponibles 
            WHERE numero_fabrica = 'YAC999999999'
        ''')
        print(f"🗑️ Unidades de preventa eliminadas del disponible")
        
        # Paso 2: Eliminar todos los registros de preventa
        cursor.execute('DELETE FROM preventa')
        
        # Paso 3: Insertar nuevos registros (solo si hay datos)
        if len(data) > 0:
            for item in data:
                cursor.execute('''
                    INSERT INTO preventa (
                        numero_fabrica, modelo_version, operacion, vendedor, color, informado, cancelado, asignado
                    ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s)
                ''', (
                    'YAC999999999',  # Número de fábrica fijo para preventa
                    item.get('modelo_version', ''),
                    item.get('operacion', ''),
                    item.get('vendedor', ''),
                    item.get('color', ''),
                    1 if item.get('informado') else 0,
                    1 if item.get('cancelado') else 0,
                    1 if item.get('asignado') else 0
                ))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'count': len(data), 'message': 'Bitácora guardada y disponible actualizado'})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para convertir preventa sin vendedor a disponible
@app.route('/api/preventa/convertir_disponible', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def convertir_preventa_disponible():
    """Convierte unidades de preventa SIN vendedor al módulo disponible"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Paso 1: Limpiar TODAS las unidades de preventa anteriores del disponible
        cursor.execute('''
            DELETE FROM disponibles 
            WHERE numero_fabrica = 'YAC999999999'
        ''')
        print(f"🗑️ Limpieza: Unidades de preventa anteriores eliminadas del disponible")
        
        # Paso 2: Obtener registros de preventa sin vendedor asignado
        cursor.execute('''
            SELECT numero_fabrica, modelo_version, operacion, color
            FROM preventa
            WHERE vendedor IS NULL OR vendedor = ''
        ''')
        preventas = cursor.fetchall()
        
        if len(preventas) == 0:
            conn.commit()  # Confirmar la limpieza aunque no haya nada que agregar
            release_db_connection(conn)
            return jsonify({'success': True, 'count': 0, 'message': 'No hay unidades sin vendedor. Preventa anterior limpiada del disponible.'})
        
        # Calcular fecha de entrega estimada (3 meses adelante)
        from datetime import datetime, timedelta
        fecha_entrega = datetime.now() + timedelta(days=90)  # 3 meses
        fecha_entrega_str = fecha_entrega.strftime('%Y-%m-%d')
        
        # Obtener precios de la base de datos
        cursor.execute('SELECT modelo, precio_ars, descuento FROM precios')
        precios_data = {}
        for row in cursor.fetchall():
            precios_data[row['modelo']] = {
                'precio_ars': row['precio_ars'],
                'descuento': row['descuento']
            }
        
        # Insertar en disponibles
        count = 0
        for prev in preventas:
            modelo = prev['modelo_version']
            
            # Obtener precio BASE (sin descuento)
            # Los descuentos se aplicarán en el Módulo 4
            precio = 0
            if modelo in precios_data:
                precio = precios_data[modelo]['precio_ars']
            
            cursor.execute('''
                INSERT INTO disponibles (
                    numero_fabrica, modelo_version, color, ubicacion, 
                    entrega_estimada, precio_disponible, operacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s)
            ''', (
                prev['numero_fabrica'],  # Usar el número de fábrica de preventa (YAC999999999)
                modelo,
                prev['color'],
                "Preventa",
                fecha_entrega_str,
                precio,
                prev['operacion']
            ))
            count += 1
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'count': count, 'message': f'{count} unidades de preventa agregadas a disponible'})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


# Módulo 3: Procesador de Excel
@app.route('/modulo3')
@login_required
@module_permission_required('planeamiento')
def modulo3():
    return render_template('modulo3.html')


@app.route('/procesar_excel', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def procesar_excel():
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se encontró ningún archivo'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó ningún archivo'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'error': 'Tipo de archivo no permitido. Use .xlsx o .xls'}), 400
        
        # Obtener lista de unidades postergadas desde la base de datos
        conn = get_db_connection()
        cursor = conn.cursor()
        cursor.execute('SELECT numero_fabrica FROM unidades_postergadas')
        rows = cursor.fetchall()
        release_db_connection(conn)
        
        unidades_postergadas = [row['numero_fabrica'] for row in rows]
        print(f"Unidades postergadas desde BD: {unidades_postergadas}")
        
        # Guardar archivo temporalmente
        filename = secure_filename(file.filename)
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        
        # Debug: verificar permisos y ruta
        print(f"📁 Intentando guardar en: {filepath}")
        print(f"📁 Directorio uploads existe: {os.path.exists(app.config['UPLOAD_FOLDER'])}")
        print(f"📁 Permisos del directorio: {oct(os.stat(app.config['UPLOAD_FOLDER']).st_mode)}")
        
        try:
            file.save(filepath)
            print(f"✅ Archivo guardado exitosamente en {filepath}")
        except Exception as save_error:
            print(f"❌ Error al guardar archivo: {save_error}")
            return jsonify({'error': f'Error al guardar archivo: {str(save_error)}'}), 500
        
        # Leer el archivo Excel
        df = pd.read_excel(filepath, header=None)
        
        # Paso 1: Eliminar las primeras 8 filas
        df = df.iloc[8:]
        
        # Paso 2: Eliminar la columna C (índice 2, ya que comienza en 0)
        if df.shape[1] > 2:  # Verificar que existe la columna C
            df = df.drop(df.columns[2], axis=1)
        
        # Paso 3: Eliminar filas completamente vacías
        df = df.dropna(how='all')
        
        # Paso 4: Ordenar por la primera columna (columna A)
        df = df.sort_values(by=df.columns[0], ascending=True)
        
        # Resetear el índice
        df = df.reset_index(drop=True)
        
        # Nombres de las columnas (primeras 15 columnas visibles)
        column_names = [
            'Nº Fábrica', 'Nº Chasis', 'Modelo/Versión', 'Color', 
            'Fecha Finanzas', 'Despacho Estimado', 'Entrega Estimada',
            'Fecha Recepción', 'Ubicación', 'Días Stock', 'Precio p/ Disponible',
            'Cód. Cliente', 'Cliente', 'Vendedor', 'Operación'
        ]
        
        # Asignar nombres a las primeras 15 columnas
        for i in range(min(15, len(df.columns))):
            df.rename(columns={df.columns[i]: column_names[i]}, inplace=True)
        
        # Calcular "Entrega Estimada" = "Despacho Estimado" + 1 mes
        if 'Despacho Estimado' in df.columns:
            # Convertir a datetime manejando errores
            df['Despacho Estimado'] = pd.to_datetime(df['Despacho Estimado'], errors='coerce')
            
            # Rellenar valores vacíos con 31/12/2030
            fecha_default = pd.Timestamp('2030-12-31')
            df['Despacho Estimado'] = df['Despacho Estimado'].fillna(fecha_default)
            
            # Calcular Entrega Estimada sumando 1 mes
            df['Entrega Estimada'] = df['Despacho Estimado'] + pd.DateOffset(months=1)
        
        # **Obtener precios de la lista de precios (Módulo 1) sin aplicar descuentos**
        if 'Modelo/Versión' in df.columns and 'Precio p/ Disponible' in df.columns:
            print("🔄 Obteniendo precios desde la lista de precios (Módulo 1)...")
            
            # Obtener todos los precios de la base de datos
            conn = get_db_connection()
            cursor = conn.cursor()
            cursor.execute('SELECT modelo, precio_ars FROM precios')
            precios_data = {}
            for row in cursor.fetchall():
                modelo = row['modelo']
                precio_ars = row['precio_ars'] or 0
                precios_data[modelo] = precio_ars
                print(f"🔧 Cargado: {modelo[:50]:50} | Precio: ${precio_ars:,.0f}")
            
            conn.rollback()
            release_db_connection(conn)
            
            print(f"📊 Precios cargados: {len(precios_data)} modelos")
            
            # Asignar precio directo de la lista sin descuentos
            def asignar_precio_directo(row):
                modelo = row['Modelo/Versión']
                numero_fabrica = row.get('Nº Fábrica', 'N/A')
                
                if pd.isna(modelo):
                    print(f"⚠️ Modelo vacío, saltando...")
                    return 0
                
                modelo_str = str(modelo).strip()
                
                # Buscar coincidencia exacta
                if modelo_str not in precios_data:
                    print(f"⚠️ Modelo '{modelo_str}' NO encontrado en precios_data")
                    return 0
                
                precio = precios_data[modelo_str]
                print(f"✅ {modelo_str[:50]:50} | Nº Fábrica: {numero_fabrica} | Precio: ${precio:,.0f}")
                
                return precio
            
            # Aplicar la función para asignar precios
            df['Precio p/ Disponible'] = df.apply(asignar_precio_directo, axis=1)
            
            # Contar cuántos precios se encontraron
            precios_encontrados = (df['Precio p/ Disponible'] > 0).sum()
            print(f"✅ Precios asignados: {precios_encontrados} de {len(df)} registros")
        
        # Mantener solo las 15 columnas originales
        columnas_a_mantener = [
            'Nº Fábrica', 'Nº Chasis', 'Modelo/Versión', 'Color', 
            'Fecha Finanzas', 'Despacho Estimado', 'Entrega Estimada',
            'Fecha Recepción', 'Ubicación', 'Días Stock', 'Precio p/ Disponible',
            'Cód. Cliente', 'Cliente', 'Vendedor', 'Operación'
        ]
        # Solo mantener las columnas que existen
        columnas_existentes = [col for col in columnas_a_mantener if col in df.columns]
        df = df[columnas_existentes]
        
        print(f"📊 Columnas finales del DataFrame: {list(df.columns)}")
        
        # **Filtrar unidades postergadas ANTES de dividir por pestañas**
        if unidades_postergadas and 'Nº Fábrica' in df.columns:
            filas_antes = len(df)
            # Convertir Nº Fábrica a string y eliminar espacios
            df['Nº Fábrica'] = df['Nº Fábrica'].astype(str).str.strip()
            # Filtrar las filas que NO estén en la lista de postergadas
            df = df[~df['Nº Fábrica'].isin(unidades_postergadas)]
            filas_despues = len(df)
            print(f"Filtrado de unidades postergadas: {filas_antes} filas -> {filas_despues} filas ({filas_antes - filas_despues} eliminadas)")
        
        # Convertir la primera columna a string para el análisis de prefijos
        df['Nº Fábrica'] = df['Nº Fábrica'].astype(str)
        
        # Separar por tipo de canal de venta según el prefijo de "Nº Fábrica"
        tabs_data = []
        
        # Ventas Especiales (F)
        df_ventas_especiales = df[df['Nº Fábrica'].str.startswith('F', na=False)].copy()
        if not df_ventas_especiales.empty:
            tabs_data.append({
                'id': 'ventas-especiales',
                'name': 'Ventas Especiales (F)',
                'count': len(df_ventas_especiales),
                'columns': list(df_ventas_especiales.columns),
                'rows': df_ventas_especiales.where(pd.notnull(df_ventas_especiales), None).values.tolist()
            })
        
        # Plan de Ahorro (TPA)
        df_plan_ahorro = df[df['Nº Fábrica'].str.startswith('TPA', na=False)].copy()
        if not df_plan_ahorro.empty:
            tabs_data.append({
                'id': 'plan-ahorro',
                'name': 'Plan de Ahorro (TPA)',
                'count': len(df_plan_ahorro),
                'columns': list(df_plan_ahorro.columns),
                'rows': df_plan_ahorro.where(pd.notnull(df_plan_ahorro), None).values.tolist()
            })
        
        # Ventas Convencionales (YAC)
        df_ventas_convencionales = df[df['Nº Fábrica'].str.startswith('YAC', na=False)].copy()
        if not df_ventas_convencionales.empty:
            tabs_data.append({
                'id': 'ventas-convencionales',
                'name': 'Ventas Convencionales (YAC)',
                'count': len(df_ventas_convencionales),
                'columns': list(df_ventas_convencionales.columns),
                'rows': df_ventas_convencionales.where(pd.notnull(df_ventas_convencionales), None).values.tolist()
            })
        
        # No incluir la pestaña "Otros" - comentado para simplificar
        # df_otros = df[~df['Nº Fábrica'].str.match(r'^(F|TPA|YAC)', na=False)].copy()
        
        # Eliminar archivo temporal
        os.remove(filepath)
        
        return jsonify({
            'success': True,
            'tabs': tabs_data,
            'total_rows': len(df)
        })
    
    except Exception as e:
        # Limpiar archivo si existe
        if 'filepath' in locals() and os.path.exists(filepath):
            os.remove(filepath)
        return jsonify({'error': f'Error al procesar el archivo: {str(e)}'}), 500


# Módulos 4, 5 y 6 (placeholders)
@app.route('/modulo4')
@login_required
@module_permission_required('planeamiento')
def modulo4():
    return render_template('modulo4.html')


@app.route('/modulo5')
@login_required
@module_permission_required('planeamiento')
def modulo5():
    return render_template('modulo5.html')


@app.route('/modulo6')
@login_required
@module_permission_required('planeamiento')
def modulo6():
    return render_template('modulo6.html')


@app.route('/modulo7')
@login_required
@module_permission_required('planeamiento')
def modulo7():
    return render_template('modulo7.html')


# ==================== API MÓDULO 4: DISPONIBLES ====================

# API para obtener unidades disponibles
@app.route('/api/disponibles', methods=['GET'])
@login_required
def get_disponibles():
    # Verificar que tenga permiso de planeamiento O ventas
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        return jsonify({'error': 'No tienes permisos para acceder a esta información'}), 403
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Obtener descuentos adicionales
    cursor.execute('SELECT tipo, clave, valor FROM descuentos_adicionales')
    desc_rows = cursor.fetchall()
    descuentos_config = {}
    for row in desc_rows:
        if row['tipo'] not in descuentos_config:
            descuentos_config[row['tipo']] = {}
        descuentos_config[row['tipo']][row['clave']] = row['valor']
    
    cursor.execute('''
        SELECT d.numero_fabrica, d.numero_chasis, d.modelo_version, d.color,
               d.fecha_finanzas, d.despacho_estimado, d.entrega_estimada,
               d.fecha_recepcion, d.ubicacion, d.dias_stock, d.precio_disponible,
               d.cod_cliente, d.cliente, d.vendedor, d.operacion,
               d.precio_base, d.descuento_individual as descuento_guardado, d.descuento_adicional,
               p.familia, p.descuento as descuento_individual, p.descuento_futuro
        FROM disponibles d
        LEFT JOIN precios p ON d.modelo_version = p.modelo
        ORDER BY d.fecha_carga DESC
    ''')
    rows = cursor.fetchall()
    conn.rollback()
    release_db_connection(conn)
    
    from datetime import datetime
    from dateutil.relativedelta import relativedelta
    
    disponibles = []
    for row in rows:
        # Usar el precio_base guardado o el precio_disponible como fallback
        precio_base_guardado = row.get('precio_base', 0) or row['precio_disponible'] or 0
        
        # Usar el descuento individual que ya fue guardado al procesar el Excel
        descuento_individual_guardado = row.get('descuento_guardado', 0) or 0
        
        descuento_total = descuento_individual_guardado
        detalles_descuento = []
        
        # Determinar fecha de entrega para calcular descuentos adicionales
        fecha_entrega_parsed = None
        
        if row['entrega_estimada']:
            try:
                fecha_str = str(row['entrega_estimada'])
                # Intentar parsear la fecha
                try:
                    fecha_entrega_parsed = datetime.strptime(fecha_str, '%Y-%m-%d')
                except:
                    try:
                        if 'GMT' in fecha_str:
                            fecha_str_limpia = fecha_str.replace(' GMT', '').strip()
                            fecha_entrega_parsed = datetime.strptime(fecha_str_limpia, '%a, %d %b %Y %H:%M:%S')
                        else:
                            fecha_entrega_parsed = datetime.fromisoformat(fecha_str.replace('GMT', '').strip())
                    except:
                        pass
            except Exception as e:
                pass
        
        # Descuento por Stock
        ubicacion_actual = (row['ubicacion'] or '').strip().upper()
        desc_stock = descuentos_config.get('stock', {}).get('descuento_stock', 0)
        
        # Comparar si contiene "STOCK" en cualquier parte
        if 'STOCK' in ubicacion_actual and desc_stock > 0:
            descuento_total += desc_stock
            detalles_descuento.append(f"Stock: {desc_stock}%")
        
        # Descuento por Color - Normalizar para comparar
        color_original = (row['color'] or '').strip()
        color_normalizado = color_original.lower().replace(' ', '_')
        desc_color = descuentos_config.get('color', {}).get(color_normalizado, 0)
        if desc_color > 0:
            descuento_total += desc_color
            detalles_descuento.append(f"Color: {desc_color}%")
        
        # Descuento por Antigüedad
        if fecha_entrega_parsed:
            try:
                meses_config = descuentos_config.get('antiguedad', {}).get('meses', 3)
                desc_antiguedad = descuentos_config.get('antiguedad', {}).get('descuento', 0)
                fecha_limite = datetime.now() - relativedelta(months=int(meses_config))
                
                if fecha_entrega_parsed < fecha_limite and desc_antiguedad > 0:
                    descuento_total += desc_antiguedad
                    detalles_descuento.append(f"Antigüedad: {desc_antiguedad}%")
            except Exception as e:
                print(f"❌ Error procesando antigüedad para {row['numero_fabrica']}: {e}")
        
        # Calcular precio final con TODOS los descuentos
        precio_final = precio_base_guardado * (1 - descuento_total / 100)
        
        # Separar descuento individual (va en su propia columna) del total de adicionales
        descuento_adicional = descuento_total - descuento_individual_guardado
        
        disponibles.append({
            'numero_fabrica': row['numero_fabrica'],
            'numero_chasis': row['numero_chasis'],
            'modelo_version': row['modelo_version'],
            'color': row['color'],
            'fecha_finanzas': row['fecha_finanzas'],
            'despacho_estimado': row['despacho_estimado'],
            'entrega_estimada': row['entrega_estimada'],
            'fecha_recepcion': row['fecha_recepcion'],
            'ubicacion': row['ubicacion'],
            'dias_stock': row['dias_stock'],
            'precio_disponible': round(precio_final, 2),
            'precio_base': precio_base_guardado,
            'descuento_individual': descuento_individual_guardado,
            'descuento_adicional': round(descuento_adicional, 2),
            'detalles_descuento': ', '.join(detalles_descuento) if detalles_descuento else 'Sin descuentos',
            'cod_cliente': row['cod_cliente'],
            'cliente': row['cliente'],
            'vendedor': row['vendedor'],
            'operacion': row['operacion'],
            'familia': row['familia'] or 'SIN FAMILIA'
        })
    
    # DEBUG: Mostrar qué se está enviando al frontend
    print(f"\n🚀 ============ ENVIANDO DATOS AL FRONTEND (MÓDULO 4) ============")
    print(f"   Total de unidades: {len(disponibles)}")
    if len(disponibles) > 0:
        print(f"\n   📤 Primera unidad que se envía:")
        print(f"      Nº Fábrica: {disponibles[0]['numero_fabrica']}")
        print(f"      Modelo: {disponibles[0]['modelo_version']}")
        print(f"      Precio Base: ${disponibles[0]['precio_base']}")
        print(f"      Descuento Individual: {disponibles[0]['descuento_individual']}%")
        print(f"      Descuento Adicional: {disponibles[0]['descuento_adicional']}%")
        print(f"      Precio Final: ${disponibles[0]['precio_disponible']}")
        
        if len(disponibles) > 1:
            print(f"\n   📤 Segunda unidad que se envía:")
            print(f"      Nº Fábrica: {disponibles[1]['numero_fabrica']}")
            print(f"      Modelo: {disponibles[1]['modelo_version']}")
            print(f"      Descuento Individual: {disponibles[1]['descuento_individual']}%")
        
        if len(disponibles) > 2:
            print(f"\n   📤 Tercera unidad que se envía:")
            print(f"      Nº Fábrica: {disponibles[2]['numero_fabrica']}")
            print(f"      Modelo: {disponibles[2]['modelo_version']}")
            print(f"      Descuento Individual: {disponibles[2]['descuento_individual']}%")
    print(f"   ==================================================================\n")
    
    return jsonify(disponibles)

# API para guardar/reemplazar unidades disponibles
@app.route('/api/disponibles', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_disponibles():
    data = request.json
    
    print(f"\n🔍 ============ RECIBIENDO DATOS PARA GUARDAR EN DISPONIBLES ============")
    print(f"   Total de unidades: {len(data)}")
    if len(data) > 0:
        print(f"\n   📦 Primera unidad recibida:")
        print(f"      Nº Fábrica: {data[0].get('numero_fabrica')}")
        print(f"      Modelo: {data[0].get('modelo_version')}")
        print(f"      Precio Disponible: {data[0].get('precio_disponible')}")
        print(f"\n   🔑 TODAS las claves disponibles: {list(data[0].keys())}")
        
        # Mostrar más unidades para verificar el patrón
        if len(data) > 1:
            print(f"\n   📦 Segunda unidad recibida:")
            print(f"      Nº Fábrica: {data[1].get('numero_fabrica')}")
            print(f"      Modelo: {data[1].get('modelo_version')}")
        
        if len(data) > 2:
            print(f"\n   📦 Tercera unidad recibida:")
            print(f"      Nº Fábrica: {data[2].get('numero_fabrica')}")
            print(f"      Modelo: {data[2].get('modelo_version')}")
    print(f"   ========================================================================\n")
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        # Eliminar todos los registros anteriores
        cursor.execute('DELETE FROM disponibles')
        
        # Insertar nuevos registros
        for item in data:
            cursor.execute('''
                INSERT INTO disponibles (
                    numero_fabrica, numero_chasis, modelo_version, color,
                    fecha_finanzas, despacho_estimado, entrega_estimada,
                    fecha_recepcion, ubicacion, dias_stock, precio_disponible,
                    cod_cliente, cliente, vendedor, operacion
                ) VALUES (%s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s, %s)
            ''', (
                item.get('numero_fabrica'),
                item.get('numero_chasis'),
                item.get('modelo_version'),
                item.get('color'),
                item.get('fecha_finanzas'),
                item.get('despacho_estimado'),
                item.get('entrega_estimada'),
                item.get('fecha_recepcion'),
                item.get('ubicacion'),
                item.get('dias_stock'),
                item.get('precio_disponible'),  # Precio directo de la lista de precios sin descuentos
                item.get('cod_cliente'),
                item.get('cliente'),
                item.get('vendedor'),
                item.get('operacion')
            ))
        
        conn.commit()
        conn.rollback()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'count': len(data)})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para obtener unidades reservadas
@app.route('/api/unidades_reservadas', methods=['GET'])
@login_required
def get_unidades_reservadas():
    # Verificar que tenga permiso de planeamiento O ventas
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        return jsonify({'error': 'No tienes permisos para acceder a esta información'}), 403
    
    conn = get_db_connection()
    cursor = conn.cursor()
    cursor.execute('SELECT numero_fabrica, vendedor, fecha_agregado FROM unidades_reservadas ORDER BY fecha_agregado DESC')
    rows = cursor.fetchall()
    release_db_connection(conn)
    
    reservadas = []
    for row in rows:
        reservadas.append({
            'numero_fabrica': row['numero_fabrica'],
            'vendedor': row['vendedor'],
            'fecha_agregado': row['fecha_agregado']
        })
    return jsonify(reservadas)

# API para agregar unidad reservada
@app.route('/api/unidades_reservadas', methods=['POST'])
@login_required
def add_unidad_reservada():
    # Verificar que tenga permiso de planeamiento O ventas
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        return jsonify({'error': 'No tienes permisos para acceder a esta información'}), 403
    
    data = request.json
    numero_fabrica = data.get('numero_fabrica', '').strip()
    vendedor = data.get('vendedor', '').strip()
    
    if not numero_fabrica:
        return jsonify({'success': False, 'error': 'Número de fábrica vacío'}), 400
    
    if not vendedor:
        return jsonify({'success': False, 'error': 'Nombre del vendedor vacío'}), 400
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('INSERT INTO unidades_reservadas (numero_fabrica, vendedor) VALUES (%s, %s)', (numero_fabrica, vendedor))
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except PgIntegrityError:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': 'Número de fábrica ya existe'}), 400
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

# API para eliminar unidad reservada
@app.route('/api/unidades_reservadas/<numero_fabrica>', methods=['DELETE'])
@login_required
def delete_unidad_reservada(numero_fabrica):
    # Verificar que tenga permiso de planeamiento O ventas
    if not (current_user.has_permission('planeamiento') or current_user.has_permission('ventas')):
        return jsonify({'error': 'No tienes permisos para acceder a esta información'}), 403
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('DELETE FROM unidades_reservadas WHERE numero_fabrica = %s', (numero_fabrica,))
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API MÓDULO 6: RECAUDACIÓN ====================

@app.route('/api/recaudacion', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_recaudacion():
    """Obtener datos de recaudación desde disponibles (solo YAC - Ventas Convencionales)
    Excluye unidades con ubicación 'Preventa'"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Obtener unidades disponibles que son YAC (Ventas Convencionales)
    # Estas son las unidades que vienen del módulo 3 (Seguimiento) con cliente vacío
    # Excluimos las que tienen ubicación "Preventa"
    cursor.execute('''
        SELECT 
            numero_fabrica,
            modelo_version,
            ubicacion,
            precio_disponible
        FROM disponibles
        WHERE numero_fabrica LIKE 'YAC%'
        AND (ubicacion IS NULL OR UPPER(TRIM(ubicacion)) != 'PREVENTA')
    ''')
    
    unidades = cursor.fetchall()
    release_db_connection(conn)
    
    # Separar por Stock y No Stock según la ubicación
    en_stock = []
    no_stock = []
    
    for unidad in unidades:
        ubicacion = (unidad['ubicacion'] or '').upper().strip()
        precio = unidad['precio_disponible'] or 0
        
        data_item = {
            'numero_fabrica': unidad['numero_fabrica'],
            'modelo_version': unidad['modelo_version'],
            'ubicacion': unidad['ubicacion'],
            'precio': precio
        }
        
        # Verificar si está en stock (contiene la palabra "STOCK")
        if 'STOCK' in ubicacion:
            en_stock.append(data_item)
        else:
            no_stock.append(data_item)
    
    # Calcular totales
    total_en_stock = sum(item['precio'] for item in en_stock)
    total_no_stock = sum(item['precio'] for item in no_stock)
    total_general = total_en_stock + total_no_stock
    
    return jsonify({
        'success': True,
        'en_stock': {
            'unidades': en_stock,
            'cantidad': len(en_stock),
            'total': total_en_stock
        },
        'no_stock': {
            'unidades': no_stock,
            'cantidad': len(no_stock),
            'total': total_no_stock
        },
        'total_general': total_general,
        'cantidad_total': len(en_stock) + len(no_stock)
    })


# ==================== API BACKUP/RESTORE ====================

@app.route('/api/backup/download', methods=['GET'])
@login_required
def download_backup():
    """Descargar backup completo de la base de datos"""
    import io
    import tempfile
    import shutil
    
    try:
        # Crear archivo temporal
        temp_db = tempfile.NamedTemporaryFile(delete=False, suffix='.db')
        shutil.copy2('database.db', temp_db.name)
        temp_db.close()
        
        return send_file(
            temp_db.name,
            as_attachment=True,
            download_name=f'database_backup_{datetime.now().strftime("%Y%m%d_%H%M%S")}.db',
            mimetype='application/x-sqlite3'
        )
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/backup/upload', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def upload_backup():
    """Restaurar base de datos desde backup"""
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se encontró archivo'}), 400
        
        file = request.files['file']
        
        if file.filename == '':
            return jsonify({'error': 'No se seleccionó archivo'}), 400
        
        # Guardar backup actual por seguridad
        import shutil
        shutil.copy2('database.db', 'database_backup_before_restore.db')
        
        # Reemplazar con el nuevo archivo
        file.save('database.db')
        
        return jsonify({'success': True, 'message': 'Base de datos restaurada correctamente'})
    except Exception as e:
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API OBSERVACIONES (MÓDULO 5) ====================

@app.route('/api/observaciones/config_dias', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_config_dias():
    """Obtener configuración de días estándar y desvío por zona"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT zona, dias_estandar, dias_desvio FROM config_dias_zonas ORDER BY zona')
        rows = cursor.fetchall()
        release_db_connection(conn)
        
        # Convertir a lista para el frontend
        config = [dict(row) for row in rows]
        
        return jsonify(config)
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/observaciones/config_dias', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_config_dias():
    """Guardar configuración de días"""
    conn = None
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()
        
        zonas = data.get('zonas', [])
        
        if not zonas:
            return jsonify({'success': False, 'error': 'No se recibieron datos de zonas'}), 400
        
        for zona_data in zonas:
            cursor.execute('''
                INSERT INTO config_dias_zonas (zona, dias_estandar, dias_desvio, fecha_actualizacion)
                VALUES (%s, %s, %s, CURRENT_TIMESTAMP)
                ON CONFLICT (zona) 
                DO UPDATE SET 
                    dias_estandar = EXCLUDED.dias_estandar,
                    dias_desvio = EXCLUDED.dias_desvio,
                    fecha_actualizacion = CURRENT_TIMESTAMP
            ''', (zona_data['zona'], zona_data['dias_estandar'], zona_data['dias_desvio']))
        
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True, 'message': f'{len(zonas)} zonas guardadas'})
    except Exception as e:
        if conn:
            try:
                conn.rollback()
            except:
                pass
            release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/observaciones/matriz_codigos', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_matriz_codigos():
    """Obtener matriz de códigos de observación"""
    conn = None
    try:
        conn = get_db_connection()
        cursor = conn.cursor()
        
        cursor.execute('SELECT clase, zona, codigos, es_zona_arribo FROM matriz_codigos_obs ORDER BY clase, zona')
        rows = cursor.fetchall()
        release_db_connection(conn)
        
        # Convertir a lista para el frontend
        matriz = [dict(row) for row in rows]
        
        return jsonify(matriz)
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/observaciones/matriz_codigos', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def save_matriz_codigos():
    """Guardar matriz de códigos"""
    conn = None
    try:
        data = request.json
        conn = get_db_connection()
        cursor = conn.cursor()
        
        matriz = data.get('matriz', [])
        
        if not matriz:
            return jsonify({'success': False, 'error': 'No se recibieron datos de matriz'}), 400
        
        print(f"\n>> Guardando matriz de codigos: {len(matriz)} registros")
        
        # IMPORTANTE: Eliminar todos los registros existentes primero
        # para evitar registros huerfanos de configuraciones anteriores
        cursor.execute('DELETE FROM matriz_codigos_obs')
        print(f"   Registros anteriores eliminados")
        
        # Insertar nuevos registros
        for idx, registro in enumerate(matriz):
            try:
                # Extraer y validar valores
                clase = registro.get('clase')
                zona_raw = registro.get('zona')
                codigos = registro.get('codigos')
                es_zona_arribo = registro.get('es_zona_arribo', False)
                
                # Validaciones estrictas
                if not clase or not isinstance(clase, str):
                    raise ValueError(f"Clase inválida: {clase}")
                
                if not isinstance(zona_raw, int):
                    raise ValueError(f"Zona debe ser un número, recibido: {type(zona_raw).__name__} = {zona_raw}")
                
                zona = zona_raw
                if zona < 1 or zona > 4:
                    raise ValueError(f"Zona fuera de rango (1-4): {zona}")
                
                if codigos is None:
                    codigos = ''
                codigos = str(codigos).strip()
                
                # Convertir es_zona_arribo a booleano
                if isinstance(es_zona_arribo, bool):
                    pass  # Ya es booleano
                elif isinstance(es_zona_arribo, str):
                    es_zona_arribo = es_zona_arribo.lower() in ['true', '1', 'yes']
                elif isinstance(es_zona_arribo, (int, float)):
                    es_zona_arribo = bool(es_zona_arribo)
                else:
                    es_zona_arribo = False
                
                print(f"  [{idx}] {clase} Zona {zona}: '{codigos}' (arribo={es_zona_arribo})")
                
                cursor.execute('''
                    INSERT INTO matriz_codigos_obs (clase, zona, codigos, es_zona_arribo, fecha_actualizacion)
                    VALUES (%s, %s, %s, %s, CURRENT_TIMESTAMP)
                ''', (clase, zona, codigos, es_zona_arribo))
                
            except ValueError as val_err:
                error_msg = f"Error de validacion en registro {idx}: {val_err}"
                print(f"  ERROR: {error_msg}")
                print(f"     Registro completo: {registro}")
                conn.rollback()
                release_db_connection(conn)
                return jsonify({'success': False, 'error': error_msg}), 400
            except Exception as row_error:
                error_msg = f"Error en registro {idx}: {str(row_error)}"
                print(f"  ERROR: {error_msg}")
                print(f"     Registro completo: {registro}")
                conn.rollback()
                release_db_connection(conn)
                return jsonify({'success': False, 'error': error_msg}), 500
        
        conn.commit()
        release_db_connection(conn)
        print(f">> OK - Matriz guardada exitosamente: {len(matriz)} registros")
        return jsonify({'success': True, 'message': f'{len(matriz)} registros guardados'})
        
    except Exception as e:
        error_msg = f"Error general guardando matriz: {str(e)}"
        print(f"ERROR: {error_msg}")
        if conn:
            try:
                conn.rollback()
            except:
                pass
            release_db_connection(conn)
        return jsonify({'success': False, 'error': error_msg}), 500

@app.route('/api/observaciones/registrar_cambio', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def registrar_cambio_observacion():
    """Registrar cambio de código de observación"""
    data = request.json
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        operacion = data.get('operacion')
        codigo_nuevo = data.get('codigo_nuevo')
        zona_nueva = data.get('zona_nueva')
        ejecutivo = data.get('ejecutivo')
        
        # Obtener último cambio
        cursor.execute('''
            SELECT codigo_nuevo as codigo, zona_nueva as zona
            FROM auditoria_observaciones
            WHERE operacion = %s
            ORDER BY fecha_cambio DESC
            LIMIT 1
        ''', (operacion,))
        
        ultimo = cursor.fetchone()
        codigo_anterior = ultimo['codigo'] if ultimo else None
        zona_anterior = ultimo['zona'] if ultimo else None
        
        # Detectar retroceso
        es_retroceso = False
        if zona_anterior and zona_nueva and zona_nueva < zona_anterior:
            es_retroceso = True
        
        # Insertar auditoría
        cursor.execute('''
            INSERT INTO auditoria_observaciones 
            (operacion, codigo_anterior, codigo_nuevo, zona_anterior, zona_nueva, ejecutivo, es_retroceso)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
        ''', (operacion, codigo_anterior, codigo_nuevo, zona_anterior, zona_nueva, ejecutivo, es_retroceso))
        
        # Actualizar estadísticas
        cursor.execute('''
            INSERT INTO stats_operaciones (operacion, cantidad_cambios, cantidad_retrocesos, marcado_sospechoso)
            VALUES (%s, 1, %s, %s)
            ON CONFLICT (operacion)
            DO UPDATE SET
                cantidad_cambios = stats_operaciones.cantidad_cambios + 1,
                cantidad_retrocesos = stats_operaciones.cantidad_retrocesos + CASE WHEN %s THEN 1 ELSE 0 END,
                marcado_sospechoso = CASE WHEN stats_operaciones.cantidad_retrocesos + CASE WHEN %s THEN 1 ELSE 0 END > 1 THEN TRUE ELSE FALSE END,
                fecha_actualizacion = CURRENT_TIMESTAMP
        ''', (operacion, 1 if es_retroceso else 0, es_retroceso, es_retroceso, es_retroceso))
        
        conn.commit()
        release_db_connection(conn)
        return jsonify({'success': True, 'es_retroceso': es_retroceso})
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500

@app.route('/api/observaciones/stats/<operacion>', methods=['GET'])
@login_required
@module_permission_required('planeamiento')
def get_stats_operacion(operacion):
    """Obtener estadísticas de una operación"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    try:
        cursor.execute('''
            SELECT cantidad_cambios, cantidad_retrocesos, marcado_sospechoso
            FROM stats_operaciones
            WHERE operacion = %s
        ''', (operacion,))
        
        row = cursor.fetchone()
        release_db_connection(conn)
        
        if row:
            return jsonify({
                'success': True,
                'stats': {
                    'cantidad_cambios': row['cantidad_cambios'],
                    'cantidad_retrocesos': row['cantidad_retrocesos'],
                    'marcado_sospechoso': row['marcado_sospechoso']
                }
            })
        else:
            return jsonify({
                'success': True,
                'stats': {
                    'cantidad_cambios': 0,
                    'cantidad_retrocesos': 0,
                    'marcado_sospechoso': False
                }
            })
    except Exception as e:
        conn.rollback()
        release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API BUSINESS INTELLIGENCE ====================

@app.route('/api/bi/upload_databases', methods=['POST'])
@login_required
def upload_bi_databases():
    """Cargar y procesar archivos de bases de datos para BI"""
    try:
        files = request.files
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Solo verificar archivos de patentamientos (ignorar entregas)
        required_files = [
            'argentina-marca', 'argentina-modelo', 
            'mendoza-marca', 'mendoza-modelo'
        ]
        
        for file_key in required_files:
            if file_key not in files:
                return jsonify({'success': False, 'error': f'Falta el archivo: {file_key}'}), 400
            # Verificar que el archivo tenga nombre válido
            file = files[file_key]
            if not file.filename or file.filename == '':
                return jsonify({'success': False, 'error': f'Archivo inválido: {file_key}'}), 400
        
        # Procesar archivos de patentamientos
        import pandas as pd
        from dateutil import parser as date_parser
        
        # Diccionario para convertir nombres de meses en español
        meses_es = {
            'enero': 1, 'febrero': 2, 'marzo': 3, 'abril': 4,
            'mayo': 5, 'junio': 6, 'julio': 7, 'agosto': 8,
            'septiembre': 9, 'octubre': 10, 'noviembre': 11, 'diciembre': 12,
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4,
            'may': 5, 'jun': 6, 'jul': 7, 'ago': 8,
            'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
        }
        
        # Limpiar tablas anteriores
        cursor.execute('DELETE FROM bi_patentamientos_argentina_marca')
        cursor.execute('DELETE FROM bi_patentamientos_argentina_modelo')
        cursor.execute('DELETE FROM bi_patentamientos_mendoza_marca')
        cursor.execute('DELETE FROM bi_patentamientos_mendoza_modelo')
        
        print("🔄 Comenzando procesamiento de archivos...")
        
        # Procesar cada archivo
        for file_key in ['argentina-marca', 'argentina-modelo', 'mendoza-marca', 'mendoza-modelo']:
            file = files[file_key]
            print(f"\n📂 Procesando archivo: {file_key} ({file.filename})")
            
            # Leer Excel sin forzar tipos
            df = pd.read_excel(file, dtype=str)
            print(f"   ✓ Archivo leído: {len(df)} filas, {len(df.columns)} columnas")
            print(f"   Columnas detectadas: {list(df.columns)}")
            
            # La primera columna es el nombre (marca o modelo)
            nombre_columna = df.columns[0]
            fechas_columnas = df.columns[1:]
            
            print(f"   Columna de nombres: '{nombre_columna}'")
            print(f"   Primeras 3 fechas: {list(fechas_columnas[:3])}")
            
            # Determinar tabla destino
            if file_key == 'argentina-marca':
                tabla = 'bi_patentamientos_argentina_marca'
            elif file_key == 'argentina-modelo':
                tabla = 'bi_patentamientos_argentina_modelo'
            elif file_key == 'mendoza-marca':
                tabla = 'bi_patentamientos_mendoza_marca'
            else:  # mendoza-modelo
                tabla = 'bi_patentamientos_mendoza_modelo'
            
            registros_insertados = 0
            errores = 0
            
            # Insertar datos
            for idx, row in df.iterrows():
                nombre = str(row[nombre_columna]).strip()
                
                # Saltar filas vacías
                if pd.isna(nombre) or nombre == '' or nombre == 'nan':
                    continue
                
                # Insertar cada mes
                for fecha_col in fechas_columnas:
                    valor = row[fecha_col]
                    
                    # Convertir valor a número
                    try:
                        if pd.isna(valor) or valor == '' or valor == 'nan':
                            cantidad = 0
                        else:
                            cantidad = int(float(str(valor).replace(',', '.')))
                    except:
                        cantidad = 0
                    
                    # Parsear fecha (puede venir como "enero-15", "01/15", o timestamp de pandas)
                    try:
                        fecha_str = str(fecha_col).strip().lower()
                        
                        # Caso 1: Formato "enero-15" o "ene-15"
                        if '-' in fecha_str:
                            partes = fecha_str.split('-')
                            mes_nombre = partes[0].strip()
                            anio = partes[1].strip()
                            
                            # Buscar el mes en el diccionario
                            mes = meses_es.get(mes_nombre, None)
                            if mes is None:
                                print(f"   ⚠️ Mes no reconocido: '{mes_nombre}'")
                                continue
                            
                            # Convertir año a 4 dígitos
                            if len(anio) == 2:
                                anio = '20' + anio
                            
                            fecha = f"{anio}-{str(mes).zfill(2)}-01"
                        
                        # Caso 2: Formato "01/15" o "1/15"
                        elif '/' in fecha_str:
                            partes = fecha_str.split('/')
                            mes = int(partes[0])
                            anio = partes[1].strip()
                            
                            if len(anio) == 2:
                                anio = '20' + anio
                            
                            fecha = f"{anio}-{str(mes).zfill(2)}-01"
                        
                        # Caso 3: Timestamp de pandas
                        elif isinstance(fecha_col, pd.Timestamp):
                            fecha = fecha_col.strftime('%Y-%m-%d')
                        
                        # Caso 4: Intentar parsear con dateutil
                        else:
                            try:
                                fecha_obj = date_parser.parse(fecha_str, dayfirst=True)
                                fecha = fecha_obj.strftime('%Y-%m-01')
                            except:
                                print(f"   ⚠️ Fecha no reconocida: '{fecha_str}'")
                                continue
                        
                        # Insertar en base de datos
                        cursor.execute(f'''
                            INSERT INTO {tabla} (nombre, fecha, cantidad)
                            VALUES (%s, %s, %s)
                        ''', (nombre, fecha, cantidad))
                        
                        registros_insertados += 1
                        
                    except Exception as e:
                        errores += 1
                        if errores < 5:  # Mostrar solo los primeros 5 errores
                            print(f"   ⚠️ Error procesando fecha '{fecha_col}': {e}")
            
            print(f"   ✅ Insertados: {registros_insertados} registros | Errores: {errores}")
        
        conn.commit()
        
        # Obtener contadores de filas
        row_counts = {}
        cursor.execute('SELECT COUNT(*) as total FROM bi_patentamientos_argentina_marca')
        row_counts['argentina-marca'] = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as total FROM bi_patentamientos_argentina_modelo')
        row_counts['argentina-modelo'] = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as total FROM bi_patentamientos_mendoza_marca')
        row_counts['mendoza-marca'] = cursor.fetchone()['total']
        
        cursor.execute('SELECT COUNT(*) as total FROM bi_patentamientos_mendoza_modelo')
        row_counts['mendoza-modelo'] = cursor.fetchone()['total']
        
        release_db_connection(conn)
        
        # INVALIDAR CACHÉ cuando se cargan nuevos datos
        patentamientos_cache['data'] = None
        patentamientos_cache['timestamp'] = 0
        print("🗑️ Caché de patentamientos invalidado")
        
        return jsonify({
            'success': True, 
            'message': 'Archivos procesados correctamente',
            'rowCounts': row_counts
        })
    except Exception as e:
        if 'conn' in locals():
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bi/patentamientos', methods=['GET'])
@login_required
def get_bi_patentamientos():
    """Obtener datos de patentamientos DIRECTAMENTE desde CSV - CON CACHÉ"""
    try:
        # Verificar si el caché es válido
        current_time = time()
        force_refresh = request.args.get('refresh', 'false').lower() == 'true'
        
        if (not force_refresh and 
            patentamientos_cache['data'] is not None and 
            (current_time - patentamientos_cache['timestamp']) < patentamientos_cache['ttl']):
            print("✅ Sirviendo datos desde CACHÉ")
            return jsonify({'success': True, 'data': patentamientos_cache['data'], 'from_cache': True})
        
        print("🔄 Cargando datos desde CSV...")
        
        # Leer datos desde CSV
        result = {
            'argentina_marca': get_patentamientos_from_csv('Mercado Argentino MARCA.csv', 'marca'),
            'argentina_modelo': get_patentamientos_from_csv('Mercado Argentino MODELO.csv', 'modelo'),
            'mendoza_marca': get_patentamientos_from_csv('Mercado Mendoza MARCA.csv', 'marca'),
            'mendoza_modelo': get_patentamientos_from_csv('Mercado Mendoza MODELO.csv', 'modelo')
        }
        
        # Guardar en caché
        patentamientos_cache['data'] = result
        patentamientos_cache['timestamp'] = current_time
        print(f"💾 Datos guardados en caché (válido por {patentamientos_cache['ttl']}s)")
        
        return jsonify({'success': True, 'data': result, 'from_cache': False})
    except FileNotFoundError as e:
        return jsonify({'success': False, 'error': f'Archivos CSV no encontrados: {str(e)}'}), 404
    except Exception as e:
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


def get_patentamientos_from_csv(filename, tipo):
    """Leer y procesar datos de patentamientos desde CSV"""
    csv_path = os.path.join(os.path.dirname(__file__), 'Patentamientos', filename)
    
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"Archivo no encontrado: {csv_path}")
    
    # Leer CSV con diferentes encodings
    df = None
    for encoding in ['latin-1', 'cp1252', 'iso-8859-1', 'utf-8']:
        try:
            df = pd.read_csv(csv_path, delimiter=';', encoding=encoding)
            break
        except:
            continue
    
    if df is None:
        raise Exception(f"No se pudo leer el archivo {filename}")
    
    # Primera columna es el nombre (marca o modelo)
    nombres_col = df.columns[0]
    
    # Resto de columnas son fechas (ene-15, feb-15, etc.)
    fecha_cols = df.columns[1:]
    
    # Convertir nombres de columnas de fecha a formato MM/YY
    meses_map = {
        'ene': '01', 'feb': '02', 'mar': '03', 'abr': '04',
        'may': '05', 'jun': '06', 'jul': '07', 'ago': '08',
        'sep': '09', 'oct': '10', 'nov': '11', 'dic': '12'
    }
    
    fechas_formateadas = []
    for col in fecha_cols:
        col_lower = str(col).lower().strip()
        if '-' in col_lower:
            mes_str, anio_str = col_lower.split('-')
            mes = meses_map.get(mes_str.strip())
            anio = anio_str.strip()
            if mes and len(anio) == 2:
                fechas_formateadas.append(f"{mes}/{anio}")
            else:
                fechas_formateadas.append(col)
        else:
            fechas_formateadas.append(col)
    
    if tipo == 'marca':
        return get_datos_por_marca(df, nombres_col, fecha_cols, fechas_formateadas)
    else:
        return get_datos_por_modelo(df, nombres_col, fecha_cols, fechas_formateadas)


def get_datos_por_marca(df, nombres_col, fecha_cols, fechas_formateadas):
    """Procesar datos mostrando TODAS las marcas individualmente"""
    result = {'labels': fechas_formateadas}
    
    # Primero, construir un diccionario temporal con todas las marcas
    marcas_dict = {}
    
    # Procesar cada fila
    for idx, row in df.iterrows():
        nombre = str(row[nombres_col]).strip()
        marca_normalizada = nombre.lower()
        
        if marca_normalizada not in marcas_dict:
            marcas_dict[marca_normalizada] = [0] * len(fecha_cols)
        
        for i, fecha_col in enumerate(fecha_cols):
            valor = row[fecha_col]
            try:
                cantidad = int(float(valor)) if pd.notna(valor) else 0
            except:
                cantidad = 0
            
            marcas_dict[marca_normalizada][i] += cantidad
    
    # Agregar todas las marcas al resultado
    result.update(marcas_dict)
    
    return result


def get_datos_por_modelo(df, nombres_col, fecha_cols, fechas_formateadas):
    """Procesar datos por modelo con Top 5"""
    # Obtener datos de la última columna para Top 5
    ultima_col = fecha_cols[-1]
    
    # Crear lista de (nombre, valor_ultimo_mes) y ordenar
    modelos_ultimo_mes = []
    for idx, row in df.iterrows():
        nombre = str(row[nombres_col]).strip()
        valor = row[ultima_col]
        try:
            cantidad = int(float(valor)) if pd.notna(valor) else 0
        except:
            cantidad = 0
        
        if cantidad > 0:
            modelos_ultimo_mes.append({'nombre': nombre, 'cantidad': cantidad})
    
    # Ordenar y obtener Top 5
    modelos_ultimo_mes.sort(key=lambda x: x['cantidad'], reverse=True)
    top5_ultimo_mes = modelos_ultimo_mes[:5]
    
    # Procesar todos los modelos con sus valores históricos
    modelos = []
    for idx, row in df.iterrows():
        nombre = str(row[nombres_col]).strip()
        valores = []
        
        for fecha_col in fecha_cols:
            valor = row[fecha_col]
            try:
                cantidad = int(float(valor)) if pd.notna(valor) else 0
            except:
                cantidad = 0
            valores.append(cantidad)
        
        modelos.append({
            'nombre': nombre,
            'valores': valores
        })
    
    return {
        'labels': fechas_formateadas,
        'modelos': modelos,
        'top5_ultimo_mes': top5_ultimo_mes
    }


def get_patentamientos_marca(cursor, tabla):
    """Obtener datos agrupados por marca - OPTIMIZADO con una consulta"""
    # Obtener todas las fechas distintas ordenadas
    cursor.execute(f'SELECT DISTINCT fecha FROM {tabla} ORDER BY fecha')
    fechas = [row['fecha'].strftime('%m/%y') for row in cursor.fetchall()]
    
    marcas = ['TOYOTA', 'FORD', 'FIAT', 'VOLKSWAGEN', 'CHEVROLET', 'PEUGEOT']
    result = {'labels': fechas}
    
    # Inicializar diccionarios para cada marca
    for marca in marcas:
        result[marca.lower()] = [0] * len(fechas)
    result['otros'] = [0] * len(fechas)
    
    # Obtener todos los datos en una sola consulta
    cursor.execute(f'SELECT fecha, nombre, cantidad FROM {tabla} ORDER BY fecha')
    
    fecha_to_index = {fecha: idx for idx, fecha in enumerate(fechas)}
    
    for row in cursor.fetchall():
        fecha_str = row['fecha'].strftime('%m/%y')
        nombre_upper = row['nombre'].upper()
        cantidad = row['cantidad']
        idx = fecha_to_index.get(fecha_str)
        
        if idx is None:
            continue
        
        # Clasificar por marca
        marca_encontrada = False
        for marca in marcas:
            if marca in nombre_upper:
                result[marca.lower()][idx] += cantidad
                marca_encontrada = True
                break
        
        if not marca_encontrada:
            result['otros'][idx] += cantidad
    
    return result


def get_patentamientos_modelo(cursor, tabla):
    """Obtener datos por modelo - OPTIMIZADO con una sola consulta"""
    # Obtener fechas
    cursor.execute(f'SELECT DISTINCT fecha FROM {tabla} ORDER BY fecha')
    fechas_rows = cursor.fetchall()
    fechas = [row['fecha'].strftime('%m/%y') for row in fechas_rows]
    fechas_dict = {row['fecha'].strftime('%m/%y'): row['fecha'] for row in fechas_rows}
    
    # Obtener la última fecha disponible
    cursor.execute(f'SELECT MAX(fecha) as ultima_fecha FROM {tabla}')
    ultima_fecha = cursor.fetchone()['ultima_fecha']
    
    # Obtener Top 5 del último mes
    cursor.execute(f'''
        SELECT nombre, cantidad
        FROM {tabla}
        WHERE fecha = %s
        ORDER BY cantidad DESC
        LIMIT 5
    ''', (ultima_fecha,))
    
    top5_ultimo_mes = [{'nombre': row['nombre'], 'cantidad': row['cantidad']} for row in cursor.fetchall()]
    
    # OPTIMIZACIÓN: Obtener TODOS los datos en una sola consulta
    cursor.execute(f'''
        SELECT nombre, fecha, cantidad
        FROM {tabla}
        ORDER BY nombre, fecha
    ''')
    
    # Construir estructura de datos eficientemente
    modelos_dict = {}
    for row in cursor.fetchall():
        nombre = row['nombre']
        fecha_str = row['fecha'].strftime('%m/%y')
        cantidad = row['cantidad']
        
        if nombre not in modelos_dict:
            modelos_dict[nombre] = {fecha_str: cantidad}
        else:
            modelos_dict[nombre][fecha_str] = cantidad
    
    # Convertir a formato esperado por el frontend
    modelos = []
    for nombre, valores_dict in modelos_dict.items():
        valores = [valores_dict.get(fecha, 0) for fecha in fechas]
        modelos.append({
            'nombre': nombre,
            'valores': valores
        })
    
    return {
        'labels': fechas,
        'modelos': modelos,
        'top5_ultimo_mes': top5_ultimo_mes
    }


@app.route('/api/bi/save_manual_data', methods=['POST'])
@login_required
def save_manual_data():
    """Guardar solo MES ACTUAL - Actualización rápida"""
    conn = None
    try:
        data = request.get_json()
        data_type = data.get('dataType')
        headers = data.get('headers')
        rows = data.get('rows')
        
        if not all([data_type, headers, rows]):
            return jsonify({'success': False, 'error': 'Datos incompletos'}), 400
        
        # Mapeo de tablas
        tabla_map = {
            'argentina-marca': 'bi_patentamientos_argentina_marca',
            'argentina-modelo': 'bi_patentamientos_argentina_modelo',
            'mendoza-marca': 'bi_patentamientos_mendoza_marca',
            'mendoza-modelo': 'bi_patentamientos_mendoza_modelo'
        }
        
        if data_type not in tabla_map:
            return jsonify({'success': False, 'error': 'Tipo de datos inválido'}), 400
        
        tabla = tabla_map[data_type]
        
        # Diccionario de meses
        meses_es = {
            'ene': 1, 'feb': 2, 'mar': 3, 'abr': 4,
            'may': 5, 'jun': 6, 'jul': 7, 'ago': 8,
            'sep': 9, 'oct': 10, 'nov': 11, 'dic': 12
        }
        
        print(f"\n📝 Actualizando MES ACTUAL para {data_type}")
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # IDENTIFICAR EL ÚLTIMO MES (columna más reciente)
        ultima_columna_idx = len(headers) - 1
        fecha_col = str(headers[ultima_columna_idx]).strip().lower()
        
        print(f"   📅 Última columna detectada: {fecha_col}")
        
        # Parsear la fecha del último mes
        if '-' in fecha_col:
            partes = fecha_col.split('-')
            mes_nombre = partes[0].strip()
            anio = partes[1].strip()
            
            mes = meses_es.get(mes_nombre)
            if not mes:
                return jsonify({'success': False, 'error': f'Mes no reconocido: {mes_nombre}'}), 400
            
            if len(anio) == 2:
                anio = '20' + anio if int(anio) <= 50 else '19' + anio
            
            fecha_actualizar = f"{anio}-{str(mes).zfill(2)}-01"
        else:
            return jsonify({'success': False, 'error': 'Formato de fecha no reconocido'}), 400
        
        print(f"   🎯 Actualizando datos para: {fecha_actualizar}")
        
        # Preparar registros SOLO del último mes
        registros = []
        
        for row_data in rows:
            nombre = str(row_data[0]).strip()
            if not nombre or nombre.lower() in ['', 'nan', 'none']:
                continue
            
            # Obtener valor del último mes
            if ultima_columna_idx < len(row_data):
                valor_str = row_data[ultima_columna_idx]
                
                try:
                    if not valor_str or str(valor_str).lower() in ['', 'nan', 'none']:
                        cantidad = 0
                    else:
                        cantidad = int(float(str(valor_str).replace(',', '').replace('.', '')))
                except:
                    cantidad = 0
                
                registros.append((nombre, fecha_actualizar, cantidad))
        
        # Insertar/Actualizar en batch
        print(f"   🔄 Actualizando {len(registros)} marcas/modelos...")
        
        values_placeholders = ','.join(['(%s, %s, %s)'] * len(registros))
        flat_values = [item for tupla in registros for item in tupla]
        
        cursor.execute(f'''
            INSERT INTO {tabla} (nombre, fecha, cantidad)
            VALUES {values_placeholders}
            ON CONFLICT (nombre, fecha) 
            DO UPDATE SET cantidad = EXCLUDED.cantidad, fecha_carga = CURRENT_TIMESTAMP
        ''', flat_values)
        
        conn.commit()
        
        cursor.execute(f'SELECT COUNT(*) as total FROM {tabla}')
        total = cursor.fetchone()['total']
        
        release_db_connection(conn)
        
        # INVALIDAR CACHÉ cuando se actualizan datos manualmente
        patentamientos_cache['data'] = None
        patentamientos_cache['timestamp'] = 0
        print("🗑️ Caché de patentamientos invalidado")
        
        print(f"   ✅ Actualizado: {len(registros)} registros para {fecha_actualizar}")
        
        return jsonify({
            'success': True,
            'message': f'Mes actual actualizado: {fecha_actualizar}',
            'recordsSaved': len(registros),
            'totalRecords': total,
            'updatedMonth': fecha_actualizar
        })
        
    except Exception as e:
        print(f"❌ Error: {e}")
        import traceback
        traceback.print_exc()
        
        if conn:
            conn.rollback()
            release_db_connection(conn)
        
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bi/get_saved_data/<data_type>', methods=['GET'])
@login_required
def get_saved_data(data_type):
    """Obtener datos guardados de una tabla específica"""
    try:
        # Determinar tabla origen
        tabla_map = {
            'argentina-marca': 'bi_patentamientos_argentina_marca',
            'argentina-modelo': 'bi_patentamientos_argentina_modelo',
            'mendoza-marca': 'bi_patentamientos_mendoza_marca',
            'mendoza-modelo': 'bi_patentamientos_mendoza_modelo'
        }
        
        if data_type not in tabla_map:
            return jsonify({'success': False, 'error': 'Tipo de datos inválido'}), 400
        
        tabla = tabla_map[data_type]
        
        conn = get_db_connection()
        cursor = conn.cursor()
        
        # Verificar si hay datos
        cursor.execute(f'SELECT COUNT(*) as total FROM {tabla}')
        total = cursor.fetchone()['total']
        
        if total == 0:
            release_db_connection(conn)
            return jsonify({'success': True, 'hasData': False})
        
        # Obtener todas las fechas únicas ordenadas
        cursor.execute(f'SELECT DISTINCT fecha FROM {tabla} ORDER BY fecha')
        fechas_raw = cursor.fetchall()
        
        # Formatear fechas como "ene-15", "feb-15", etc.
        meses_nombres = {
            1: 'ene', 2: 'feb', 3: 'mar', 4: 'abr', 5: 'may', 6: 'jun',
            7: 'jul', 8: 'ago', 9: 'sep', 10: 'oct', 11: 'nov', 12: 'dic'
        }
        
        fechas = []
        for row in fechas_raw:
            fecha_obj = row['fecha']
            mes_nombre = meses_nombres[fecha_obj.month]
            anio_corto = str(fecha_obj.year)[-2:]
            fechas.append(f"{mes_nombre}-{anio_corto}")
        
        # Obtener todos los nombres únicos ordenados
        cursor.execute(f'SELECT DISTINCT nombre FROM {tabla} ORDER BY nombre')
        nombres_raw = cursor.fetchall()
        nombres = [row['nombre'] for row in nombres_raw]
        
        # Obtener TODOS los datos en una sola consulta (OPTIMIZACIÓN CRÍTICA)
        cursor.execute(f'SELECT nombre, fecha, cantidad FROM {tabla} ORDER BY nombre, fecha')
        todos_los_datos = cursor.fetchall()
        
        # Crear diccionario para acceso rápido: {(nombre, fecha): cantidad}
        datos_dict = {}
        for row in todos_los_datos:
            key = (row['nombre'], row['fecha'])
            datos_dict[key] = row['cantidad']
        
        # Construir matriz de datos usando el diccionario
        rows_data = []
        for nombre in nombres:
            row = [nombre]
            
            # Para cada fecha, obtener la cantidad del diccionario
            for fecha_obj_wrapper in fechas_raw:
                fecha_obj = fecha_obj_wrapper['fecha']
                cantidad = datos_dict.get((nombre, fecha_obj), 0)
                row.append(cantidad)
            
            rows_data.append(row)
        
        # Obtener fecha de última actualización
        cursor.execute(f'SELECT MAX(fecha_carga) as ultima_actualizacion FROM {tabla}')
        ultima_act = cursor.fetchone()['ultima_actualizacion']
        
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'hasData': True,
            'headers': ['Marca/Modelo'] + fechas,
            'rows': rows_data,
            'totalRecords': total,
            'lastUpdate': ultima_act.strftime('%Y-%m-%d %H:%M:%S') if ultima_act else None
        })
        
    except Exception as e:
        print(f"❌ Error obteniendo datos guardados: {e}")
        if 'conn' in locals():
            release_db_connection(conn)
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/bi/load_historical_data', methods=['POST'])
@login_required
def load_historical_data():
    """Invalidar caché y forzar recarga desde CSV"""
    patentamientos_cache['data'] = None
    patentamientos_cache['timestamp'] = 0
    return jsonify({'success': True, 'message': 'Caché invalidado. Los datos se recargarán en la próxima consulta.'})


@app.route('/api/bi/load_status', methods=['GET'])
@login_required
def get_load_status():
    """Verificar estado de los archivos CSV"""
    csv_dir = os.path.join(os.path.dirname(__file__), 'Patentamientos')
    archivos = [
        'Mercado Argentino MARCA.csv',
        'Mercado Argentino MODELO.csv',
        'Mercado Mendoza MARCA.csv',
        'Mercado Mendoza MODELO.csv'
    ]
    
    status = []
    for archivo in archivos:
        path = os.path.join(csv_dir, archivo)
        if os.path.exists(path):
            stat = os.stat(path)
            status.append({
                'archivo': archivo,
                'existe': True,
                'tamano': stat.st_size,
                'ultima_modificacion': datetime.fromtimestamp(stat.st_mtime).strftime('%Y-%m-%d %H:%M:%S')
            })
        else:
            status.append({
                'archivo': archivo,
                'existe': False
            })
    
    return jsonify({
        'archivos': status,
        'cache_activo': patentamientos_cache['data'] is not None
    })


# ================================
# RETAIL Y PLAN DE NEGOCIO ROUTES
# ================================

def extract_family(modelo):
    """Extract vehicle family from model name"""
    modelo_upper = str(modelo).upper()
    
    # Priority order matters (check COROLLA CROSS before COROLLA)
    if 'COROLLA CROSS' in modelo_upper:
        return 'COROLLA CROSS'
    elif 'YARIS CROSS' in modelo_upper:
        return 'YARIS CROSS'
    elif 'COROLLA' in modelo_upper:
        return 'COROLLA'
    elif 'HILUX' in modelo_upper:
        return 'HILUX'
    elif 'SW4' in modelo_upper:
        return 'SW4'
    elif 'YARIS' in modelo_upper:
        return 'YARIS'
    elif 'RAV' in modelo_upper or 'RAV4' in modelo_upper:
        return 'RAV 4'
    elif 'HIACE' in modelo_upper:
        return 'HIACE'
    else:
        return 'OTROS'


@app.route('/bi/retail_plan')
@login_required
def retail_plan():
    """Página del módulo Retail y Plan de Negocio"""
    return render_template('bi_retail_plan.html')


@app.route('/api/retail/plan', methods=['GET'])
@login_required
def get_retail_plan():
    """Obtener objetivos del plan de negocio"""
    try:
        anio = request.args.get('anio', 2025, type=int)
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT familia, convencional_objetivo, especificas_objetivo, tpa_objetivo
            FROM retail_plan
            WHERE anio = %s
            ORDER BY familia
        """, (anio,))
        
        resultados = cur.fetchall()
        plan = {}
        for row in resultados:
            plan[row['familia']] = {
                'convencional': row['convencional_objetivo'],
                'especificas': row['especificas_objetivo'],
                'tpa': row['tpa_objetivo'],
                'total': row['convencional_objetivo'] + row['especificas_objetivo'] + row['tpa_objetivo']
            }
        
        cur.close()
        conn.close()
        
        return jsonify({'success': True, 'plan': plan})
    except Exception as e:
        print(f"Error obteniendo plan: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/retail/plan', methods=['POST'])
@login_required
def update_retail_plan():
    """Actualizar objetivos del plan de negocio"""
    try:
        data = request.get_json()
        anio = data.get('anio', 2025)
        familia = data['familia']
        convencional = data['convencional']
        especificas = data['especificas']
        tpa = data['tpa']
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            INSERT INTO retail_plan 
            (anio, familia, convencional_objetivo, especificas_objetivo, tpa_objetivo, updated_at)
            VALUES (%s, %s, %s, %s, %s, CURRENT_TIMESTAMP)
            ON CONFLICT (anio, familia) 
            DO UPDATE SET 
                convencional_objetivo = EXCLUDED.convencional_objetivo,
                especificas_objetivo = EXCLUDED.especificas_objetivo,
                tpa_objetivo = EXCLUDED.tpa_objetivo,
                updated_at = CURRENT_TIMESTAMP
        """, (anio, familia, convencional, especificas, tpa))
        
        conn.commit()
        cur.close()
        conn.close()
        
        return jsonify({'success': True})
    except Exception as e:
        print(f"Error actualizando plan: {e}")
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/retail/sales_vs_plan', methods=['GET'])
@login_required
def sales_vs_plan():
    """Comparar ventas reales vs plan de negocio con desvío acumulado diario"""
    try:
        anio = request.args.get('anio', 2025, type=int)
        
        # Promedios mensuales de patentamiento histórico y días por mes
        datos_mensuales = {
            1: {'porcentaje': 11.54, 'dias': 31},   # Enero
            2: {'porcentaje': 7.29, 'dias': 28},    # Febrero (ajustar para bisiestos)
            3: {'porcentaje': 8.37, 'dias': 31},    # Marzo
            4: {'porcentaje': 7.76, 'dias': 30},    # Abril
            5: {'porcentaje': 8.37, 'dias': 31},    # Mayo
            6: {'porcentaje': 9.25, 'dias': 30},    # Junio
            7: {'porcentaje': 9.25, 'dias': 31},    # Julio
            8: {'porcentaje': 9.43, 'dias': 31},    # Agosto
            9: {'porcentaje': 8.65, 'dias': 30},    # Septiembre
            10: {'porcentaje': 8.80, 'dias': 31},   # Octubre
            11: {'porcentaje': 7.79, 'dias': 30},   # Noviembre
            12: {'porcentaje': 3.99, 'dias': 31}    # Diciembre
        }
        
        # Calcular porcentaje acumulado hasta hoy considerando días exactos
        from datetime import datetime
        hoy = datetime.now()
        anio_actual = hoy.year
        
        # Si el año solicitado es anterior al año actual, usar 100% (año completo)
        if anio < anio_actual:
            porcentaje_esperado_acumulado = 100.0
            mes_actual = 12
            dia_actual = 31
            print(f"🗓️ Año {anio} (completo) | Porcentaje esperado acumulado: {porcentaje_esperado_acumulado:.2f}%")
        else:
            # Para el año actual, usar los días transcurridos
            mes_actual = hoy.month
            dia_actual = hoy.day
            
            # Acumular meses completos anteriores
            porcentaje_esperado_acumulado = sum(datos_mensuales[m]['porcentaje'] for m in range(1, mes_actual))
            
            # Calcular porcentaje diario del mes actual
            porcentaje_diario_mes_actual = datos_mensuales[mes_actual]['porcentaje'] / datos_mensuales[mes_actual]['dias']
            
            # Agregar los días transcurridos del mes actual
            porcentaje_esperado_acumulado += (porcentaje_diario_mes_actual * dia_actual)
            
            print(f"🗓️ Fecha: {dia_actual}/{mes_actual}/{anio} | Porcentaje esperado acumulado: {porcentaje_esperado_acumulado:.2f}%")
        
        # Leer CSV de retail
        csv_path = os.path.join(os.path.dirname(__file__), 'Retail y Plan de Negocio', 'Retail y Plan de Negocio.csv')
        
        # Try different encodings
        df = None
        for encoding in ['utf-8', 'latin-1', 'cp1252', 'iso-8859-1']:
            try:
                df = pd.read_csv(csv_path, sep=';', encoding=encoding)
                break
            except:
                continue
        
        if df is None:
            return jsonify({'success': False, 'error': 'No se pudo leer el archivo CSV'}), 500
        
        # Extract family and sale type
        df['Familia'] = df['Modelo / Versión'].apply(extract_family)
        
        sale_type_map = {
            'YAC': 'Convencional',
            'TPA': 'Plan Ahorro',
            'F02': 'Vtas Especiales'
        }
        df['Tipo_Venta'] = df['Orden'].str[:3].map(sale_type_map)
        
        # Filter by year if Fecha column exists
        if 'Fecha' in df.columns:
            df['Fecha'] = pd.to_datetime(df['Fecha'], format='%d/%m/%Y', errors='coerce')
            df = df[df['Fecha'].dt.year == anio]
        
        # Group by family and sale type
        ventas_reales = df.groupby(['Familia', 'Tipo_Venta']).size().unstack(fill_value=0)
        
        # Get plan objectives
        conn = get_db_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT familia, convencional_objetivo, especificas_objetivo, tpa_objetivo
            FROM retail_plan
            WHERE anio = %s
            ORDER BY familia
        """, (anio,))
        
        plan_data = {}
        resultados_db = cur.fetchall()
        
        for row in resultados_db:
            familia = row['familia']
            plan_data[familia] = {
                'Convencional': row['convencional_objetivo'],
                'Vtas Especiales': row['especificas_objetivo'],
                'Plan Ahorro': row['tpa_objetivo']
            }
        
        cur.close()
        conn.close()
        
        # Build comparison data with accumulated deviation logic
        familias = list(set(list(ventas_reales.index) + list(plan_data.keys())))
        resultados = []
        
        for familia in familias:
            real_conv = ventas_reales.loc[familia, 'Convencional'] if familia in ventas_reales.index and 'Convencional' in ventas_reales.columns else 0
            real_espec = ventas_reales.loc[familia, 'Vtas Especiales'] if familia in ventas_reales.index and 'Vtas Especiales' in ventas_reales.columns else 0
            real_tpa = ventas_reales.loc[familia, 'Plan Ahorro'] if familia in ventas_reales.index and 'Plan Ahorro' in ventas_reales.columns else 0
            
            plan_conv = plan_data.get(familia, {}).get('Convencional', 0)
            plan_espec = plan_data.get(familia, {}).get('Vtas Especiales', 0)
            plan_tpa = plan_data.get(familia, {}).get('Plan Ahorro', 0)
            
            # Calcular objetivo acumulado esperado hasta ahora (enteros)
            objetivo_acum_conv = int(plan_conv * porcentaje_esperado_acumulado / 100)
            objetivo_acum_espec = int(plan_espec * porcentaje_esperado_acumulado / 100)
            objetivo_acum_tpa = int(plan_tpa * porcentaje_esperado_acumulado / 100)
            
            # Calcular desvío: (real - objetivo_acumulado) / objetivo_acumulado * 100
            def calcular_desvio(real, objetivo_acum):
                if objetivo_acum == 0:
                    return 0
                return round(((real - objetivo_acum) / objetivo_acum * 100), 1)
            
            resultados.append({
                'familia': familia,
                'convencional': {
                    'real': int(real_conv),
                    'objetivo_total': plan_conv,
                    'objetivo_acumulado': objetivo_acum_conv,
                    'desvio': calcular_desvio(real_conv, objetivo_acum_conv)
                },
                'especiales': {
                    'real': int(real_espec),
                    'objetivo_total': plan_espec,
                    'objetivo_acumulado': objetivo_acum_espec,
                    'desvio': calcular_desvio(real_espec, objetivo_acum_espec)
                },
                'tpa': {
                    'real': int(real_tpa),
                    'objetivo_total': plan_tpa,
                    'objetivo_acumulado': objetivo_acum_tpa,
                    'desvio': calcular_desvio(real_tpa, objetivo_acum_tpa)
                },
                'total': {
                    'real': int(real_conv + real_espec + real_tpa),
                    'objetivo_total': plan_conv + plan_espec + plan_tpa,
                    'objetivo_acumulado': objetivo_acum_conv + objetivo_acum_espec + objetivo_acum_tpa,
                    'desvio': calcular_desvio(
                        real_conv + real_espec + real_tpa,
                        objetivo_acum_conv + objetivo_acum_espec + objetivo_acum_tpa
                    )
                }
            })
        
        return jsonify({
            'success': True, 
            'datos': resultados,
            'fecha_calculo': f"{dia_actual}/{mes_actual}/{anio}",
            'porcentaje_acumulado': round(porcentaje_esperado_acumulado, 2),
            'detalles_mensuales': [
                {
                    'mes': mes,
                    'nombre': ['Enero', 'Febrero', 'Marzo', 'Abril', 'Mayo', 'Junio',
                              'Julio', 'Agosto', 'Septiembre', 'Octubre', 'Noviembre', 'Diciembre'][mes-1],
                    'porcentaje': datos_mensuales[mes]['porcentaje'],
                    'dias': datos_mensuales[mes]['dias'],
                    'porcentaje_diario': round(datos_mensuales[mes]['porcentaje'] / datos_mensuales[mes]['dias'], 3)
                }
                for mes in range(1, 13)
            ]
        })
    
    except Exception as e:
        print(f"Error en sales_vs_plan: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


# ==================== API MÓDULO 7: ANÁLISIS DE SEGUIMIENTO 2 ====================

@app.route('/api/modulo7/procesar', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def modulo7_procesar():
    """Procesar archivo Excel con hojas hoy, ayer e integra"""
    try:
        if 'file' not in request.files:
            return jsonify({'success': False, 'error': 'No se recibió ningún archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'success': False, 'error': 'No se seleccionó ningún archivo'}), 400
        
        if not allowed_file(file.filename):
            return jsonify({'success': False, 'error': 'Tipo de archivo no permitido'}), 400
        
        # Guardar archivo temporalmente
        filename = secure_filename(file.filename)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        temp_filename = f"modulo7_{timestamp}_{filename}"
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], temp_filename)
        file.save(filepath)
        
        # Procesar el archivo
        resultado = procesar_excel_modulo7(filepath)
        
        if resultado['success']:
            return jsonify(resultado)
        else:
            return jsonify(resultado), 500
            
    except Exception as e:
        print(f"❌ Error en modulo7_procesar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/modulo7/actualizar', methods=['POST'])
@login_required
@module_permission_required('planeamiento')
def modulo7_actualizar():
    """Actualizar valores editados por el usuario"""
    try:
        data = request.json
        filepath = data.get('filepath')
        cambios = data.get('cambios', [])
        
        if not filepath or not os.path.exists(filepath):
            return jsonify({'success': False, 'error': 'Archivo no encontrado'}), 404
        
        # Aplicar cambios
        resultado = aplicar_cambios_modulo7(filepath, cambios)
        
        return jsonify(resultado)
        
    except Exception as e:
        print(f"❌ Error en modulo7_actualizar: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'success': False, 'error': str(e)}), 500


@app.route('/api/modulo7/descargar/<filename>')
@login_required
@module_permission_required('planeamiento')
def modulo7_descargar(filename):
    """Descargar archivo procesado"""
    try:
        filepath = os.path.join(app.config['UPLOAD_FOLDER'], filename)
        if not os.path.exists(filepath):
            return jsonify({'error': 'Archivo no encontrado'}), 404
        
        return send_file(
            filepath,
            as_attachment=True,
            download_name=f'analisis_seguimiento_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx'
        )
        
    except Exception as e:
        print(f"❌ Error en modulo7_descargar: {e}")
        return jsonify({'error': str(e)}), 500


def procesar_excel_modulo7(filepath):
    """
    Procesa el archivo Excel replicando todas las macros de Excel en Python puro
    """
    try:
        from openpyxl import load_workbook
        from openpyxl.styles import Font
        from openpyxl.utils import get_column_letter
        from dateutil.relativedelta import relativedelta
        
        print("📂 Cargando archivo Excel...")
        wb = load_workbook(filepath)
        
        # Verificar que existan las hojas requeridas
        required_sheets = ['hoy', 'ayer', 'integra']
        for sheet_name in required_sheets:
            if sheet_name not in wb.sheetnames:
                return {'success': False, 'error': f'La hoja "{sheet_name}" no existe en el archivo'}
        
        ws_hoy = wb['hoy']
        ws_ayer = wb['ayer']
        ws_integra = wb['integra']
        
        print("🗑️  Paso 1: Eliminando primeras 8 filas de 'hoy'...")
        ws_hoy.delete_rows(1, 8)
        
        print("🔍 Paso 2: Eliminando filas con 'BAJA' en columna AB...")
        filas_a_eliminar = []
        for row in range(ws_hoy.max_row, 0, -1):
            valor_ab = ws_hoy.cell(row=row, column=28).value  # Columna AB = 28
            if valor_ab and 'BAJA' in str(valor_ab).upper():
                filas_a_eliminar.append(row)
        
        for row in filas_a_eliminar:
            ws_hoy.delete_rows(row, 1)
        
        print(f"   ✓ {len(filas_a_eliminar)} filas eliminadas")
        
        print("📊 Paso 3: Ordenando por columna M...")
        # Convertir a lista para ordenar
        data = []
        header_row = [cell.value for cell in ws_hoy[1]]
        for row in ws_hoy.iter_rows(min_row=2, values_only=True):
            data.append(list(row))
        
        # Ordenar por columna M (índice 12) - convertir a string para comparar
        def ordenar_key(x):
            valor = x[12]
            if valor is None:
                return ''
            # Convertir todo a string para comparación consistente
            return str(valor).lower()
        
        data.sort(key=ordenar_key)
        
        # Limpiar hoja y reescribir
        ws_hoy.delete_rows(2, ws_hoy.max_row)
        for row_data in data:
            ws_hoy.append(row_data)
        
        print("🔧 Paso 4: Reestructurando columnas...")
        # Eliminar columnas W (23), AF (32), AG (33)
        # Eliminar en orden inverso para no afectar índices
        for col_idx in [33, 32, 23]:
            ws_hoy.delete_cols(col_idx, 1)
        
        # Insertar columnas según macro: H(1 col), O(4 cols), AD(4 cols)
        ws_hoy.insert_cols(8, 1)   # 1 columna en H
        ws_hoy.insert_cols(15, 4)  # 4 columnas en O
        ws_hoy.insert_cols(30, 4)  # 4 columnas en AD
        
        # Cambiar fuente a Calibri 11
        for row in ws_hoy.iter_rows():
            for cell in row:
                cell.font = Font(name='Calibri', size=11)
        
        # Autoajustar ancho de columnas (aproximado)
        for column in ws_hoy.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if cell.value:
                        max_length = max(max_length, len(str(cell.value)))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws_hoy.column_dimensions[column_letter].width = adjusted_width
        
        print("📋 Paso 5: Copiando fila 1 de 'ayer' a 'hoy'...")
        from copy import copy
        for col_idx, cell in enumerate(ws_ayer[1], start=1):
            target_cell = ws_hoy.cell(row=1, column=col_idx)
            target_cell.value = cell.value
            # Copiar estilos de forma segura
            if cell.has_style:
                target_cell.font = copy(cell.font)
                target_cell.border = copy(cell.border)
                target_cell.fill = copy(cell.fill)
                target_cell.number_format = copy(cell.number_format)
                target_cell.protection = copy(cell.protection)
                target_cell.alignment = copy(cell.alignment)
        
        print("🎨 Paso 6: Copiando formato de columnas X e Y de 'ayer'...")
        ultima_fila_hoy = ws_hoy.max_row
        # Obtener estilos de referencia de la fila 2 de 'ayer'
        ref_cell_x = ws_ayer.cell(row=2, column=24)
        ref_cell_y = ws_ayer.cell(row=2, column=25)
        
        for row_idx in range(2, ultima_fila_hoy + 1):
            # Columna X (24)
            target_x = ws_hoy.cell(row=row_idx, column=24)
            target_x.number_format = ref_cell_x.number_format
            if ref_cell_x.has_style:
                target_x.font = copy(ref_cell_x.font)
            
            # Columna Y (25)
            target_y = ws_hoy.cell(row=row_idx, column=25)
            target_y.number_format = ref_cell_y.number_format
            if ref_cell_y.has_style:
                target_y.font = copy(ref_cell_y.font)
        
        print("📅 Paso 7: Calculando columna H (Fecha estimada)...")
        for row_idx in range(2, ultima_fila_hoy + 1):
            fecha_g = ws_hoy.cell(row=row_idx, column=7).value  # Columna G
            if fecha_g and isinstance(fecha_g, datetime):
                # EDATE(G, 1) - sumar 1 mes
                nueva_fecha = fecha_g + relativedelta(months=1)
                # Validar día 31 o año 1900
                if nueva_fecha.day == 31 or nueva_fecha.year == 1900:
                    ws_hoy.cell(row=row_idx, column=8).value = None
                else:
                    ws_hoy.cell(row=row_idx, column=8).value = nueva_fecha
            else:
                ws_hoy.cell(row=row_idx, column=8).value = None
        
        print("🔗 Paso 8: Aplicando VLOOKUP (columnas O, P, Q, AE, AF, AG con 'ayer')...")
        # Crear diccionario de búsqueda desde 'ayer'
        dict_ayer = {}
        for row in ws_ayer.iter_rows(min_row=2, values_only=False):
            operacion = row[13].value  # Columna N (operación) = índice 13
            if operacion:
                dict_ayer[operacion] = {
                    'O': row[14].value,  # Columna O
                    'P': row[15].value,  # Columna P
                    'Q': row[16].value,  # Columna Q
                    'AE': row[30].value if len(row) > 30 else None,  # Columna AE
                    'AF': row[31].value if len(row) > 31 else None,  # Columna AF
                    'AG': row[32].value if len(row) > 32 else None,  # Columna AG
                }
        
        # Aplicar VLOOKUP en 'hoy'
        # Después de reestructuración: O=19, P=20, Q=21, AE=38, AF=39, AG=40
        for row_idx in range(2, ultima_fila_hoy + 1):
            operacion_hoy = ws_hoy.cell(row=row_idx, column=14).value  # Columna N
            if operacion_hoy and operacion_hoy in dict_ayer:
                datos = dict_ayer[operacion_hoy]
                ws_hoy.cell(row=row_idx, column=19).value = datos.get('O')  # Columna O (ahora 19)
                ws_hoy.cell(row=row_idx, column=20).value = datos.get('P')  # Columna P (ahora 20)
                ws_hoy.cell(row=row_idx, column=21).value = datos.get('Q')  # Columna Q (ahora 21)
                ws_hoy.cell(row=row_idx, column=38).value = datos.get('AE')  # Columna AE (ahora 38)
                ws_hoy.cell(row=row_idx, column=39).value = datos.get('AF')  # Columna AF (ahora 39)
                ws_hoy.cell(row=row_idx, column=40).value = datos.get('AG')  # Columna AG (ahora 40)
        
        print("🔗 Paso 9: Aplicando VLOOKUP (columna R con 'integra')...")
        # Crear diccionario de búsqueda desde 'integra'
        dict_integra = {}
        for row in ws_integra.iter_rows(min_row=2, values_only=True):
            if row[0]:  # Columna 1: operación
                dict_integra[row[0]] = row[11] if len(row) > 11 else None  # Columna 12: ubicación
        
        # Aplicar VLOOKUP en columna R de 'hoy'
        # Después de reestructuración: R=22
        for row_idx in range(2, ultima_fila_hoy + 1):
            operacion_hoy = ws_hoy.cell(row=row_idx, column=14).value  # Columna N
            if operacion_hoy and operacion_hoy in dict_integra:
                ws_hoy.cell(row=row_idx, column=22).value = dict_integra[operacion_hoy]  # Columna R (ahora 22)
        
        print("💰 Paso 10: Calculando efectivo pendiente (columna AD)...")
        # Después de reestructuración: Z=30, AA=31, AB=32, AC=33, AD=34
        for row_idx in range(2, ultima_fila_hoy + 1):
            z = ws_hoy.cell(row=row_idx, column=30).value or 0  # Columna Z (ahora 30)
            aa = ws_hoy.cell(row=row_idx, column=31).value or 0  # Columna AA (ahora 31)
            ab = ws_hoy.cell(row=row_idx, column=32).value or 0  # Columna AB (ahora 32)
            ac = ws_hoy.cell(row=row_idx, column=33).value or 0  # Columna AC (ahora 33)
            
            # Convertir a números
            try:
                z = float(z) if z else 0
                aa = float(aa) if aa else 0
                ab = float(ab) if ab else 0
                ac = float(ac) if ac else 0
            except:
                z, aa, ab, ac = 0, 0, 0, 0
            
            resultado = z - aa - ab - ac
            ws_hoy.cell(row=row_idx, column=30).value = max(0, resultado)  # No negativos
        
        print("📊 Paso 11: Creando tabla de análisis...")
        # La tabla ya está creada desde A1 hasta AO y última fila
        
        print("🔍 Paso 12: Aplicando filtro avanzado...")
        # Filtrar filas donde:
        # - Columna P (Estado) no esté vacía O
        # - Columna Q (Estado Seg Op) no esté vacía O
        # - Columnas AE, AF, AG tengan algún valor
        
        filas_filtradas = []
        print(f"🔍 DEBUG: Leyendo columnas desde fila 2 hasta {ultima_fila_hoy}")
        
        # CÁLCULO POSICIÓN AJ: Original AJ=36, luego:
        # -3 (eliminar W,AF,AG) = 33
        # +1 (insertar H) = 34
        # +4 (insertar O) = 38  
        # +4 (insertar AD) = 42
        # Por lo tanto AJ ahora está en columna 42
        
        for row_idx in range(2, ultima_fila_hoy + 1):
            # Posiciones después de reestructuración:
            p = ws_hoy.cell(row=row_idx, column=20).value  # Estado (P ahora 20)
            q = ws_hoy.cell(row=row_idx, column=21).value  # Estado Seg Op (Q ahora 21)
            ae = ws_hoy.cell(row=row_idx, column=38).value  # EF Cancelad (AE ahora 38)
            af = ws_hoy.cell(row=row_idx, column=39).value  # usado comp (AF ahora 39)
            ag = ws_hoy.cell(row=row_idx, column=40).value  # cred liq (AG ahora 40)
            aj = ws_hoy.cell(row=row_idx, column=42).value  # Observaciones (AJ ahora 42)
            
            # Debug primeras 3 filas
            if row_idx <= 4:
                print(f"  Fila {row_idx}: AJ(col 42) = '{aj}' | P={p} | Q={q}")
            
            if p or q or ae or af or ag:
                fila_data = {
                    'row_idx': row_idx,
                    'operacion': ws_hoy.cell(row=row_idx, column=14).value,  # N sigue en 14
                    'ejecutivo': ws_hoy.cell(row=row_idx, column=19).value,  # O ahora 19
                    'estado': p,
                    'estado_seg_op': q,
                    'ef_cancelad': ae,
                    'usado_comp': af,
                    'cred_liq': ag,
                    'observaciones': aj
                }
                filas_filtradas.append(fila_data)
        
        # Guardar archivo procesado
        # Forzar el cálculo de fórmulas al abrir
        wb.calculation.calcMode = 'auto'
        wb.save(filepath)
        
        print(f"✅ Procesamiento completado: {len(filas_filtradas)} filas filtradas")
        
        return {
            'success': True,
            'filepath': filepath,
            'filename': os.path.basename(filepath),
            'total_filas': ultima_fila_hoy - 1,
            'filas_filtradas': len(filas_filtradas),
            'datos_filtrados': filas_filtradas
        }
        
    except Exception as e:
        print(f"❌ Error en procesar_excel_modulo7: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


def aplicar_cambios_modulo7(filepath, cambios):
    """
    Aplica los cambios editados por el usuario al archivo Excel
    """
    try:
        from openpyxl import load_workbook
        
        print(f"📝 Aplicando {len(cambios)} cambios...")
        
        wb = load_workbook(filepath)
        ws_hoy = wb['hoy']
        
        cambios_aplicados = 0
        for cambio in cambios:
            row_idx = cambio.get('row_idx')
            campo = cambio.get('campo')
            valor = cambio.get('valor')
            
            if not row_idx or not campo:
                continue
            
            # Mapear campo a columna (solo campos editables, observaciones es solo lectura)
            # Posiciones después de reestructuración
            col_map = {
                'ejecutivo': 19,      # O (ahora 19)
                'estado': 20,         # P (ahora 20)
                'estado_seg_op': 21,  # Q (ahora 21)
                'ef_cancelad': 38,    # AE (ahora 38)
                'usado_comp': 39,     # AF (ahora 39)
                'cred_liq': 40        # AG (ahora 40)
            }
            
            if campo in col_map:
                ws_hoy.cell(row=row_idx, column=col_map[campo]).value = valor
                cambios_aplicados += 1
        
        # Forzar el cálculo de fórmulas al abrir
        wb.calculation.calcMode = 'auto'
        wb.save(filepath)
        
        print(f"✅ {cambios_aplicados} cambios aplicados correctamente")
        
        return {
            'success': True,
            'cambios_aplicados': cambios_aplicados
        }
        
    except Exception as e:
        print(f"❌ Error en aplicar_cambios_modulo7: {e}")
        import traceback
        traceback.print_exc()
        return {'success': False, 'error': str(e)}


# ==================== APIs MÓDULO USADOS ====================

# --- APIs Marcas ---
@app.route('/api/usados/marcas', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_marcas_usados():
    """Obtener todas las marcas de usados"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nombre
            FROM marcas_usados
            WHERE activo = TRUE
            ORDER BY nombre ASC
        """)
        marcas = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(m) for m in marcas])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/marcas', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_marca_usados():
    """Crear nueva marca"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO marcas_usados (nombre)
            VALUES (%s)
            RETURNING id
        """, (data['nombre'],))
        
        marca_id = cursor.fetchone()['id']
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': marca_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# --- APIs Modelos ---
@app.route('/api/usados/modelos/<int:marca_id>', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_modelos_usados(marca_id):
    """Obtener modelos de una marca"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nombre
            FROM modelos_usados
            WHERE marca_id = %s AND activo = TRUE
            ORDER BY nombre ASC
        """, (marca_id,))
        modelos = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(m) for m in modelos])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/modelos', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_modelo_usados():
    """Crear nuevo modelo"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO modelos_usados (marca_id, nombre)
            VALUES (%s, %s)
            RETURNING id
        """, (data['marca_id'], data['nombre']))
        
        modelo_id = cursor.fetchone()['id']
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': modelo_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# --- APIs Ingresos ---
@app.route('/api/usados/ingresos', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_ingresos_usados():
    """Obtener todos los ingresos"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                i.id,
                i.dominio,
                i.limpieza_requerida,
                i.fecha_ingreso,
                i.estado,
                i.observaciones,
                m.nombre as marca_nombre,
                mo.nombre as modelo_nombre
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            WHERE i.activo = TRUE
            ORDER BY i.fecha_ingreso DESC
        """)
        ingresos = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(i) for i in ingresos])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/ingresos', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_ingreso_usados():
    """Crear nuevo ingreso de vehículo y programarlo automáticamente"""
    conn = None
    try:
        from datetime import datetime, timedelta
        
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Insertar ingreso (siempre clasificación='USADOS', es_stock_fijo=FALSE)
        cursor.execute("""
            INSERT INTO ingresos_usados 
            (dominio, marca_id, modelo_id, limpieza_requerida, observaciones, usuario_ingreso, estado, clasificacion, es_stock_fijo)
            VALUES (%s, %s, %s, %s, %s, %s, 'Playa Lavadero', 'USADOS', FALSE)
            RETURNING id
        """, (
            data['dominio'],
            data['marca_id'],
            data['modelo_id'],
            data['limpieza_requerida'],
            data.get('observaciones'),
            current_user.username
        ))
        
        ingreso_id = cursor.fetchone()['id']
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, NULL, 'Playa Lavadero', %s)
        """, (ingreso_id, current_user.username))
        
        print(f"DEBUG: Vehículo {data['dominio']} ingresado a Playa Lavadero")
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': ingreso_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500


@app.route('/api/usados/ingresos/<int:ingreso_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
def vender_vehiculo_usados(ingreso_id):
    """Eliminar vehículo del stock (venta/egreso)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que el vehículo existe
        cursor.execute("""
            SELECT dominio, clasificacion FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        
        vehiculo = cursor.fetchone()
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        # Registrar en historial antes de eliminar
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario, observaciones)
            SELECT %s, estado, 'VENDIDO', %s, 'Vehículo vendido - Egreso del stock'
            FROM ingresos_usados WHERE id = %s
        """, (ingreso_id, current_user.username, ingreso_id))
        
        # Eliminar planificaciones asociadas (si existen)
        cursor.execute("""
            DELETE FROM planificacion_operaciones WHERE ingreso_id = %s
        """, (ingreso_id,))
        
        # Marcar como inactivo en lugar de eliminar (soft delete)
        cursor.execute("""
            UPDATE ingresos_usados 
            SET activo = FALSE, estado = 'VENDIDO'
            WHERE id = %s
        """, (ingreso_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'mensaje': f'Vehículo {vehiculo["dominio"]} vendido exitosamente'
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en vender_vehiculo_usados: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# --- APIs Stock (KINTO, TEST DRIVE, USADOS) ---
@app.route('/api/usados/stock', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_stock():
    """Obtener todos los vehículos del stock (KINTO, TEST DRIVE, USADOS)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                i.id,
                i.dominio,
                i.clasificacion,
                i.es_stock_fijo,
                i.estado,
                i.fecha_ingreso,
                m.nombre as marca_nombre,
                mo.nombre as modelo_nombre,
                (SELECT MAX(tiempo_fin_real) 
                 FROM planificacion_operaciones po 
                 WHERE po.ingreso_id = i.id AND po.tiempo_fin_real IS NOT NULL) as ultimo_lavado,
                COALESCE(
                    (SELECT MIN(fecha_planificada)
                     FROM planificacion_operaciones po
                     WHERE po.ingreso_id = i.id 
                       AND po.completado = FALSE),
                    (SELECT MIN(fecha_reserva::DATE + hora_inicio::TIME)
                     FROM reservas_kinto rk
                     WHERE rk.dominio = i.dominio
                       AND rk.estado = 'Reservado')
                ) as proximo_turno
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            WHERE i.activo = TRUE
            ORDER BY 
                CASE i.clasificacion
                    WHEN 'KINTO' THEN 1
                    WHEN 'TEST DRIVE' THEN 2
                    WHEN 'USADOS' THEN 3
                END,
                i.dominio ASC
        """)
        stock = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(s) for s in stock])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_stock():
    """Crear nuevo vehículo de stock fijo (KINTO o TEST DRIVE)"""
    conn = None
    try:
        data = request.get_json()
        clasificacion = data.get('clasificacion')
        
        # Validar que solo se puedan crear KINTO o TEST DRIVE
        if clasificacion not in ['KINTO', 'TEST DRIVE']:
            return jsonify({'error': 'Solo se pueden crear vehículos KINTO o TEST DRIVE desde aquí'}), 400
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar si el dominio ya existe
        cursor.execute("SELECT id FROM ingresos_usados WHERE dominio = %s AND activo = TRUE", (data['dominio'],))
        if cursor.fetchone():
            release_db_connection(conn)
            return jsonify({'error': 'Ya existe un vehículo con ese dominio'}), 400
        
        # Obtener ID de TOYOTA (todos los KINTO/TEST DRIVE son TOYOTA)
        cursor.execute("""
            SELECT id FROM marcas_usados WHERE nombre = 'TOYOTA'
        """)
        marca_toyota = cursor.fetchone()
        marca_id = marca_toyota['id'] if marca_toyota else 1
        
        # Obtener o crear modelo (siempre con TOYOTA)
        modelo_id = None
        if 'modelo_nombre' in data:
            cursor.execute("""
                SELECT id FROM modelos_usados WHERE nombre = %s AND marca_id = %s
            """, (data['modelo_nombre'], marca_id))
            modelo = cursor.fetchone()
            
            if modelo:
                modelo_id = modelo['id']
            else:
                cursor.execute("""
                    INSERT INTO modelos_usados (nombre, marca_id)
                    VALUES (%s, %s)
                    RETURNING id
                """, (data['modelo_nombre'], marca_id))
                modelo_id = cursor.fetchone()['id']
        
        # Insertar vehículo
        cursor.execute("""
            INSERT INTO ingresos_usados 
            (dominio, marca_id, modelo_id, clasificacion, es_stock_fijo, 
             limpieza_requerida, estado, usuario_ingreso)
            VALUES (%s, %s, %s, %s, TRUE, 'Standard Plus', 'Salón', %s)
            RETURNING id
        """, (
            data['dominio'].upper(),
            marca_id,
            modelo_id,
            clasificacion,
            current_user.username
        ))
        
        ingreso_id = cursor.fetchone()['id']
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, NULL, 'Salón', %s)
        """, (ingreso_id, current_user.username))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': ingreso_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock/<int:ingreso_id>', methods=['PUT'])
@login_required
@module_permission_required('usados')
def actualizar_stock(ingreso_id):
    """Actualizar datos de vehículo KINTO o TEST DRIVE (solo supervisores)"""
    if current_user.role != 'supervisor':
        return jsonify({'error': 'No autorizado. Solo supervisores pueden editar vehículos.'}), 403
    
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que sea KINTO o TEST DRIVE
        cursor.execute("""
            SELECT clasificacion FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if vehiculo['clasificacion'] not in ['KINTO', 'TEST DRIVE']:
            release_db_connection(conn)
            return jsonify({'error': 'Solo se pueden editar vehículos KINTO y TEST DRIVE'}), 400
        
        # Obtener marca TOYOTA
        cursor.execute("""
            SELECT id FROM marcas_usados WHERE nombre = 'TOYOTA'
        """)
        marca = cursor.fetchone()
        marca_id = marca['id'] if marca else None
        
        # Obtener o crear modelo
        modelo_id = None
        if 'modelo_nombre' in data and data['modelo_nombre']:
            cursor.execute("""
                SELECT id FROM modelos_usados WHERE nombre = %s AND marca_id = %s
            """, (data['modelo_nombre'], marca_id))
            modelo = cursor.fetchone()
            
            if modelo:
                modelo_id = modelo['id']
            else:
                cursor.execute("""
                    INSERT INTO modelos_usados (nombre, marca_id)
                    VALUES (%s, %s)
                    RETURNING id
                """, (data['modelo_nombre'], marca_id))
                modelo_id = cursor.fetchone()['id']
        
        # Actualizar vehículo
        cursor.execute("""
            UPDATE ingresos_usados
            SET dominio = %s,
                modelo_id = %s
            WHERE id = %s
        """, (
            data.get('dominio', '').upper(),
            modelo_id,
            ingreso_id
        ))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock/<int:ingreso_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
def eliminar_stock(ingreso_id):
    """Eliminar vehículo KINTO o TEST DRIVE (solo supervisores)"""
    if current_user.role != 'supervisor':
        return jsonify({'error': 'No autorizado. Solo supervisores pueden eliminar vehículos.'}), 403
    
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que sea KINTO o TEST DRIVE
        cursor.execute("""
            SELECT clasificacion FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if vehiculo['clasificacion'] not in ['KINTO', 'TEST DRIVE']:
            release_db_connection(conn)
            return jsonify({'error': 'Solo se pueden eliminar vehículos KINTO y TEST DRIVE'}), 400
        
        # Marcar como inactivo en lugar de eliminar físicamente
        cursor.execute("""
            UPDATE ingresos_usados
            SET activo = FALSE
            WHERE id = %s
        """, (ingreso_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock/<int:ingreso_id>/ubicacion', methods=['PUT'])
@login_required
@module_permission_required('usados')
def cambiar_ubicacion_stock(ingreso_id):
    """Cambiar ubicación de vehículo USADOS entre Salón y Reparación (solo supervisores)"""
    if current_user.role != 'supervisor':
        return jsonify({'error': 'No autorizado. Solo supervisores pueden cambiar ubicaciones.'}), 403
    
    conn = None
    try:
        data = request.get_json()
        nueva_ubicacion = data.get('ubicacion')
        
        if nueva_ubicacion not in ['Salón', 'Reparación']:
            return jsonify({'error': 'Ubicación no válida'}), 400
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que sea USADOS
        cursor.execute("""
            SELECT clasificacion, estado FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if vehiculo['clasificacion'] != 'USADOS':
            release_db_connection(conn)
            return jsonify({'error': 'Solo vehículos USADOS pueden cambiar a Reparación'}), 400
        
        # Actualizar ubicación
        estado_anterior = vehiculo['estado'] or 'Salón'
        
        cursor.execute("""
            UPDATE ingresos_usados
            SET estado = %s
            WHERE id = %s
        """, (nueva_ubicacion, ingreso_id))
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, %s, %s, %s)
        """, (ingreso_id, estado_anterior, nueva_ubicacion, current_user.username))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock/solicitar-turno', methods=['POST'])
@login_required
@module_permission_required('usados')
def solicitar_turno_stock():
    """Programar turno de lavado para KINTO o TEST DRIVE directamente al Gantt"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que sea KINTO o TEST DRIVE
        cursor.execute("""
            SELECT clasificacion, dominio FROM ingresos_usados WHERE id = %s
        """, (data['ingreso_id'],))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if vehiculo['clasificacion'] not in ['KINTO', 'TEST DRIVE']:
            release_db_connection(conn)
            return jsonify({'error': 'Solo se pueden solicitar turnos para KINTO y TEST DRIVE'}), 400
        
        # Validar conflictos (máximo 2 operaciones simultáneas)
        from datetime import datetime
        h_ini, m_ini = map(int, data['hora_inicio'].split(':')[:2])
        h_fin, m_fin = map(int, data['hora_fin'].split(':')[:2])
        minuto_inicio = h_ini * 60 + m_ini
        minuto_fin = h_fin * 60 + m_fin
        
        # Validar horarios según día de semana
        fecha_obj = datetime.strptime(str(data['fecha']), '%Y-%m-%d')
        dia_semana = fecha_obj.weekday()
        
        if dia_semana == 5:  # Sábado
            if minuto_inicio < 9 * 60 or minuto_fin > 13 * 60:
                release_db_connection(conn)
                return jsonify({'error': 'Los sábados solo se permiten turnos entre 9:00 y 13:00'}), 400
        else:
            # Validar franja de descanso si es jornada cortada
            cursor.execute("""
                SELECT tipo_jornada FROM config_jornadas_lavadero WHERE fecha = %s
            """, (data['fecha'],))
            jornada = cursor.fetchone()
            es_jornada_cortada = not jornada or jornada['tipo_jornada'] != 'completa'
            
            if es_jornada_cortada:
                if (minuto_inicio < 16 * 60 and minuto_fin > 13 * 60):
                    release_db_connection(conn)
                    return jsonify({'error': 'No se pueden asignar turnos en la franja de descanso (13:00-16:00)'}), 400
        
        # Contar conflictos
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
            AND (
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s
                 AND EXTRACT(HOUR FROM hora_fin) * 60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) >= %s
                 AND EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (data['fecha'], minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        conflictos_usados = cursor.fetchone()['count']
        
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM reservas_kinto
            WHERE fecha_reserva = %s AND estado != 'Cancelado'
            AND (
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s
                 AND EXTRACT(HOUR FROM hora_fin) * 60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) >= %s
                 AND EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (data['fecha'], minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        conflictos_kinto = cursor.fetchone()['count']
        
        if (conflictos_usados + conflictos_kinto) >= 2:
            release_db_connection(conn)
            return jsonify({'error': 'No hay espacio disponible. Máximo 2 operaciones simultáneas.'}), 400
        
        # Obtener último orden de ejecución
        cursor.execute("""
            SELECT COALESCE(MAX(orden_ejecucion), 0) as max_orden
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
        """, (data['fecha'],))
        max_orden = cursor.fetchone()['max_orden']
        
        # Insertar planificación
        cursor.execute("""
            INSERT INTO planificacion_operaciones
            (ingreso_id, fecha_planificada, hora_inicio, hora_fin, 
             posicion_lavadero, operacion_lavado_id, orden_ejecucion)
            VALUES (%s, %s, %s, %s, 1, %s, %s)
            RETURNING id
        """, (
            data['ingreso_id'],
            data['fecha'],
            data['hora_inicio'],
            data['hora_fin'],
            data['operacion_id'],
            max_orden + 1
        ))
        
        planificacion_id = cursor.fetchone()['id']
        
        # Actualizar estado del vehículo
        cursor.execute("""
            UPDATE ingresos_usados
            SET estado = 'Lavado Programado'
            WHERE id = %s
        """, (data['ingreso_id'],))
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, 'Salón', 'Lavado Programado', %s)
        """, (data['ingreso_id'], current_user.username))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': planificacion_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en solicitar_turno_stock: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/stock/solicitar-turno-gap', methods=['POST'])
@login_required
@module_permission_required('usados')
def solicitar_turno_gap_stock():
    """Programar turno desde calendario (gap) con Standard Plus fijo"""
    conn = None
    try:
        data = request.get_json()
        ingreso_id = data['ingreso_id']
        fecha = data['fecha']
        hora_inicio = data['hora_inicio']
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que sea KINTO o TEST DRIVE
        cursor.execute("""
            SELECT clasificacion, dominio FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if vehiculo['clasificacion'] not in ['KINTO', 'TEST DRIVE']:
            release_db_connection(conn)
            return jsonify({'error': 'Solo se pueden solicitar turnos para KINTO y TEST DRIVE'}), 400
        
        # Obtener ID y duración de Standard Plus
        cursor.execute("""
            SELECT id, duracion_minutos 
            FROM config_operaciones_lavado 
            WHERE nombre = 'Standard Plus'
        """)
        standard_plus = cursor.fetchone()
        if not standard_plus:
            release_db_connection(conn)
            return jsonify({'error': 'Standard Plus no configurado'}), 400
        
        operacion_id = standard_plus['id']
        duracion = standard_plus['duracion_minutos']
        
        # Calcular hora_fin
        from datetime import datetime, timedelta
        hora_inicio_dt = datetime.strptime(hora_inicio, '%H:%M')
        hora_fin_dt = hora_inicio_dt + timedelta(minutes=duracion)
        hora_fin = hora_fin_dt.strftime('%H:%M:%S')
        hora_inicio_full = hora_inicio + ':00'
        
        # Validar conflictos
        h_ini, m_ini = map(int, hora_inicio.split(':'))
        minuto_inicio = h_ini * 60 + m_ini
        minuto_fin = minuto_inicio + duracion
        
        # Validar día de semana
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana = fecha_obj.weekday()
        
        if dia_semana == 5:  # Sábado
            if minuto_inicio < 9 * 60 or minuto_fin > 13 * 60:
                release_db_connection(conn)
                return jsonify({'error': 'Los sábados solo se permiten turnos entre 9:00 y 13:00'}), 400
        else:
            # Validar franja de descanso si es jornada cortada
            cursor.execute("""
                SELECT tipo_jornada FROM config_jornadas_lavadero WHERE fecha = %s
            """, (fecha,))
            jornada = cursor.fetchone()
            es_jornada_cortada = not jornada or jornada['tipo_jornada'] != 'completa'
            
            if es_jornada_cortada:
                if (minuto_inicio < 16 * 60 and minuto_fin > 13 * 60):
                    release_db_connection(conn)
                    return jsonify({'error': 'No se pueden asignar turnos en la franja de descanso (13:00-16:00)'}), 400
        
        # Verificar conflictos con operaciones Usados
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
            AND (
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s 
                 AND EXTRACT(HOUR FROM hora_fin)*60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) >= %s 
                 AND EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        
        conflictos = cursor.fetchone()['count']
        
        # Verificar conflictos con otras reservas Kinto
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM reservas_kinto
            WHERE fecha_reserva = %s
            AND estado = 'Reservado'
            AND (
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s 
                 AND EXTRACT(HOUR FROM hora_fin)*60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) >= %s 
                 AND EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        
        conflictos_kinto = cursor.fetchone()['count']
        
        if (conflictos + conflictos_kinto) >= 2:
            release_db_connection(conn)
            return jsonify({'error': 'No hay espacio disponible. Máximo 2 operaciones simultáneas.'}), 400
        
        # Bug 3: Verificar que este vehículo no tenga ya un turno programado para el mismo día
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM planificacion_operaciones
            WHERE ingreso_id = %s
            AND fecha_planificada::DATE = %s
        """, (ingreso_id, fecha))
        
        turno_existente = cursor.fetchone()['count']
        if turno_existente > 0:
            release_db_connection(conn)
            return jsonify({'error': f'El vehículo {vehiculo["dominio"]} ya tiene un turno programado para este día'}), 400
        
        # Obtener último orden de ejecución
        cursor.execute("""
            SELECT COALESCE(MAX(orden_ejecucion), 0) as max_orden
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
        """, (fecha,))
        max_orden = cursor.fetchone()['max_orden']
        
        # Insertar en planificacion_operaciones
        cursor.execute("""
            INSERT INTO planificacion_operaciones
            (ingreso_id, fecha_planificada, hora_inicio, hora_fin, 
             posicion_lavadero, operacion_lavado_id, orden_ejecucion)
            VALUES (%s, %s::date + %s::time, %s, %s, 1, %s, %s)
            RETURNING id
        """, (
            ingreso_id,
            fecha,
            hora_inicio_full,  # Para fecha_planificada (fecha + hora)
            hora_inicio_full,  # Para hora_inicio
            hora_fin,          # Para hora_fin
            operacion_id,
            max_orden + 1
        ))
        
        planificacion_id = cursor.fetchone()['id']
        
        # Actualizar estado del vehículo
        cursor.execute("""
            UPDATE ingresos_usados
            SET estado = 'Lavado Programado'
            WHERE id = %s
        """, (ingreso_id,))
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, 'Salón', 'Lavado Programado', %s)
        """, (ingreso_id, current_user.username))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'id': planificacion_id,
            'mensaje': f'Turno programado para {vehiculo["dominio"]}'
        }), 201
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en solicitar_turno_gap_stock: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# --- APIs Operaciones de Lavado (Configuración) ---
@app.route('/api/usados/operaciones', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_operaciones_lavado():
    """Obtener todas las operaciones de lavado"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT id, nombre, duracion_minutos, es_sistema
            FROM config_operaciones_lavado
            WHERE activo = TRUE
            ORDER BY es_sistema DESC, nombre ASC
        """)
        operaciones = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(o) for o in operaciones])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/operaciones', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_operacion_lavado():
    """Crear nueva operación personalizada"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO config_operaciones_lavado (nombre, duracion_minutos, es_sistema)
            VALUES (%s, %s, FALSE)
            RETURNING id
        """, (data['nombre'], data['duracion_minutos']))
        
        operacion_id = cursor.fetchone()['id']
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': operacion_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/operaciones/<int:operacion_id>', methods=['PUT'])
@login_required
@module_permission_required('usados')
def actualizar_operacion_lavado(operacion_id):
    """Actualizar duración de operación"""
    conn = None
    try:
        data = request.get_json()
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE config_operaciones_lavado
            SET duracion_minutos = %s
            WHERE id = %s
        """, (data['duracion_minutos'], operacion_id))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/operaciones/<int:operacion_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
def eliminar_operacion_lavado(operacion_id):
    """Eliminar operación personalizada (no del sistema)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que no sea del sistema
        cursor.execute("SELECT es_sistema FROM config_operaciones_lavado WHERE id = %s", (operacion_id,))
        operacion = cursor.fetchone()
        
        if operacion and operacion['es_sistema']:
            release_db_connection(conn)
            return jsonify({'error': 'No se pueden eliminar operaciones del sistema'}), 400
        
        cursor.execute("""
            UPDATE config_operaciones_lavado
            SET activo = FALSE
            WHERE id = %s
        """, (operacion_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# --- APIs Cola y Planificación ---
@app.route('/api/usados/cola', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_cola_lavadero():
    """Obtener vehículos en estado Playa Lavadero (cola)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                i.id,
                i.dominio,
                i.limpieza_requerida,
                i.fecha_ingreso,
                m.nombre as marca_nombre,
                mo.nombre as modelo_nombre
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            WHERE i.estado = 'Playa Lavadero' 
            AND i.activo = TRUE
            AND i.clasificacion = 'USADOS'
            ORDER BY i.fecha_ingreso ASC
        """)
        cola = cursor.fetchall()
        release_db_connection(conn)
        return jsonify([dict(v) for v in cola])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/planificacion/<fecha>', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_planificacion(fecha):
    """Obtener planificación del día (Usados + Kinto)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Operaciones Usados (incluye KINTO, TEST DRIVE, USADOS programados y tareas cíclicas)
        # Muestra TODAS las operaciones del día (pendientes y completadas)
        # Las completadas se renderizan con barra gris y checkmark
        # Las tareas cíclicas se muestran pero no aparecen en el panel de lavadero
        cursor.execute("""
            SELECT 
                p.id,
                p.fecha_planificada::DATE as fecha,
                p.hora_inicio::TEXT as hora_inicio,
                p.hora_fin::TEXT as hora_fin,
                p.posicion_lavadero,
                p.orden_ejecucion,
                co.duracion_minutos,
                COALESCE(i.dominio, co.nombre) as dominio,
                COALESCE(i.limpieza_requerida, co.nombre) as limpieza_requerida,
                CASE 
                    WHEN p.es_tarea_ciclica = TRUE THEN 'TAREA_CICLICA'
                    WHEN i.id IS NULL THEN 'TAREA_MANUAL'
                    ELSE COALESCE(i.clasificacion, 'USADOS')
                END as clasificacion,
                co.nombre as operacion_nombre,
                COALESCE(i.estado, 'Completado') as estado,
                CASE 
                    WHEN p.es_tarea_ciclica = TRUE THEN 'ciclica'
                    WHEN i.id IS NULL THEN 'manual' 
                    ELSE 'usados' 
                END as tipo,
                p.tiempo_fin_real,
                COALESCE(p.es_tarea_ciclica, FALSE) as es_tarea_ciclica
            FROM planificacion_operaciones p
            LEFT JOIN ingresos_usados i ON p.ingreso_id = i.id
            LEFT JOIN config_operaciones_lavado co ON p.operacion_lavado_id = co.id
            WHERE p.fecha_planificada::DATE = %s
        """, (fecha,))
        operaciones_usados = cursor.fetchall()
        
        # Reservas Kinto
        cursor.execute("""
            SELECT 
                k.id,
                k.fecha_reserva as fecha,
                k.hora_inicio::TEXT as hora_inicio,
                k.hora_fin::TEXT as hora_fin,
                1 as posicion_lavadero,
                NULL as orden_ejecucion,
                k.duracion_minutos,
                k.dominio,
                'Standard Plus' as limpieza_requerida,
                'Standard Plus' as operacion_nombre,
                k.estado,
                'kinto' as tipo
            FROM reservas_kinto k
            WHERE k.fecha_reserva = %s
            AND k.estado = 'Reservado'
        """, (fecha,))
        reservas_kinto = cursor.fetchall()
        
        # Combinar y ordenar por hora
        todas = list(operaciones_usados) + list(reservas_kinto)
        todas.sort(key=lambda x: x['hora_inicio'])
        
        release_db_connection(conn)
        return jsonify([dict(op) for op in todas])
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/planificacion/<int:operacion_id>', methods=['PUT'])
@login_required
@module_permission_required('usados')
def actualizar_horario_operacion(operacion_id):
    """Actualizar horarios de una operación (Usados o Kinto) y reordenar"""
    conn = None
    try:
        data = request.get_json()
        tipo = data.get('tipo', 'usados')
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener fecha actual de la operación según tipo
        if tipo == 'kinto':
            cursor.execute("""
                SELECT fecha_reserva as fecha
                FROM reservas_kinto
                WHERE id = %s
            """, (operacion_id,))
        else:
            cursor.execute("""
                SELECT fecha_planificada::DATE as fecha
                FROM planificacion_operaciones
                WHERE id = %s
            """, (operacion_id,))
        
        operacion_actual = cursor.fetchone()
        if not operacion_actual:
            release_db_connection(conn)
            return jsonify({'error': 'Operación no encontrada'}), 404
        
        fecha_operacion = data.get('fecha', operacion_actual['fecha'])
        
        # Validar conflictos: contar operaciones simultáneas (Usados + Kinto)
        # Convertir horas a minutos
        h_ini, m_ini = map(int, data['hora_inicio'].split(':')[:2])
        h_fin, m_fin = map(int, data['hora_fin'].split(':')[:2])
        minuto_inicio = h_ini * 60 + m_ini
        minuto_fin = h_fin * 60 + m_fin
        
        # ===== VALIDACIÓN 1: Verificar día de la semana (sábados solo 9:00-13:00) =====
        from datetime import datetime
        fecha_obj = datetime.strptime(str(fecha_operacion), '%Y-%m-%d')
        dia_semana = fecha_obj.weekday()  # 0=Lunes, 5=Sábado, 6=Domingo
        
        if dia_semana == 5:  # Sábado
            # Horario permitido: 9:00-13:00 (540 a 780 minutos)
            if minuto_inicio < 9 * 60 or minuto_fin > 13 * 60:
                release_db_connection(conn)
                return jsonify({'error': '❌ Los sábados solo se permiten turnos entre 9:00 y 13:00'}), 400
        
        # ===== VALIDACIÓN 2: Verificar franjas de descanso (solo jornadas cortadas) =====
        cursor.execute("""
            SELECT tipo_jornada FROM config_jornadas_lavadero
            WHERE fecha = %s
        """, (fecha_operacion,))
        jornada_config = cursor.fetchone()
        
        # Si NO hay jornada completa configurada (o sea, es jornada cortada)
        es_jornada_cortada = not jornada_config or jornada_config['tipo_jornada'] != 'completa'
        
        if es_jornada_cortada and dia_semana != 5:  # No aplicar a sábados (ya tienen su horario especial)
            # Horario cortado: 8:30-13:00 (510-780) y 16:00-20:00 (960-1200)
            # Zona prohibida: 13:00-16:00 (780-960)
            
            # Verificar si el turno cae en la zona de descanso
            if (minuto_inicio < 16 * 60 and minuto_fin > 13 * 60):
                release_db_connection(conn)
                return jsonify({'error': '❌ No se pueden asignar turnos en la franja de descanso (13:00-16:00). Configura una jornada completa si es necesario.'}), 400
        
        # Contar conflictos en Usados
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
            AND NOT (id = %s AND %s = 'usados')
            AND (
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s
                 AND EXTRACT(HOUR FROM hora_fin) * 60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) >= %s
                 AND EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha_operacion, operacion_id, tipo, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        conflictos_usados = cursor.fetchone()['count']
        
        # Contar conflictos en Kinto
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM reservas_kinto
            WHERE fecha_reserva = %s
            AND NOT (id = %s AND %s = 'kinto')
            AND (
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s
                 AND EXTRACT(HOUR FROM hora_fin) * 60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) >= %s
                 AND EXTRACT(HOUR FROM hora_inicio) * 60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha_operacion, operacion_id, tipo, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        conflictos_kinto = cursor.fetchone()['count']
        
        total_conflictos = conflictos_usados + conflictos_kinto
        
        if total_conflictos >= 2:
            release_db_connection(conn)
            return jsonify({'error': 'No hay espacio disponible. Máximo 2 operaciones simultáneas.'}), 400
        
        # Actualizar según tipo
        if tipo == 'kinto':
            if 'fecha' in data:
                cursor.execute("""
                    UPDATE reservas_kinto
                    SET hora_inicio = %s, hora_fin = %s, fecha_reserva = %s
                    WHERE id = %s
                """, (data['hora_inicio'], data['hora_fin'], data['fecha'], operacion_id))
            else:
                cursor.execute("""
                    UPDATE reservas_kinto
                    SET hora_inicio = %s, hora_fin = %s
                    WHERE id = %s
                """, (data['hora_inicio'], data['hora_fin'], operacion_id))
        else:
            if 'fecha' in data:
                cursor.execute("""
                    UPDATE planificacion_operaciones
                    SET hora_inicio = %s, hora_fin = %s, fecha_planificada = %s
                    WHERE id = %s
                """, (data['hora_inicio'], data['hora_fin'], data['fecha'], operacion_id))
            else:
                cursor.execute("""
                    UPDATE planificacion_operaciones
                    SET hora_inicio = %s, hora_fin = %s
                    WHERE id = %s
                """, (data['hora_inicio'], data['hora_fin'], operacion_id))
            
            # Reordenar operaciones de Usados del día
            cursor.execute("""
                WITH operaciones_ordenadas AS (
                    SELECT id, ROW_NUMBER() OVER (ORDER BY hora_inicio ASC) as nuevo_orden
                    FROM planificacion_operaciones
                    WHERE fecha_planificada::DATE = %s
                )
                UPDATE planificacion_operaciones p
                SET orden_ejecucion = o.nuevo_orden
                FROM operaciones_ordenadas o
                WHERE p.id = o.id
            """, (fecha_operacion,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500


@app.route('/api/usados/planificacion/<int:operacion_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
@usados_section_required('planificacion')
def dar_baja_operacion(operacion_id):
    """Dar de baja una operación programada (Usados o Kinto)"""
    conn = None
    try:
        data = request.get_json()
        tipo = data.get('tipo', 'usados')
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Eliminar según tipo
        if tipo == 'kinto':
            # Eliminar de reservas_kinto
            cursor.execute("""
                DELETE FROM reservas_kinto
                WHERE id = %s
                RETURNING dominio
            """, (operacion_id,))
            
            deleted = cursor.fetchone()
            if not deleted:
                release_db_connection(conn)
                return jsonify({'error': 'Operación no encontrada'}), 404
        else:
            # Para operaciones de Usados, obtener el ingreso_id antes de eliminar
            cursor.execute("""
                SELECT ingreso_id FROM planificacion_operaciones
                WHERE id = %s
            """, (operacion_id,))
            
            result = cursor.fetchone()
            if not result:
                release_db_connection(conn)
                return jsonify({'error': 'Operación no encontrada'}), 404
            
            ingreso_id = result['ingreso_id']
            
            # Eliminar la operación
            cursor.execute("""
                DELETE FROM planificacion_operaciones
                WHERE id = %s
                RETURNING id
            """, (operacion_id,))
            
            deleted = cursor.fetchone()
            if not deleted:
                release_db_connection(conn)
                return jsonify({'error': 'Operación no encontrada'}), 404
            
            # Actualizar el estado del ingreso de vuelta a "Playa Lavadero"
            if ingreso_id:
                cursor.execute("""
                    UPDATE ingresos_usados
                    SET estado = 'Playa Lavadero'
                    WHERE id = %s
                """, (ingreso_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'message': 'Turno dado de baja exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500


@app.route('/api/usados/programar-automatico', methods=['POST'])
@login_required
@module_permission_required('usados')
def programar_automatico():
    """Algoritmo inteligente para programar lavados en huecos disponibles"""
    conn = None
    try:
        data = request.get_json()
        fecha_inicial = data['fecha']
        orden_manual = data.get('orden_vehiculos', [])
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # 1. Obtener vehículos en cola (FIFO) - SOLO clasificación USADOS
        if orden_manual:
            # Usar orden manual (FIFO con prioridad usuario)
            cursor.execute("""
                SELECT id, limpieza_requerida, dominio
                FROM ingresos_usados
                WHERE id = ANY(%s) 
                AND estado = 'Playa Lavadero' 
                AND activo = TRUE
                AND clasificacion = 'USADOS'
            """, (orden_manual,))
            vehiculos_dict = {v['id']: v for v in cursor.fetchall()}
            cola = [vehiculos_dict[vid] for vid in orden_manual if vid in vehiculos_dict]
        else:
            # Orden por defecto (FIFO por fecha_ingreso)
            cursor.execute("""
                SELECT id, limpieza_requerida, dominio
                FROM ingresos_usados
                WHERE estado = 'Playa Lavadero' 
                AND activo = TRUE
                AND clasificacion = 'USADOS'
                ORDER BY fecha_ingreso ASC
            """)
            cola = cursor.fetchall()
        
        if not cola:
            release_db_connection(conn)
            return jsonify({'error': 'No hay vehículos en cola'}), 400
        
        # 2. Obtener duraciones de operaciones
        cursor.execute("""
            SELECT nombre, duracion_minutos
            FROM config_operaciones_lavado
            WHERE activo = TRUE
        """)
        duraciones = {row['nombre']: row['duracion_minutos'] for row in cursor.fetchall()}
        
        # Mapeo de limpieza requerida a operación
        mapeo_operaciones = {
            'Standard': 'Standard',
            'Standard Plus': 'Standard Plus',
            'Standard Light': 'Standard Light',
            'Entrega': 'Entrega',
            'Repaso': 'Repaso',
            'Kinto': 'Standard Plus'  # Kinto usa Standard Plus (50 min)
        }
        
        # 3. Función auxiliar para obtener horarios laborales del día
        def obtener_horarios_dia(fecha):
            # Verificar día de la semana
            from datetime import datetime
            fecha_obj = datetime.strptime(str(fecha), '%Y-%m-%d')
            dia_semana = fecha_obj.weekday()  # 0=Lunes, 5=Sábado, 6=Domingo
            
            # Los sábados tienen horario especial: 9:00-13:00
            if dia_semana == 5:
                return [(9 * 60, 13 * 60)]  # 9:00 a 13:00
            
            # Verificar si hay jornada completa configurada
            cursor.execute("""
                SELECT tipo_jornada FROM config_jornadas_lavadero
                WHERE fecha = %s
            """, (fecha,))
            
            jornada = cursor.fetchone()
            
            if jornada and jornada['tipo_jornada'] == 'completa':
                # Jornada completa: 8:30 a 20:00
                return [(8 * 60 + 30, 20 * 60)]
            else:
                # Por defecto: 8:30-13:00 y 16:00-20:00 (jornada cortada)
                return [(8 * 60 + 30, 13 * 60), (16 * 60, 20 * 60)]
        
        # 4. Función para obtener operaciones ya programadas (Usados + Kinto)
        def obtener_operaciones_programadas(fecha):
            # Obtener Usados
            cursor.execute("""
                SELECT po.hora_inicio::TEXT, po.hora_fin::TEXT
                FROM planificacion_operaciones po
                WHERE po.fecha_planificada::DATE = %s
            """, (fecha,))
            usados = cursor.fetchall()
            
            # Obtener Kinto
            cursor.execute("""
                SELECT hora_inicio::TEXT, hora_fin::TEXT
                FROM reservas_kinto
                WHERE fecha_reserva::DATE = %s AND estado != 'Cancelado'
            """, (fecha,))
            kinto = cursor.fetchall()
            
            # Convertir todas a minutos
            operaciones = []
            for op in list(usados) + list(kinto):
                h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                operaciones.append((h_ini * 60 + m_ini, h_fin * 60 + m_fin))
            
            # Ordenar por hora de inicio
            operaciones.sort()
            return operaciones
        
        # 5. Función para encontrar huecos disponibles en un día
        def encontrar_huecos(fecha, duracion_necesaria):
            horarios = obtener_horarios_dia(fecha)
            if not horarios:
                print(f"⚠️ DEBUG: No hay horarios definidos para {fecha}")
                return []
            
            operaciones = obtener_operaciones_programadas(fecha)
            huecos = []
            
            # Verificar si la fecha es hoy y obtener minutos actuales
            from datetime import datetime
            ahora = datetime.now()
            fecha_obj = datetime.strptime(str(fecha), '%Y-%m-%d')
            es_hoy = fecha_obj.date() == ahora.date()
            minutos_actuales = ahora.hour * 60 + ahora.minute if es_hoy else 0
            
            print(f"🔍 DEBUG: Buscando huecos en {fecha} para {duracion_necesaria} min")
            print(f"   Es hoy: {es_hoy}, Minutos actuales: {minutos_actuales} ({minutos_actuales//60:02d}:{minutos_actuales%60:02d})")
            print(f"   Horarios del día: {horarios}")
            print(f"   Operaciones programadas: {len(operaciones)}")
            
            for bloque_inicio, bloque_fin in horarios:
                tiempo_disponible = bloque_inicio
                
                print(f"   Bloque: {bloque_inicio//60:02d}:{bloque_inicio%60:02d} - {bloque_fin//60:02d}:{bloque_fin%60:02d}")
                
                # Si es hoy y el bloque inicia antes de la hora actual, ajustar
                if es_hoy and tiempo_disponible < minutos_actuales:
                    tiempo_disponible = minutos_actuales
                    print(f"   → Ajustado inicio a hora actual: {tiempo_disponible//60:02d}:{tiempo_disponible%60:02d}")
                
                # Si el bloque ya pasó completamente, saltarlo
                if es_hoy and bloque_fin <= minutos_actuales:
                    print(f"   → Bloque ya pasó, saltando")
                    continue
                
                # Revisar cada operación en este bloque
                for op_inicio, op_fin in operaciones:
                    # Si la operación está dentro del bloque
                    if op_fin <= bloque_inicio or op_inicio >= bloque_fin:
                        continue  # No afecta este bloque
                    
                    # Hay hueco antes de esta operación?
                    if tiempo_disponible + duracion_necesaria <= op_inicio:
                        huecos.append(tiempo_disponible)
                    
                    # Actualizar tiempo disponible
                    tiempo_disponible = max(tiempo_disponible, op_fin)
                
                # Revisar si queda espacio al final del bloque
                if tiempo_disponible + duracion_necesaria <= bloque_fin:
                    huecos.append(tiempo_disponible)
                    print(f"   ✅ Hueco encontrado: {tiempo_disponible//60:02d}:{tiempo_disponible%60:02d}")
                else:
                    print(f"   ❌ No cabe: necesita {duracion_necesaria} min, disponible: {bloque_fin - tiempo_disponible} min")
            
            print(f"   Total huecos encontrados: {len(huecos)}")
            return huecos
        
        # 6. Programar cada vehículo de la cola
        from datetime import datetime, timedelta
        programados = 0
        fecha_actual = fecha_inicial
        max_dias_busqueda = 30  # Buscar hasta 30 días adelante
        
        for vehiculo in cola:
            limpieza = vehiculo['limpieza_requerida']
            operacion_nombre = mapeo_operaciones.get(limpieza, 'Standard')
            duracion = duraciones.get(operacion_nombre, 60)
            
            # Obtener ID de operación
            cursor.execute("SELECT id FROM config_operaciones_lavado WHERE nombre = %s", (operacion_nombre,))
            operacion_row = cursor.fetchone()
            if not operacion_row:
                cursor.execute("SELECT id FROM config_operaciones_lavado WHERE nombre = 'Standard'")
                operacion_row = cursor.fetchone()
            operacion_id = operacion_row['id']
            
            # Buscar primer hueco disponible
            programado = False
            fecha_busqueda = fecha_actual
            
            for dia in range(max_dias_busqueda):
                huecos = encontrar_huecos(fecha_busqueda, duracion)
                
                if huecos:
                    # Tomar el primer hueco disponible
                    minuto_inicio = huecos[0]
                    minuto_fin = minuto_inicio + duracion
                    
                    hora_inicio_op = f"{minuto_inicio // 60:02d}:{minuto_inicio % 60:02d}:00"
                    hora_fin_op = f"{minuto_fin // 60:02d}:{minuto_fin % 60:02d}:00"
                    
                    # Insertar en planificación
                    cursor.execute("""
                        INSERT INTO planificacion_operaciones
                        (ingreso_id, fecha_planificada, hora_inicio, hora_fin, posicion_lavadero, operacion_lavado_id, orden_ejecucion)
                        VALUES (%s, %s, %s, %s, %s, %s, %s)
                    """, (vehiculo['id'], fecha_busqueda, hora_inicio_op, hora_fin_op, 1, operacion_id, programados + 1))
                    
                    # Cambiar estado del vehículo
                    cursor.execute("""
                        UPDATE ingresos_usados
                        SET estado = 'Lavado Programado'
                        WHERE id = %s
                    """, (vehiculo['id'],))
                    
                    # Registrar en historial
                    cursor.execute("""
                        INSERT INTO historial_estados_usados
                        (ingreso_id, estado_anterior, estado_nuevo, usuario)
                        VALUES (%s, 'Playa Lavadero', 'Lavado Programado', %s)
                    """, (vehiculo['id'], current_user.username))
                    
                    programados += 1
                    programado = True
                    print(f"✅ {vehiculo['dominio']} programado para {fecha_busqueda} a las {hora_inicio_op}")
                    break
                
                # Siguiente día
                fecha_obj = datetime.strptime(fecha_busqueda, '%Y-%m-%d')
                fecha_obj += timedelta(days=1)
                fecha_busqueda = fecha_obj.strftime('%Y-%m-%d')
            
            if not programado:
                print(f"⚠️ No se pudo programar {vehiculo['dominio']} (sin espacio en {max_dias_busqueda} días)")
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True, 
            'programados': programados,
            'total_cola': len(cola),
            'mensaje': f'Se programaron {programados} de {len(cola)} vehículos'
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en programar_automatico: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/optimizar-gantt', methods=['POST'])
@login_required
@module_permission_required('usados')
def optimizar_gantt():
    """Elimina tiempos muertos entre operaciones del Gantt"""
    conn = None
    try:
        data = request.get_json()
        fecha = data['fecha']
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # 1. Obtener SOLO operaciones de Usados del día ordenadas por hora_inicio
        # Las operaciones de Kinto NO deben ser optimizadas
        cursor.execute("""
            SELECT po.id, po.hora_inicio::TEXT, po.hora_fin::TEXT,
                   col.duracion_minutos, i.estado
            FROM planificacion_operaciones po
            JOIN config_operaciones_lavado col ON po.operacion_lavado_id = col.id
            JOIN ingresos_usados i ON po.ingreso_id = i.id
            WHERE po.fecha_planificada::DATE = %s
            ORDER BY po.hora_inicio ASC
        """, (fecha,))
        operaciones = cursor.fetchall()
        
        if not operaciones:
            release_db_connection(conn)
            return jsonify({'error': 'No hay operaciones de Usados para optimizar'}), 400
        
        # 2. Obtener turno del día para saber hora de inicio
        from datetime import datetime
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana_num = fecha_obj.weekday()
        
        cursor.execute("""
            SELECT hora_inicio::TEXT as hora_inicio
            FROM turnos_lavadero
            WHERE dia_semana = %s AND activo = TRUE
            LIMIT 1
        """, (dia_semana_num,))
        turno = cursor.fetchone()
        
        if not turno:
            release_db_connection(conn)
            return jsonify({'error': 'No hay turno configurado'}), 400
        
        # 3. Obtener configuración de jornada para esta fecha
        cursor.execute("""
            SELECT tipo_jornada
            FROM config_jornadas_lavadero
            WHERE fecha = %s
        """, (fecha,))
        config_jornada = cursor.fetchone()
        
        # Determinar horarios según tipo de jornada
        if config_jornada and config_jornada['tipo_jornada'] == 'corrida':
            # Jornada corrida: 08:30 - 17:30
            minuto_inicio = 8 * 60 + 30
            minuto_fin_jornada = 17 * 60 + 30
            tiene_descanso = False
        else:
            # Jornada cortada (por defecto): 08:30 - 13:00 y 16:00 - 20:00
            minuto_inicio = 8 * 60 + 30
            minuto_fin_manana = 13 * 60
            minuto_inicio_tarde = 16 * 60
            minuto_fin_jornada = 20 * 60
            tiene_descanso = True
        
        # 4. Reorganizar operaciones eliminando tiempos muertos
        # Obtener hora actual para no mover operaciones que ya pasaron o están en curso
        from datetime import datetime
        ahora = datetime.now()
        # Solo considerar hora actual si estamos en la misma fecha
        hoy = datetime.now().date()
        fecha_operacion = fecha_obj.date()
        es_hoy = (hoy == fecha_operacion)
        
        if es_hoy:
            minuto_actual = ahora.hour * 60 + ahora.minute
        else:
            minuto_actual = -1  # No filtrar por hora si es fecha futura
        
        tiempo_actual = minuto_inicio
        operaciones_movidas = 0
        
        print(f"[OPTIMIZAR] Procesando {len(operaciones)} operaciones")
        print(f"[OPTIMIZAR] Fecha: {fecha}, Es hoy: {es_hoy}, Tiene descanso: {tiene_descanso}")
        
        for operacion in operaciones:
            duracion = operacion['duracion_minutos']
            estado = operacion.get('estado', 'Reservado')
            
            # Parsear hora_inicio de la operación
            hora_op_str = operacion['hora_inicio']
            h_op, m_op = map(int, hora_op_str.split(':')[:2])
            minuto_inicio_op = h_op * 60 + m_op
            
            print(f"[OPTIMIZAR] Op {operacion['id']}: {hora_op_str} ({duracion}min) Estado: {estado}")
            
            # Si la operación está completada (Salón o Completado), no moverla - queda fija
            if estado in ['Salón', 'Completado']:
                hora_fin_str = operacion['hora_fin']
                h_fin, m_fin = map(int, hora_fin_str.split(':')[:2])
                tiempo_actual = h_fin * 60 + m_fin
                print(f"[OPTIMIZAR]   ⏩ OMITIR: Completada")
                continue
            
            # Si la operación ya terminó o está en curso (solo para hoy), mantener su horario original
            if es_hoy and minuto_inicio_op <= minuto_actual:
                # Esta operación ya pasó o está en ejecución, no moverla
                hora_fin_str = operacion['hora_fin']
                h_fin, m_fin = map(int, hora_fin_str.split(':')[:2])
                tiempo_actual = h_fin * 60 + m_fin
                print(f"[OPTIMIZAR]   ⏩ OMITIR: Ya pasó o en curso")
                continue
            
            # Verificar si cabe en horario laboral
            if tiene_descanso:
                # Jornada cortada: verificar si cabe en turno actual o debe saltar al siguiente
                hora_fin_prevista = tiempo_actual + duracion
                
                # Si estamos en la mañana y la operación terminaría después de las 13:00
                if tiempo_actual < minuto_fin_manana and hora_fin_prevista > minuto_fin_manana:
                    # Mover al inicio de la tarde
                    tiempo_actual = minuto_inicio_tarde
                    hora_fin_prevista = tiempo_actual + duracion
            else:
                # Jornada corrida: verificar que no se pase de las 17:30
                hora_fin_prevista = tiempo_actual + duracion
                if hora_fin_prevista > minuto_fin_jornada:
                    # No cabe más operaciones hoy
                    break
            
            # Nueva hora inicio = tiempo_actual
            hora_inicio_nueva = f"{tiempo_actual // 60:02d}:{tiempo_actual % 60:02d}:00"
            tiempo_actual += duracion
            hora_fin_nueva = f"{tiempo_actual // 60:02d}:{tiempo_actual % 60:02d}:00"
            
            print(f"[OPTIMIZAR]   Nueva: {hora_inicio_nueva} - {hora_fin_nueva}")
            
            # Solo actualizar si realmente cambió
            if hora_inicio_nueva != operacion['hora_inicio']:
                print(f"[OPTIMIZAR]   ✅ ACTUALIZANDO (era {operacion['hora_inicio']})")
                cursor.execute("""
                    UPDATE planificacion_operaciones
                    SET hora_inicio = %s, hora_fin = %s
                    WHERE id = %s
                """, (hora_inicio_nueva, hora_fin_nueva, operacion['id']))
                operaciones_movidas += 1
            else:
                print(f"[OPTIMIZAR]   ⏸️ NO CAMBIA (ya está en {hora_inicio_nueva})")
        
        print(f"[OPTIMIZAR] COMMIT: {operaciones_movidas} operaciones movidas")
        conn.commit()
        release_db_connection(conn)
        
        if operaciones_movidas == 0:
            return jsonify({
                'success': True,
                'operaciones': 0,
                'mensaje': 'No hay operaciones para optimizar (todas ya están optimizadas o en curso)'
            }), 200
        
        return jsonify({
            'success': True,
            'operaciones': operaciones_movidas,
            'mensaje': f'Gantt optimizado: {operaciones_movidas} operaciones sin tiempos muertos'
        }), 200
    
    except Exception as e:
        print(f"Error en optimizar_gantt: {e}")
        if conn:
            conn.rollback()
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

# --- APIs Panel Lavadero ---
@app.route('/api/usados/lavadero/<fecha>', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_operaciones_lavadero(fecha):
    """Obtener operaciones programadas del día para panel de operarios (Usados + Kinto)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Operaciones Usados (incluye tareas cíclicas y operaciones normales NO completadas)
        cursor.execute("""
            SELECT 
                p.id,
                p.fecha_planificada::DATE as fecha,
                p.hora_inicio::TEXT as hora_inicio,
                p.hora_fin::TEXT as hora_fin,
                p.posicion_lavadero,
                p.orden_ejecucion,
                co.duracion_minutos,
                COALESCE(i.dominio, co.nombre) as dominio,
                i.observaciones,
                CASE 
                    WHEN p.es_tarea_ciclica = TRUE THEN 'TAREA_CICLICA'
                    ELSE COALESCE(i.clasificacion, 'USADOS')
                END as clasificacion,
                m.nombre as marca_nombre,
                mo.nombre as modelo_nombre,
                co.nombre as operacion_nombre,
                CASE 
                    WHEN p.es_tarea_ciclica = TRUE THEN 'ciclica'
                    ELSE 'usados'
                END as tipo,
                COALESCE(p.es_tarea_ciclica, FALSE) as es_tarea_ciclica,
                co.nombre as nombre_tarea_ciclica
            FROM planificacion_operaciones p
            LEFT JOIN ingresos_usados i ON p.ingreso_id = i.id
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            LEFT JOIN config_operaciones_lavado co ON p.operacion_lavado_id = co.id
            WHERE p.fecha_planificada::DATE = %s
            AND p.tiempo_fin_real IS NULL
            AND (p.es_tarea_ciclica = TRUE OR i.estado = 'Lavado Programado' OR i.estado IS NULL)
            ORDER BY p.orden_ejecucion ASC
        """, (fecha,))
        operaciones_usados = cursor.fetchall() or []
        
        # Reservas Kinto
        cursor.execute("""
            SELECT 
                k.id,
                k.fecha_reserva::DATE as fecha,
                k.hora_inicio::TEXT as hora_inicio,
                k.hora_fin::TEXT as hora_fin,
                1 as posicion_lavadero,
                ROW_NUMBER() OVER (ORDER BY k.hora_inicio) as orden_ejecucion,
                k.duracion_minutos,
                k.dominio,
                k.observaciones,
                'KINTO' as clasificacion,
                'KINTO' as marca_nombre,
                'Alquiler' as modelo_nombre,
                'Standard Plus' as operacion_nombre,
                'kinto' as tipo
            FROM reservas_kinto k
            WHERE k.fecha_reserva::DATE = %s
            AND k.estado = 'Reservado'
            ORDER BY k.hora_inicio ASC
        """, (fecha,))
        reservas_kinto = cursor.fetchall() or []
        
        # Combinar y ordenar por hora
        todas = list(operaciones_usados) + list(reservas_kinto)
        todas.sort(key=lambda x: x['hora_inicio'] if x['hora_inicio'] else '')
        
        release_db_connection(conn)
        return jsonify([dict(o) for o in todas])
    except Exception as e:
        print(f"Error en get_operaciones_lavadero: {str(e)}")
        import traceback
        traceback.print_exc()
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/completar-lavado/<int:operacion_id>', methods=['POST'])
@login_required
@module_permission_required('usados')
def completar_lavado(operacion_id):
    """Completar lavado y mover vehículo a Salón (Usados o Kinto)"""
    conn = None
    try:
        data = request.get_json() or {}
        tipo = data.get('tipo', 'usados')
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        if tipo == 'kinto':
            # Actualizar estado de reserva Kinto a Completado
            cursor.execute("""
                UPDATE reservas_kinto
                SET estado = 'Completado'
                WHERE id = %s
                RETURNING dominio, fecha_reserva, hora_fin
            """, (operacion_id,))
            
            reserva = cursor.fetchone()
            if not reserva:
                release_db_connection(conn)
                return jsonify({'error': 'Reserva Kinto no encontrada'}), 404
            
            # TAMBIÉN marcar como completada en planificacion_operaciones
            cursor.execute("""
                UPDATE planificacion_operaciones p
                SET tiempo_fin_real = CURRENT_TIMESTAMP
                FROM ingresos_usados i
                WHERE p.ingreso_id = i.id
                AND i.dominio = %s
                AND i.clasificacion = 'KINTO'
                AND p.fecha_planificada::DATE = %s
                AND p.tiempo_fin_real IS NULL
            """, (reserva['dominio'], reserva['fecha_reserva']))
            
            conn.commit()
            release_db_connection(conn)
            return jsonify({
                'success': True,
                'mensaje': f'Lavado Kinto completado para {reserva["dominio"]}'
            }), 200
        
        # Operación Usados
        cursor.execute("""
            SELECT ingreso_id
            FROM planificacion_operaciones
            WHERE id = %s
        """, (operacion_id,))
        
        planificacion = cursor.fetchone()
        if not planificacion:
            release_db_connection(conn)
            return jsonify({'error': 'Operación no encontrada'}), 404
        
        ingreso_id = planificacion['ingreso_id']
        
        # Cambiar estado del vehículo a "Salón"
        cursor.execute("""
            UPDATE ingresos_usados
            SET estado = 'Salón'
            WHERE id = %s
        """, (ingreso_id,))
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, 'Lavado Programado', 'Salón', %s)
        """, (ingreso_id, current_user.username))
        
        # Registrar tiempo real de finalización en planificación
        cursor.execute("""
            UPDATE planificacion_operaciones
            SET tiempo_fin_real = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (operacion_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en completar_lavado: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/control-salon', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_control_salon():
    """Obtener vehículos en Salón con información de último lavado"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener vehículos en estado Salón con información del último lavado
        # SOLO USADOS (KINTO y TEST DRIVE se gestionan desde Stock)
        # EXCLUIR vehículos en Reparación
        cursor.execute("""
            SELECT 
                i.id,
                i.dominio,
                m.nombre as marca,
                mo.nombre as modelo,
                i.fecha_ingreso,
                MAX(po.tiempo_fin_real) as ultimo_lavado,
                COALESCE(DATE_PART('day', NOW() - MAX(po.tiempo_fin_real)), 0) as dias_en_salon
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            LEFT JOIN planificacion_operaciones po ON i.id = po.ingreso_id
            WHERE i.estado = 'Salón' AND i.activo = TRUE AND i.clasificacion = 'USADOS'
            GROUP BY i.id, i.dominio, m.nombre, mo.nombre, i.fecha_ingreso
            ORDER BY ultimo_lavado DESC NULLS LAST
        """)
        
        vehiculos = cursor.fetchall()
        
        release_db_connection(conn)
        
        return jsonify([dict(v) for v in vehiculos]), 200
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_control_salon: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/programar-desde-salon', methods=['POST'])
@login_required
@module_permission_required('usados')
def programar_desde_salon():
    """Programar operación (Entrega o Repaso) desde Control Salón buscando primer hueco disponible"""
    from datetime import datetime, timedelta
    conn = None
    try:
        data = request.get_json()
        ingreso_id = data['ingreso_id']
        tipo_operacion = data['tipo_operacion']  # 'Entrega' o 'Repaso'
        fecha_inicial = data.get('fecha_inicial', datetime.now().strftime('%Y-%m-%d'))
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener ID de la operación de lavado
        cursor.execute("""
            SELECT id, duracion_minutos 
            FROM config_operaciones_lavado 
            WHERE nombre = %s
        """, (tipo_operacion,))
        operacion = cursor.fetchone()
        
        if not operacion:
            release_db_connection(conn)
            return jsonify({'error': f'Operación "{tipo_operacion}" no encontrada'}), 404
        
        operacion_id = operacion['id']
        duracion = operacion['duracion_minutos']
        
        # Función auxiliar para obtener horarios laborales del día
        def obtener_horarios_dia(fecha):
            # Horario fijo: 8:30 a 20:00 todos los días
            return [(8 * 60 + 30, 20 * 60)]
        
        # Función para obtener operaciones programadas
        def obtener_operaciones_programadas(fecha):
            cursor.execute("""
                SELECT po.hora_inicio::TEXT, po.hora_fin::TEXT
                FROM planificacion_operaciones po
                WHERE po.fecha_planificada::DATE = %s
            """, (fecha,))
            usados = cursor.fetchall()
            
            cursor.execute("""
                SELECT hora_inicio::TEXT, hora_fin::TEXT
                FROM reservas_kinto
                WHERE fecha_reserva::DATE = %s AND estado != 'Cancelado'
            """, (fecha,))
            kinto = cursor.fetchall()
            
            operaciones = []
            for op in list(usados) + list(kinto):
                h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                operaciones.append((h_ini * 60 + m_ini, h_fin * 60 + m_fin))
            
            operaciones.sort()
            return operaciones
        
        # Buscar primer hueco disponible
        def encontrar_primer_hueco(fecha, duracion_necesaria):
            horarios = obtener_horarios_dia(fecha)
            if not horarios:
                return None
            
            operaciones = obtener_operaciones_programadas(fecha)
            
            for bloque_inicio, bloque_fin in horarios:
                tiempo_disponible = bloque_inicio
                
                for op_inicio, op_fin in operaciones:
                    if op_fin <= bloque_inicio or op_inicio >= bloque_fin:
                        continue
                    
                    if tiempo_disponible + duracion_necesaria <= op_inicio:
                        return tiempo_disponible
                    
                    tiempo_disponible = max(tiempo_disponible, op_fin)
                
                if tiempo_disponible + duracion_necesaria <= bloque_fin:
                    return tiempo_disponible
            
            return None
        
        # Buscar fecha con espacio disponible
        fecha_busqueda = fecha_inicial
        fecha_programada = None
        minuto_inicio = None
        max_dias = 30
        
        for dia in range(max_dias):
            minuto = encontrar_primer_hueco(fecha_busqueda, duracion)
            if minuto is not None:
                fecha_programada = fecha_busqueda
                minuto_inicio = minuto
                break
            
            fecha_obj = datetime.strptime(fecha_busqueda, '%Y-%m-%d')
            fecha_obj += timedelta(days=1)
            fecha_busqueda = fecha_obj.strftime('%Y-%m-%d')
        
        if not fecha_programada:
            release_db_connection(conn)
            return jsonify({'error': f'No hay espacio disponible en los próximos {max_dias} días'}), 400
        
        # Crear programación
        minuto_fin = minuto_inicio + duracion
        hora_inicio = f"{minuto_inicio // 60:02d}:{minuto_inicio % 60:02d}:00"
        hora_fin = f"{minuto_fin // 60:02d}:{minuto_fin % 60:02d}:00"
        
        cursor.execute("""
            INSERT INTO planificacion_operaciones
            (ingreso_id, fecha_planificada, hora_inicio, hora_fin, posicion_lavadero, operacion_lavado_id, orden_ejecucion)
            VALUES (%s, %s, %s, %s, %s, %s, 
                (SELECT COALESCE(MAX(orden_ejecucion), 0) + 1 FROM planificacion_operaciones WHERE fecha_planificada::DATE = %s)
            )
            RETURNING id
        """, (ingreso_id, fecha_programada, hora_inicio, hora_fin, 1, operacion_id, fecha_programada))
        
        nueva_operacion = cursor.fetchone()
        
        # NO actualizar el estado del vehículo si viene del Salón
        # La nueva operación (Entrega o Repaso) se programa pero el vehículo permanece en Salón
        # hasta que se complete esta nueva operación específica
        # Esto evita que se reactiven operaciones de lavado antiguas que ya estaban completadas
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'fecha_programada': fecha_programada,
            'hora_inicio': hora_inicio,
            'hora_fin': hora_fin,
            'operacion_id': nueva_operacion['id']
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en programar_desde_salon: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/tarea-ciclica', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_tarea_ciclica():
    """Crear tareas cíclicas recurrentes (desayunos, descansos, etc.) que bloquean espacios en el gantt"""
    from datetime import datetime, timedelta
    conn = None
    try:
        data = request.get_json()
        operacion_id = data['operacion_id']
        fecha_inicio = data['fecha_inicio']
        fecha_fin = data['fecha_fin']
        hora_inicio = data['hora_inicio']
        hora_fin = data['hora_fin']
        frecuencia = data['frecuencia']  # {tipo: 'diaria'|'laborable'|'sabados'|'personalizada', dias: [1,2,3,4,5]}
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Verificar que la operación existe
        cursor.execute("""
            SELECT nombre, duracion_minutos 
            FROM config_operaciones_lavado 
            WHERE id = %s
        """, (operacion_id,))
        operacion = cursor.fetchone()
        
        if not operacion:
            release_db_connection(conn)
            return jsonify({'error': 'Operación no encontrada'}), 404
        
        # Generar lista de fechas según frecuencia
        fecha_inicio_dt = datetime.strptime(fecha_inicio, '%Y-%m-%d')
        fecha_fin_dt = datetime.strptime(fecha_fin, '%Y-%m-%d')
        fechas_a_crear = []
        
        tipo_frecuencia = frecuencia.get('tipo', 'laborable')
        
        fecha_actual = fecha_inicio_dt
        while fecha_actual <= fecha_fin_dt:
            incluir_fecha = False
            dia_semana = fecha_actual.weekday()  # 0=Lun, 1=Mar, 2=Mié, 3=Jue, 4=Vie, 5=Sáb, 6=Dom
            
            if tipo_frecuencia == 'laborable':
                # Lunes a Viernes
                incluir_fecha = dia_semana < 5
            elif tipo_frecuencia == 'lun_mie_vie':
                # Lunes (0), Miércoles (2), Viernes (4)
                incluir_fecha = dia_semana in [0, 2, 4]
            elif tipo_frecuencia == 'mar_jue_sab':
                # Martes (1), Jueves (3), Sábado (5)
                incluir_fecha = dia_semana in [1, 3, 5]
            elif tipo_frecuencia == 'sabados':
                # Solo sábados
                incluir_fecha = dia_semana == 5
            
            if incluir_fecha:
                fechas_a_crear.append(fecha_actual.strftime('%Y-%m-%d'))
            
            fecha_actual += timedelta(days=1)
        
        # Crear tareas para cada fecha
        tareas_creadas = 0
        tareas_con_conflicto = 0
        
        for fecha in fechas_a_crear:
            # Verificar si hay conflictos de horario en esta fecha con otras operaciones
            cursor.execute("""
                SELECT COUNT(*) as conflictos
                FROM planificacion_operaciones
                WHERE fecha_planificada::DATE = %s
                AND (
                    (hora_inicio < %s AND hora_fin > %s) OR
                    (hora_inicio < %s AND hora_fin > %s) OR
                    (hora_inicio >= %s AND hora_fin <= %s)
                )
            """, (fecha, hora_fin, hora_inicio, hora_fin, hora_fin, hora_inicio, hora_fin))
            
            conflicto = cursor.fetchone()
            if conflicto and conflicto['conflictos'] > 0:
                tareas_con_conflicto += 1
                continue  # Saltar esta fecha si hay conflicto
            
            # Crear la tarea cíclica (sin vehículo, es_tarea_ciclica=TRUE)
            cursor.execute("""
                INSERT INTO planificacion_operaciones
                (ingreso_id, fecha_planificada, hora_inicio, hora_fin, posicion_lavadero, operacion_lavado_id, orden_ejecucion, es_tarea_ciclica, fecha_creacion)
                VALUES (NULL, %s, %s, %s, 1, %s, 
                    (SELECT COALESCE(MAX(orden_ejecucion), 0) + 1 FROM planificacion_operaciones WHERE fecha_planificada::DATE = %s),
                    TRUE, CURRENT_TIMESTAMP
                )
            """, (fecha, hora_inicio, hora_fin, operacion_id, fecha))
            
            tareas_creadas += 1
        
        conn.commit()
        release_db_connection(conn)
        
        mensaje = f'{tareas_creadas} tarea(s) cíclica(s) creada(s) exitosamente'
        if tareas_con_conflicto > 0:
            mensaje += f'. {tareas_con_conflicto} fecha(s) omitida(s) por conflictos de horario'
        
        return jsonify({
            'success': True,
            'tareas_creadas': tareas_creadas,
            'tareas_con_conflicto': tareas_con_conflicto,
            'mensaje': mensaje
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en crear_tarea_ciclica: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/enviar-a-cola-desde-salon', methods=['POST'])
@login_required
@module_permission_required('usados')
def enviar_a_cola_desde_salon():
    """Enviar vehículo desde Salón a Cola con operación específica (Entrega o Repaso)"""
    conn = None
    try:
        data = request.get_json()
        ingreso_id = data['ingreso_id']
        tipo_operacion = data['tipo_operacion']  # 'Entrega' o 'Repaso'
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Validar que Entrega y Repaso SOLO se apliquen a clasificación='USADOS'
        cursor.execute("""
            SELECT clasificacion FROM ingresos_usados WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        if tipo_operacion in ['Entrega', 'Repaso'] and vehiculo['clasificacion'] != 'USADOS':
            release_db_connection(conn)
            return jsonify({'error': f'{tipo_operacion} solo se puede aplicar a vehículos USADOS, no a {vehiculo["clasificacion"]}'}), 400
        
        # Actualizar estado y limpieza_requerida del vehículo
        cursor.execute("""
            UPDATE ingresos_usados
            SET estado = 'Playa Lavadero',
                limpieza_requerida = %s
            WHERE id = %s
        """, (tipo_operacion, ingreso_id))
        
        # Registrar en historial
        cursor.execute("""
            INSERT INTO historial_estados_usados
            (ingreso_id, estado_anterior, estado_nuevo, usuario)
            VALUES (%s, 'Salón', 'Playa Lavadero', %s)
        """, (ingreso_id, current_user.username))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'mensaje': f'Vehículo enviado a cola para {tipo_operacion}'
        }), 200
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en enviar_a_cola_desde_salon: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/stock/<int:ingreso_id>/historial-lavados', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_historial_lavados_vehiculo(ingreso_id):
    """Obtener historial de lavados de un vehículo específico (completados y programados)"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener información del vehículo
        cursor.execute("""
            SELECT dominio, clasificacion 
            FROM ingresos_usados 
            WHERE id = %s
        """, (ingreso_id,))
        vehiculo = cursor.fetchone()
        
        if not vehiculo:
            release_db_connection(conn)
            return jsonify({'error': 'Vehículo no encontrado'}), 404
        
        dominio = vehiculo['dominio']
        clasificacion = vehiculo['clasificacion']
        
        # Obtener lavados completados y programados
        cursor.execute("""
            SELECT 
                co.nombre as operacion,
                TO_CHAR(p.fecha_planificada, 'YYYY-MM-DD') as fecha_programada,
                TO_CHAR(p.hora_inicio, 'HH24:MI') as hora_inicio,
                TO_CHAR(p.hora_fin, 'HH24:MI') as hora_fin,
                TO_CHAR(p.tiempo_fin_real AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_completado,
                CASE 
                    WHEN p.tiempo_fin_real IS NOT NULL THEN 'Completado'
                    WHEN p.fecha_planificada::DATE = CURRENT_DATE THEN 'En Curso'
                    WHEN p.fecha_planificada::DATE < CURRENT_DATE THEN 'Vencido'
                    ELSE 'Programado'
                END as estado
            FROM planificacion_operaciones p
            LEFT JOIN config_operaciones_lavado co ON p.operacion_lavado_id = co.id
            WHERE p.ingreso_id = %s
            ORDER BY 
                CASE 
                    WHEN p.tiempo_fin_real IS NOT NULL THEN 1
                    ELSE 0
                END,
                p.fecha_planificada DESC, 
                p.hora_inicio DESC
        """, (ingreso_id,))
        
        lavados = cursor.fetchall()
        
        # Si es KINTO, también obtener reservas
        reservas_kinto = []
        if clasificacion == 'KINTO':
            cursor.execute("""
                SELECT 
                    'Standard Plus' as operacion,
                    TO_CHAR(fecha_reserva, 'YYYY-MM-DD') as fecha_programada,
                    TO_CHAR(hora_inicio, 'HH24:MI') as hora_inicio,
                    TO_CHAR(hora_fin, 'HH24:MI') as hora_fin,
                    CASE WHEN estado = 'Completado' 
                        THEN TO_CHAR(fecha_reserva + hora_fin AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS')
                        ELSE NULL 
                    END as fecha_completado,
                    CASE 
                        WHEN estado = 'Completado' THEN 'Completado'
                        WHEN fecha_reserva::DATE = CURRENT_DATE THEN 'En Curso'
                        WHEN fecha_reserva::DATE < CURRENT_DATE THEN 'Vencido'
                        ELSE 'Programado'
                    END as estado
                FROM reservas_kinto
                WHERE dominio = %s AND estado != 'Cancelado'
                ORDER BY fecha_reserva DESC, hora_inicio DESC
            """, (dominio,))
            reservas_kinto = cursor.fetchall()
        
        release_db_connection(conn)
        
        # Combinar y ordenar
        todos_lavados = [dict(l) for l in lavados] + [dict(r) for r in reservas_kinto]
        todos_lavados.sort(key=lambda x: (
            0 if x['estado'] == 'Completado' else 1,
            x['fecha_completado'] if x['fecha_completado'] else x['fecha_programada']
        ), reverse=True)
        
        return jsonify(todos_lavados), 200
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        return jsonify({'error': str(e)}), 500


@app.route('/api/usados/registro-lavados', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_registro_lavados():
    """Obtener registro histórico de todas las operaciones de lavado por vehículo"""
    conn = None
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        dominio_filtro = request.args.get('dominio')
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Query base para obtener TODAS las operaciones + vehículos en cola + reservas KINTO
        # UNION de tres consultas:
        # 1. Operaciones ya programadas (de planificacion_operaciones)
        # 2. Vehículos en cola sin programar aún (de ingresos_usados en estado 'Playa Lavadero')
        # 3. Reservas KINTO (de reservas_kinto)
        query = """
            SELECT 
                i.dominio,
                i.clasificacion,
                m.nombre as marca,
                mo.nombre as modelo,
                TO_CHAR(i.fecha_ingreso AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_alta_sistema,
                co.nombre as operacion,
                TO_CHAR(p.fecha_creacion AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_programada,
                TO_CHAR(p.hora_inicio, 'HH24:MI') as hora_inicio,
                TO_CHAR(p.hora_fin, 'HH24:MI') as hora_fin,
                TO_CHAR(p.tiempo_fin_real AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_ejecucion,
                co.duracion_minutos,
                i.estado as estado_vehiculo,
                CASE 
                    WHEN p.tiempo_fin_real IS NOT NULL THEN 'Completado'
                    WHEN p.fecha_planificada::DATE = CURRENT_DATE AND i.estado = 'Lavado Programado' THEN 'En Curso'
                    WHEN p.fecha_planificada::DATE < CURRENT_DATE AND p.tiempo_fin_real IS NULL THEN 'Vencido'
                    ELSE 'Pendiente'
                END as estado_operacion,
                p.id as operacion_id,
                1 as orden
            FROM planificacion_operaciones p
            LEFT JOIN ingresos_usados i ON p.ingreso_id = i.id
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            LEFT JOIN config_operaciones_lavado co ON p.operacion_lavado_id = co.id
            WHERE i.id IS NOT NULL
            
            UNION ALL
            
            SELECT 
                i.dominio,
                i.clasificacion,
                m.nombre as marca,
                mo.nombre as modelo,
                TO_CHAR(i.fecha_ingreso AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_alta_sistema,
                i.limpieza_requerida as operacion,
                NULL as fecha_programada,
                NULL as hora_inicio,
                NULL as hora_fin,
                NULL as fecha_ejecucion,
                NULL as duracion_minutos,
                i.estado as estado_vehiculo,
                'En Cola' as estado_operacion,
                NULL as operacion_id,
                2 as orden
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            WHERE i.estado = 'Playa Lavadero' 
            AND i.activo = TRUE
            AND NOT EXISTS (
                SELECT 1 FROM planificacion_operaciones p2 
                WHERE p2.ingreso_id = i.id AND p2.tiempo_fin_real IS NULL
            )
            
            UNION ALL
            
            SELECT 
                k.dominio,
                'KINTO' as clasificacion,
                'KINTO' as marca,
                'Alquiler' as modelo,
                TO_CHAR(k.fecha_reserva AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_alta_sistema,
                'Standard Plus' as operacion,
                TO_CHAR(k.fecha_reserva AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') as fecha_programada,
                TO_CHAR(k.hora_inicio, 'HH24:MI') as hora_inicio,
                TO_CHAR(k.hora_fin, 'HH24:MI') as hora_fin,
                CASE WHEN k.estado = 'Completado' THEN TO_CHAR(k.fecha_reserva + k.hora_fin AT TIME ZONE 'UTC' AT TIME ZONE 'America/Argentina/Buenos_Aires', 'YYYY-MM-DD HH24:MI:SS') ELSE NULL END as fecha_ejecucion,
                k.duracion_minutos,
                k.estado as estado_vehiculo,
                CASE 
                    WHEN k.estado = 'Completado' THEN 'Completado'
                    WHEN k.fecha_reserva::DATE = CURRENT_DATE THEN 'En Curso'
                    WHEN k.fecha_reserva::DATE < CURRENT_DATE THEN 'Vencido'
                    ELSE 'Pendiente'
                END as estado_operacion,
                k.id as operacion_id,
                1 as orden
            FROM reservas_kinto k
        """
        
        # Construir WHERE para la primera parte del UNION (operaciones programadas)
        where_programadas = ""
        params = []
        
        if fecha_desde and fecha_hasta:
            where_programadas = " AND p.fecha_planificada BETWEEN %s AND %s"
            params.extend([fecha_desde, fecha_hasta])
        
        if dominio_filtro:
            where_programadas += " AND i.dominio ILIKE %s"
            params.append(f'%{dominio_filtro}%')
        
        # Construir WHERE para la segunda parte del UNION (cola)
        where_cola = ""
        if dominio_filtro:
            where_cola = " AND i.dominio ILIKE %s"
            params.append(f'%{dominio_filtro}%')
        
        # Construir WHERE para la tercera parte del UNION (KINTO)
        where_kinto = " WHERE k.estado != 'Cancelado'"  # Excluir canceladas
        if fecha_desde and fecha_hasta:
            where_kinto += " AND k.fecha_reserva BETWEEN %s AND %s"
            params.extend([fecha_desde, fecha_hasta])
        
        if dominio_filtro:
            where_kinto += " AND k.dominio ILIKE %s"
            params.append(f'%{dominio_filtro}%')
        
        # Aplicar filtros a las tres partes del UNION
        query = query.replace("WHERE i.id IS NOT NULL", f"WHERE i.id IS NOT NULL{where_programadas}")
        query = query.replace("AND NOT EXISTS", f"{where_cola} AND NOT EXISTS")
        query = query.replace("FROM reservas_kinto k", f"FROM reservas_kinto k{where_kinto}")
        
        query += " ORDER BY orden ASC, fecha_alta_sistema DESC"
        
        cursor.execute(query, params)
        registros = cursor.fetchall()
        
        release_db_connection(conn)
        return jsonify([dict(r) for r in registros]), 200
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_registro_lavados: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# --- APIs Reportes ---
@app.route('/api/usados/reportes', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_reportes_usados():
    """Generar reportes y métricas del módulo usados"""
    conn = None
    try:
        fecha_desde = request.args.get('desde')
        fecha_hasta = request.args.get('hasta')
        
        if not fecha_desde or not fecha_hasta:
            return jsonify({'error': 'Fechas requeridas'}), 400
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # 1. Total de vehículos en el período
        cursor.execute("""
            SELECT COUNT(*) as total
            FROM ingresos_usados
            WHERE fecha_ingreso BETWEEN %s AND %s
            AND activo = TRUE
        """, (fecha_desde, fecha_hasta))
        total_vehiculos = cursor.fetchone()['total']
        
        # 2. Completados (en Salón)
        cursor.execute("""
            SELECT COUNT(*) as completados
            FROM ingresos_usados
            WHERE fecha_ingreso BETWEEN %s AND %s
            AND estado = 'Salón'
            AND activo = TRUE
        """, (fecha_desde, fecha_hasta))
        completados = cursor.fetchone()['completados']
        
        # 3. Tiempo promedio en sistema (de Playa Lavadero a Salón)
        cursor.execute("""
            SELECT 
                AVG(EXTRACT(EPOCH FROM (h2.fecha_cambio - h1.fecha_cambio))/3600) as promedio_horas
            FROM historial_estados_usados h1
            INNER JOIN historial_estados_usados h2 ON h1.ingreso_id = h2.ingreso_id
            INNER JOIN ingresos_usados i ON i.id = h1.ingreso_id
            WHERE h1.estado_nuevo = 'Playa Lavadero'
            AND h2.estado_nuevo = 'Salón'
            AND i.fecha_ingreso BETWEEN %s AND %s
            AND h2.fecha_cambio > h1.fecha_cambio
        """, (fecha_desde, fecha_hasta))
        promedio_result = cursor.fetchone()
        tiempo_promedio = round(promedio_result['promedio_horas'] or 0, 1)
        
        # 4. Eficiencia (completados / total)
        eficiencia = round((completados / total_vehiculos * 100) if total_vehiculos > 0 else 0, 1)
        
        # 5. Tiempos por tipo de limpieza
        cursor.execute("""
            SELECT 
                i.limpieza_requerida,
                AVG(EXTRACT(EPOCH FROM (h2.fecha_cambio - h1.fecha_cambio))/3600) as promedio_horas
            FROM historial_estados_usados h1
            INNER JOIN historial_estados_usados h2 ON h1.ingreso_id = h2.ingreso_id
            INNER JOIN ingresos_usados i ON i.id = h1.ingreso_id
            WHERE h1.estado_nuevo = 'Playa Lavadero'
            AND h2.estado_nuevo = 'Salón'
            AND i.fecha_ingreso BETWEEN %s AND %s
            AND h2.fecha_cambio > h1.fecha_cambio
            GROUP BY i.limpieza_requerida
            ORDER BY promedio_horas DESC
        """, (fecha_desde, fecha_hasta))
        tiempos_por_tipo = [
            {
                'limpieza_requerida': row['limpieza_requerida'],
                'promedio_hrs': round(row['promedio_horas'], 1)
            }
            for row in cursor.fetchall()
        ]
        
        # 6. Estado actual del inventario
        cursor.execute("""
            SELECT 
                estado,
                COUNT(*) as cantidad
            FROM ingresos_usados
            WHERE activo = TRUE
            GROUP BY estado
            ORDER BY cantidad DESC
        """)
        estado_actual = [dict(row) for row in cursor.fetchall()]
        
        # 7. Detalle de vehículos
        cursor.execute("""
            SELECT 
                i.dominio,
                i.limpieza_requerida,
                i.fecha_ingreso,
                i.estado,
                m.nombre as marca_nombre,
                mo.nombre as modelo_nombre,
                EXTRACT(EPOCH FROM (NOW() - i.fecha_ingreso))/3600 as horas_sistema
            FROM ingresos_usados i
            LEFT JOIN marcas_usados m ON i.marca_id = m.id
            LEFT JOIN modelos_usados mo ON i.modelo_id = mo.id
            WHERE i.fecha_ingreso BETWEEN %s AND %s
            AND i.activo = TRUE
            ORDER BY i.fecha_ingreso DESC
        """, (fecha_desde, fecha_hasta))
        
        detalle_vehiculos = []
        for row in cursor.fetchall():
            horas = row['horas_sistema']
            if horas < 24:
                tiempo_texto = f"{int(horas)}h"
            else:
                dias = int(horas / 24)
                tiempo_texto = f"{dias}d"
            
            detalle_vehiculos.append({
                'dominio': row['dominio'],
                'marca_nombre': row['marca_nombre'],
                'modelo_nombre': row['modelo_nombre'],
                'limpieza_requerida': row['limpieza_requerida'],
                'estado': row['estado'],
                'fecha_ingreso': row['fecha_ingreso'].isoformat() if row['fecha_ingreso'] else None,
                'tiempo_en_sistema': tiempo_texto
            })
        
        release_db_connection(conn)
        
        return jsonify({
            'total_vehiculos': total_vehiculos,
            'completados': completados,
            'tiempo_promedio_hrs': f"{tiempo_promedio}",
            'eficiencia': eficiencia,
            'tiempos_por_tipo': tiempos_por_tipo,
            'estado_actual': estado_actual,
            'detalle_vehiculos': detalle_vehiculos
        })
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_reportes_usados: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ==================== KINTO (AUTOS DE ALQUILER) ====================

@app.route('/api/kinto/gaps/<fecha>', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_kinto_gaps(fecha):
    """Obtener huecos disponibles en el Gantt para reservas Kinto (Standard Plus)"""
    conn = None
    try:
        from datetime import datetime
        
        # Validar día de la semana
        fecha_obj = datetime.strptime(fecha, '%Y-%m-%d')
        dia_semana = fecha_obj.weekday()  # 0=Lunes, 6=Domingo
        
        # Domingos: No se trabaja
        if dia_semana == 6:  # Domingo
            return []
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener duración de Standard Plus
        cursor.execute("""
            SELECT duracion_minutos 
            FROM config_operaciones_lavado 
            WHERE nombre = 'Standard Plus'
        """)
        standard_plus = cursor.fetchone()
        duracion_kinto = standard_plus['duracion_minutos'] if standard_plus else 50
        
        # Obtener operaciones del día (Usados + Kinto) ordenadas
        cursor.execute("""
            SELECT 
                po.hora_inicio::TEXT as hora_inicio,
                po.hora_fin::TEXT as hora_fin,
                col.duracion_minutos
            FROM planificacion_operaciones po
            JOIN config_operaciones_lavado col ON po.operacion_lavado_id = col.id
            WHERE po.fecha_planificada::DATE = %s
            ORDER BY po.hora_inicio ASC
        """, (fecha,))
        operaciones = cursor.fetchall()
        
        # También obtener reservas Kinto del día
        cursor.execute("""
            SELECT hora_inicio::TEXT, hora_fin::TEXT, duracion_minutos
            FROM reservas_kinto
            WHERE fecha_reserva = %s
            AND estado = 'Reservado'
            ORDER BY hora_inicio ASC
        """, (fecha,))
        reservas_kinto = cursor.fetchall()
        
        # Combinar operaciones y reservas
        todas_operaciones = list(operaciones) + list(reservas_kinto)
        todas_operaciones.sort(key=lambda x: x['hora_inicio'])
        
        # Obtener configuración de jornada para esta fecha
        cursor.execute("""
            SELECT tipo_jornada, 
                   hora_inicio_manana::TEXT, hora_fin_manana::TEXT,
                   hora_inicio_tarde::TEXT, hora_fin_tarde::TEXT
            FROM config_jornadas_lavadero
            WHERE fecha = %s
        """, (fecha,))
        config_jornada = cursor.fetchone()
        
        release_db_connection(conn)
        
        # Obtener hora actual si es hoy
        ahora = datetime.now()
        es_hoy = fecha_obj.date() == ahora.date()
        minuto_actual = ahora.hour * 60 + ahora.minute if es_hoy else 0
        
        # Sábados: Horario especial 9:00 - 13:00 (solo mañana)
        if dia_semana == 5:  # Sábado
            hora_inicio_lavadero = 9 * 60  # 09:00
            hora_fin_lavadero = 13 * 60     # 13:00
            tiene_descanso = False
            
            # Para sábados, usar lógica simple de jornada corrida pero con horario 9-13
            gaps = []
            if not todas_operaciones:
                tiempo_segmento = max(hora_inicio_lavadero, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_lavadero
                while tiempo_segmento + duracion_kinto <= hora_fin_lavadero:
                    gaps.append({
                        'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                        'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                        'duracion_minutos': duracion_kinto,
                        'minuto_inicio': tiempo_segmento,
                        'minuto_fin': tiempo_segmento + duracion_kinto
                    })
                    tiempo_segmento += duracion_kinto
            else:
                tiempo_actual = max(hora_inicio_lavadero, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_lavadero
                for op in todas_operaciones:
                    h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                    h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                    minuto_inicio_op = h_ini * 60 + m_ini
                    minuto_fin_op = h_fin * 60 + m_fin
                    
                    if tiempo_actual < minuto_inicio_op:
                        tiempo_segmento = tiempo_actual
                        while tiempo_segmento + duracion_kinto <= minuto_inicio_op and tiempo_segmento + duracion_kinto <= hora_fin_lavadero:
                            gaps.append({
                                'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                                'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                                'duracion_minutos': duracion_kinto,
                                'minuto_inicio': tiempo_segmento,
                                'minuto_fin': tiempo_segmento + duracion_kinto
                            })
                            tiempo_segmento += duracion_kinto
                    
                    tiempo_actual = max(tiempo_actual, minuto_fin_op)
                
                if tiempo_actual < hora_fin_lavadero:
                    tiempo_segmento = tiempo_actual
                    while tiempo_segmento + duracion_kinto <= hora_fin_lavadero:
                        gaps.append({
                            'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                            'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                            'duracion_minutos': duracion_kinto,
                            'minuto_inicio': tiempo_segmento,
                            'minuto_fin': tiempo_segmento + duracion_kinto
                        })
                        tiempo_segmento += duracion_kinto
            
            return gaps
        
        # Determinar horarios según tipo de jornada (Lunes a Viernes)
        if config_jornada and config_jornada['tipo_jornada'] == 'corrida':
            # Jornada corrida: 08:30 - 17:30
            hora_inicio_lavadero = 8 * 60 + 30  # 08:30
            hora_fin_lavadero = 17 * 60 + 30     # 17:30
            tiene_descanso = False
        else:
            # Jornada cortada (por defecto): 08:30 - 13:00 y 16:00 - 20:00
            hora_inicio_manana = 8 * 60 + 30  # 08:30
            hora_fin_manana = 13 * 60          # 13:00
            hora_inicio_tarde = 16 * 60        # 16:00
            hora_fin_tarde = 20 * 60           # 20:00
            tiene_descanso = True
        
        gaps = []
        
        if not todas_operaciones:
            # Todo el día está libre - dividir en segmentos
            if tiene_descanso:
                # Jornada cortada: generar gaps en dos tramos
                # Tramo mañana: 08:30 - 13:00
                tiempo_segmento = max(hora_inicio_manana, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_manana
                while tiempo_segmento + duracion_kinto <= hora_fin_manana:
                    gaps.append({
                        'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                        'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                        'duracion_minutos': duracion_kinto,
                        'minuto_inicio': tiempo_segmento,
                        'minuto_fin': tiempo_segmento + duracion_kinto
                    })
                    tiempo_segmento += duracion_kinto
                
                # Tramo tarde: 16:00 - 20:00
                tiempo_segmento = max(hora_inicio_tarde, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_tarde
                while tiempo_segmento + duracion_kinto <= hora_fin_tarde:
                    gaps.append({
                        'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                        'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                        'duracion_minutos': duracion_kinto,
                        'minuto_inicio': tiempo_segmento,
                        'minuto_fin': tiempo_segmento + duracion_kinto
                    })
                    tiempo_segmento += duracion_kinto
            else:
                # Jornada corrida: 08:30 - 17:30
                tiempo_segmento = max(hora_inicio_lavadero, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_lavadero
                while tiempo_segmento + duracion_kinto <= hora_fin_lavadero:
                    gaps.append({
                        'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                        'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                        'duracion_minutos': duracion_kinto,
                        'minuto_inicio': tiempo_segmento,
                        'minuto_fin': tiempo_segmento + duracion_kinto
                    })
                    tiempo_segmento += duracion_kinto
        else:
            # Buscar gaps entre operaciones
            if tiene_descanso:
                # Jornada cortada: procesar mañana y tarde por separado
                # TRAMO MAÑANA (08:30 - 13:00)
                tiempo_actual = max(hora_inicio_manana, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_manana
                for op in todas_operaciones:
                    h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                    h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                    minuto_inicio_op = h_ini * 60 + m_ini
                    minuto_fin_op = h_fin * 60 + m_fin
                    
                    # Solo procesar operaciones que caen en el tramo de mañana
                    if minuto_inicio_op >= hora_fin_manana:
                        break  # Ya pasamos al tramo de tarde
                    
                    if tiempo_actual < minuto_inicio_op and tiempo_actual < hora_fin_manana:
                        limite_gap = min(minuto_inicio_op, hora_fin_manana)
                        tiempo_segmento = tiempo_actual
                        while tiempo_segmento + duracion_kinto <= limite_gap:
                            gaps.append({
                                'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                                'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                                'duracion_minutos': duracion_kinto,
                                'minuto_inicio': tiempo_segmento,
                                'minuto_fin': tiempo_segmento + duracion_kinto
                            })
                            tiempo_segmento += duracion_kinto
                    
                    tiempo_actual = max(tiempo_actual, minuto_fin_op)
                
                # Gap final de la mañana
                if tiempo_actual < hora_fin_manana:
                    tiempo_segmento = tiempo_actual
                    while tiempo_segmento + duracion_kinto <= hora_fin_manana:
                        gaps.append({
                            'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                            'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                            'duracion_minutos': duracion_kinto,
                            'minuto_inicio': tiempo_segmento,
                            'minuto_fin': tiempo_segmento + duracion_kinto
                        })
                        tiempo_segmento += duracion_kinto
                
                # TRAMO TARDE (16:00 - 20:00)
                tiempo_actual = max(hora_inicio_tarde, minuto_actual + duracion_kinto) if es_hoy else hora_inicio_tarde
                for op in todas_operaciones:
                    h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                    h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                    minuto_inicio_op = h_ini * 60 + m_ini
                    minuto_fin_op = h_fin * 60 + m_fin
                    
                    # Solo procesar operaciones que caen en el tramo de tarde
                    if minuto_fin_op <= hora_inicio_tarde:
                        continue  # Esta operación fue en la mañana
                    
                    if tiempo_actual < minuto_inicio_op and minuto_inicio_op >= hora_inicio_tarde:
                        tiempo_segmento = max(tiempo_actual, hora_inicio_tarde)
                        while tiempo_segmento + duracion_kinto <= minuto_inicio_op:
                            gaps.append({
                                'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                                'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                                'duracion_minutos': duracion_kinto,
                                'minuto_inicio': tiempo_segmento,
                                'minuto_fin': tiempo_segmento + duracion_kinto
                            })
                            tiempo_segmento += duracion_kinto
                    
                    tiempo_actual = max(tiempo_actual, minuto_fin_op)
                
                # Gap final de la tarde
                if tiempo_actual < hora_fin_tarde:
                    tiempo_segmento = max(tiempo_actual, hora_inicio_tarde)
                    while tiempo_segmento + duracion_kinto <= hora_fin_tarde:
                        gaps.append({
                            'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                            'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                            'duracion_minutos': duracion_kinto,
                            'minuto_inicio': tiempo_segmento,
                            'minuto_fin': tiempo_segmento + duracion_kinto
                        })
                        tiempo_segmento += duracion_kinto
            else:
                # Jornada corrida: procesar todo el día de una vez
                tiempo_actual = hora_inicio_lavadero
                for op in todas_operaciones:
                    h_ini, m_ini = map(int, op['hora_inicio'].split(':')[:2])
                    h_fin, m_fin = map(int, op['hora_fin'].split(':')[:2])
                    minuto_inicio_op = h_ini * 60 + m_ini
                    minuto_fin_op = h_fin * 60 + m_fin
                    
                    if tiempo_actual < minuto_inicio_op:
                        tiempo_segmento = tiempo_actual
                        while tiempo_segmento + duracion_kinto <= minuto_inicio_op:
                            gaps.append({
                                'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                                'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                                'duracion_minutos': duracion_kinto,
                                'minuto_inicio': tiempo_segmento,
                                'minuto_fin': tiempo_segmento + duracion_kinto
                            })
                            tiempo_segmento += duracion_kinto
                    
                    tiempo_actual = max(tiempo_actual, minuto_fin_op)
                
                # Gap después de la última operación
                if tiempo_actual < hora_fin_lavadero:
                    tiempo_segmento = tiempo_actual
                    while tiempo_segmento + duracion_kinto <= hora_fin_lavadero:
                        gaps.append({
                            'hora_inicio': f"{tiempo_segmento // 60:02d}:{tiempo_segmento % 60:02d}",
                            'hora_fin': f"{(tiempo_segmento + duracion_kinto) // 60:02d}:{(tiempo_segmento + duracion_kinto) % 60:02d}",
                            'duracion_minutos': duracion_kinto,
                            'minuto_inicio': tiempo_segmento,
                            'minuto_fin': tiempo_segmento + duracion_kinto
                        })
                        tiempo_segmento += duracion_kinto
        
        # FILTRAR GAPS EN EL PASADO si es hoy
        from datetime import datetime
        hoy = datetime.now().date()
        
        if fecha_obj.date() == hoy:
            # Solo es hoy - filtrar gaps que ya pasaron
            hora_actual = datetime.now().hour
            minuto_actual = datetime.now().minute
            minutos_actuales = hora_actual * 60 + minuto_actual
            
            # Filtrar gaps que ya pasaron completamente
            gaps = [g for g in gaps if g['minuto_fin'] > minutos_actuales]
        
        return jsonify(gaps)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_kinto_gaps: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


@app.route('/api/kinto/reservas', methods=['POST'])
@login_required
@module_permission_required('usados')
def crear_reserva_kinto():
    """Crear reserva Kinto y programarla directamente en el Gantt"""
    conn = None
    try:
        data = request.get_json()
        fecha = data['fecha']
        hora_inicio = data['hora_inicio']
        dominio = data['dominio'].upper().strip()
        observaciones = data.get('observaciones', '')
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        # Obtener ID de Standard Plus
        cursor.execute("""
            SELECT id, duracion_minutos 
            FROM config_operaciones_lavado 
            WHERE nombre = 'Standard Plus'
        """)
        standard_plus = cursor.fetchone()
        if not standard_plus:
            release_db_connection(conn)
            return jsonify({'error': 'Standard Plus no configurado'}), 400
        
        operacion_id = standard_plus['id']
        duracion = standard_plus['duracion_minutos']
        
        # Calcular hora_fin
        from datetime import datetime, timedelta
        hora_inicio_dt = datetime.strptime(hora_inicio, '%H:%M')
        hora_fin_dt = hora_inicio_dt + timedelta(minutes=duracion)
        hora_fin = hora_fin_dt.strftime('%H:%M')
        
        # Validar que no haya conflictos (operaciones Usados o Kinto ya reservadas)
        h_ini, m_ini = map(int, hora_inicio.split(':'))
        h_fin, m_fin = map(int, hora_fin.split(':'))
        minuto_inicio = h_ini * 60 + m_ini
        minuto_fin = h_fin * 60 + m_fin
        
        # Verificar conflictos con operaciones Usados
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM planificacion_operaciones
            WHERE fecha_planificada::DATE = %s
            AND (
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s 
                 AND EXTRACT(HOUR FROM hora_fin)*60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) >= %s 
                 AND EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        
        conflictos = cursor.fetchone()['count']
        if conflictos > 0:
            release_db_connection(conn)
            return jsonify({'error': 'Conflicto con operaciones existentes'}), 400
        
        # Verificar conflictos con otras reservas Kinto
        cursor.execute("""
            SELECT COUNT(*) as count
            FROM reservas_kinto
            WHERE fecha_reserva = %s
            AND estado = 'Reservado'
            AND (
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s 
                 AND EXTRACT(HOUR FROM hora_fin)*60 + EXTRACT(MINUTE FROM hora_fin) > %s)
                OR
                (EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) >= %s 
                 AND EXTRACT(HOUR FROM hora_inicio)*60 + EXTRACT(MINUTE FROM hora_inicio) < %s)
            )
        """, (fecha, minuto_fin, minuto_inicio, minuto_inicio, minuto_fin))
        
        conflictos_kinto = cursor.fetchone()['count']
        if conflictos_kinto > 0:
            release_db_connection(conn)
            return jsonify({'error': 'Conflicto con otras reservas Kinto'}), 400
        
        # Crear reserva Kinto programada directamente
        cursor.execute("""
            INSERT INTO reservas_kinto 
            (fecha_reserva, hora_inicio, hora_fin, dominio, duracion_minutos, observaciones, usuario)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (fecha, hora_inicio, hora_fin, dominio, duracion, observaciones, current_user.username))
        
        reserva_id = cursor.fetchone()['id']
        
        # TAMBIÉN crear registro en planificacion_operaciones para el Registro de Lavados
        # Buscar el vehículo KINTO en ingresos_usados (stock fijo)
        cursor.execute("""
            SELECT id FROM ingresos_usados 
            WHERE dominio = %s AND clasificacion = 'KINTO' AND es_stock_fijo = TRUE
            LIMIT 1
        """, (dominio,))
        
        kinto_ingreso = cursor.fetchone()
        
        if kinto_ingreso:
            # Obtener último orden de ejecución
            cursor.execute("""
                SELECT COALESCE(MAX(orden_ejecucion), 0) as max_orden
                FROM planificacion_operaciones
                WHERE fecha_planificada::DATE = %s
            """, (fecha,))
            max_orden = cursor.fetchone()['max_orden']
            
            # Insertar en planificacion_operaciones
            cursor.execute("""
                INSERT INTO planificacion_operaciones
                (ingreso_id, fecha_planificada, hora_inicio, hora_fin, 
                 posicion_lavadero, operacion_lavado_id, orden_ejecucion)
                VALUES (%s, %s, %s, %s, 1, %s, %s)
            """, (
                kinto_ingreso['id'],
                fecha,
                hora_inicio,
                hora_fin,
                operacion_id,
                max_orden + 1
            ))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'reserva_id': reserva_id,
            'mensaje': f'Reserva Kinto creada para {dominio}'
        }), 201
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en crear_reserva_kinto: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500


# ============================================================================
# MÓDULO POST VENTA
# ============================================================================

@app.route('/postventa')
@login_required
@module_permission_required('postventa')
def postventa():
    """Página principal del módulo Post Venta"""
    return render_template('postventa.html')

@app.route('/postventa/gestion')
@login_required
@module_permission_required('postventa')
def postventa_gestion():
    """Gestión de Servicios - Carga de turnos y administración"""
    # Solo admin, usuario_plus y supervisor
    if current_user.role not in ['admin', 'usuario_plus', 'supervisor']:
        flash('No tienes permisos para acceder a Gestión de Servicios', 'error')
        return redirect(url_for('postventa'))
    return render_template('postventa_gestion.html')

@app.route('/postventa/ordenes')
@login_required
@module_permission_required('postventa')
def postventa_ordenes():
    """Órdenes de Trabajo - Gestión de turnos asignados"""
    return render_template('postventa_ordenes.html')

@app.route('/postventa/asesores')
@login_required
@module_permission_required('postventa')
def postventa_admin_asesores():
    """Administración de Asesores de Post Venta"""
    # Solo admin y supervisor
    if current_user.role not in ['admin', 'supervisor']:
        flash('No tienes permisos para acceder a Administración', 'error')
        return redirect(url_for('postventa'))
    return render_template('postventa_asesores.html')

@app.route('/postventa/servicios')
@login_required
@module_permission_required('postventa')
def postventa_admin_servicios():
    """Administración de Servicios de Post Venta"""
    # Solo admin y supervisor
    if current_user.role not in ['admin', 'supervisor']:
        flash('No tienes permisos para acceder a Administración', 'error')
        return redirect(url_for('postventa'))
    return render_template('postventa_admin_servicios.html')

@app.route('/postventa/reportes')
@login_required
@module_permission_required('postventa')
def postventa_reportes():
    """Registros y Reportes"""
    # Solo admin y supervisor
    if current_user.role not in ['admin', 'supervisor']:
        flash('No tienes permisos para acceder a Reportes', 'error')
        return redirect(url_for('postventa'))
    return render_template('postventa_reportes.html')

@app.route('/postventa/horarios-pactados')
@login_required
@module_permission_required('postventa')
def postventa_horarios_pactados():
    """Visualización de horarios pactados con clientes"""
    return render_template('postventa_horarios_pactados.html')

@app.route('/postventa/guia')
@login_required
@module_permission_required('postventa')
def postventa_guia():
    """Guía de Usuario de Post Venta"""
    return render_template('postventa_guia.html')

# ============================================================================
# API USADOS - PROVEEDORES
# ============================================================================

@app.route('/api/usados/proveedores', methods=['GET'])
@login_required
@module_permission_required('usados')
def get_proveedores():
    """Obtener todos los proveedores"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            SELECT id, nombre, cuit, rubro, especialidad, 
                   telefono, ubicacion, activo,
                   fecha_creacion, fecha_modificacion
            FROM usados_proveedores
            ORDER BY nombre ASC
        """)
        
        proveedores = cursor.fetchall() or []
        release_db_connection(conn)
        
        return jsonify(proveedores), 200
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_proveedores: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/proveedores', methods=['POST'])
@login_required
@module_permission_required('usados')
def create_proveedor():
    """Crear nuevo proveedor"""
    conn = None
    try:
        data = request.get_json()
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            INSERT INTO usados_proveedores
            (nombre, cuit, rubro, especialidad, telefono, ubicacion, activo)
            VALUES (%s, %s, %s, %s, %s, %s, %s)
            RETURNING id
        """, (
            data['nombre'],
            data['cuit'],
            data.get('rubro'),
            data.get('especialidad'),
            data.get('telefono'),
            data.get('ubicacion'),
            data.get('activo', True)
        ))
        
        proveedor_id = cursor.fetchone()['id']
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True, 'id': proveedor_id}), 201
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en create_proveedor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/proveedores/<int:proveedor_id>', methods=['PUT'])
@login_required
@module_permission_required('usados')
def update_proveedor(proveedor_id):
    """Actualizar proveedor existente"""
    conn = None
    try:
        data = request.get_json()
        
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            UPDATE usados_proveedores
            SET nombre = %s,
                cuit = %s,
                rubro = %s,
                especialidad = %s,
                telefono = %s,
                ubicacion = %s,
                activo = %s,
                fecha_modificacion = CURRENT_TIMESTAMP
            WHERE id = %s
        """, (
            data['nombre'],
            data['cuit'],
            data.get('rubro'),
            data.get('especialidad'),
            data.get('telefono'),
            data.get('ubicacion'),
            data.get('activo', True),
            proveedor_id
        ))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en update_proveedor: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/usados/proveedores/<int:proveedor_id>', methods=['DELETE'])
@login_required
@module_permission_required('usados')
def delete_proveedor(proveedor_id):
    """Eliminar proveedor"""
    conn = None
    try:
        conn = get_pg_connection()
        cursor = conn.cursor()
        
        cursor.execute("""
            DELETE FROM usados_proveedores
            WHERE id = %s
        """, (proveedor_id,))
        
        conn.commit()
        release_db_connection(conn)
        
        return jsonify({'success': True}), 200
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en delete_proveedor: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - SERVICIOS
# ============================================================================

@app.route('/api/postventa/servicios', methods=['GET', 'POST'])
@login_required
def get_servicios_postventa():
    """Obtener todos los servicios o crear uno nuevo"""
    if request.method == 'POST':
        return create_servicio_postventa()
    
    # GET request
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        cur.execute("""
            SELECT s.id, s.nombre, s.sector, s.ranking, s.asesor_fijo_id, s.activo,
                   a.nombre as asesor_nombre
            FROM pv_servicios s
            LEFT JOIN pv_asesores a ON s.asesor_fijo_id = a.id
            ORDER BY s.sector, s.ranking DESC, s.nombre
        """)
        
        servicios = []
        rows = cur.fetchall()
        print(f"Filas obtenidas: {len(rows)}")
        
        for row in rows:
            servicios.append({
                'id': row['id'],
                'nombre': row['nombre'],
                'sector': row['sector'],
                'ranking': row['ranking'],
                'asesor_fijo_id': row['asesor_fijo_id'],
                'activo': row['activo'],
                'asesor_nombre': row['asesor_nombre']
            })
        
        cur.close()
        release_db_connection(conn)
        print(f"Retornando {len(servicios)} servicios")
        return jsonify(servicios)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_servicios_postventa: {type(e).__name__} - {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/servicios', methods=['POST'])
@login_required
def create_servicio_postventa():
    """Crear o actualizar un servicio"""
    conn = None
    try:
        data = request.get_json()
        conn = get_db_connection()
        cur = conn.cursor()
        
        if data.get('id'):
            # Actualizar
            cur.execute("""
                UPDATE pv_servicios
                SET nombre = %s, sector = %s, ranking = %s, 
                    asesor_fijo_id = %s, activo = %s
                WHERE id = %s
            """, (data['nombre'], data['sector'], data['ranking'],
                  data.get('asesor_fijo_id'), data.get('activo', True), data['id']))
        else:
            # Crear
            cur.execute("""
                INSERT INTO pv_servicios (nombre, sector, ranking, asesor_fijo_id, activo)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            """, (data['nombre'], data['sector'], data['ranking'],
                  data.get('asesor_fijo_id'), data.get('activo', True)))
            result = cur.fetchone()
            data['id'] = result['id']
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        return jsonify(data)
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en create_servicio_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/servicios/<int:servicio_id>', methods=['DELETE', 'PUT'])
@login_required
def manage_servicio_postventa(servicio_id):
    """Actualizar o eliminar un servicio"""
    if request.method == 'PUT':
        # Actualizar servicio
        conn = None
        try:
            data = request.get_json()
            conn = get_db_connection()
            cur = conn.cursor()
            
            cur.execute("""
                UPDATE pv_servicios
                SET nombre = %s, sector = %s, ranking = %s, 
                    asesor_fijo_id = %s, activo = %s
                WHERE id = %s
            """, (data['nombre'], data['sector'], data['ranking'],
                  data.get('asesor_fijo_id'), data.get('activo', True), servicio_id))
            
            conn.commit()
            cur.close()
            release_db_connection(conn)
            return jsonify({'id': servicio_id, **data})
            
        except Exception as e:
            if conn:
                conn.rollback()
                release_db_connection(conn)
            print(f"Error actualizando servicio: {e}")
            return jsonify({'error': str(e)}), 500
    else:
        # DELETE - Desactivar servicio
        conn = None
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            
            cur.execute("UPDATE pv_servicios SET activo = FALSE WHERE id = %s", (servicio_id,))
            
            conn.commit()
            cur.close()
            release_db_connection(conn)
            return jsonify({'mensaje': 'Servicio desactivado'})
            
        except Exception as e:
            if conn:
                conn.rollback()
                release_db_connection(conn)
            print(f"Error en delete_servicio_postventa: {e}")
            return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - ASESORES
# ============================================================================

@app.route('/api/postventa/asesores', methods=['GET', 'POST'])
@login_required
def get_asesores_postventa():
    """Obtener todos los asesores o crear uno nuevo"""
    if request.method == 'POST':
        return create_asesor_postventa()
    
    # GET request
    conn = None
    try:
        conn = get_db_connection()
        cur = conn.cursor()
        
        sector = request.args.get('sector')
        
        if sector:
            cur.execute("""
                SELECT a.id, a.nombre, a.usuario_id, u.full_name as usuario_nombre, a.sector, a.activo
                FROM pv_asesores a
                LEFT JOIN users u ON a.usuario_id = u.id
                WHERE a.sector = %s AND a.activo = TRUE
                ORDER BY a.nombre
            """, (sector,))
        else:
            cur.execute("""
                SELECT a.id, a.nombre, a.usuario_id, u.full_name as usuario_nombre, a.sector, a.activo
                FROM pv_asesores a
                LEFT JOIN users u ON a.usuario_id = u.id
                WHERE a.activo = TRUE
                ORDER BY a.sector, a.nombre
            """)
        
        asesores = []
        for row in cur.fetchall():
            asesores.append({
                'id': row['id'],
                'nombre': row['nombre'],
                'usuario_id': row['usuario_id'],
                'usuario_nombre': row['usuario_nombre'],
                'sector': row['sector'],
                'activo': row['activo']
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(asesores)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_asesores_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/asesores', methods=['POST'])
@login_required
def create_asesor_postventa():
    """Crear o actualizar un asesor"""
    conn = None
    try:
        data = request.get_json()
        conn = get_db_connection()
        cur = conn.cursor()
        
        user_id = data.get('usuario_id')
        
        # Si se proporciona información de usuario, crear el usuario
        if data.get('crear_usuario') and data.get('username'):
            cur.execute('''
                INSERT INTO users (username, password_hash, role, full_name, active)
                VALUES (%s, %s, %s, %s, %s)
                RETURNING id
            ''', (
                data['username'],
                generate_password_hash(data.get('password', 'asesor123')),
                'usuario',
                data['nombre'],
                True
            ))
            user_result = cur.fetchone()
            user_id = user_result['id']
        
        if data.get('id'):
            # Actualizar
            cur.execute("""
                UPDATE pv_asesores
                SET nombre = %s, usuario_id = %s, sector = %s, activo = %s
                WHERE id = %s
            """, (data['nombre'], user_id, data['sector'],
                  data.get('activo', True), data['id']))
        else:
            # Crear
            cur.execute("""
                INSERT INTO pv_asesores (nombre, usuario_id, sector, activo)
                VALUES (%s, %s, %s, %s)
                RETURNING id
            """, (data['nombre'], user_id, data['sector'],
                  data.get('activo', True)))
            result = cur.fetchone()
            data['id'] = result['id']
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        return jsonify(data)
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en create_asesor_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/asesores/<int:asesor_id>', methods=['DELETE', 'PUT'])
@login_required
def manage_asesor_postventa(asesor_id):
    """Actualizar o eliminar un asesor"""
    if request.method == 'PUT':
        # Actualizar asesor
        conn = None
        try:
            data = request.get_json()
            conn = get_db_connection()
            cur = conn.cursor()
            
            cur.execute("""
                UPDATE pv_asesores
                SET nombre = %s, usuario_id = %s, sector = %s, activo = %s
                WHERE id = %s
            """, (data['nombre'], data['usuario_id'], data['sector'],
                  data.get('activo', True), asesor_id))
            
            conn.commit()
            cur.close()
            release_db_connection(conn)
            return jsonify({'id': asesor_id, **data})
            
        except Exception as e:
            if conn:
                conn.rollback()
                release_db_connection(conn)
            print(f"Error actualizando asesor: {e}")
            return jsonify({'error': str(e)}), 500
    else:
        # DELETE - Desactivar asesor
        conn = None
        try:
            conn = get_db_connection()
            cur = conn.cursor()
            
            cur.execute("UPDATE pv_asesores SET activo = FALSE WHERE id = %s", (asesor_id,))
            
            conn.commit()
            cur.close()
            release_db_connection(conn)
            return jsonify({'mensaje': 'Asesor desactivado'})
            
        except Exception as e:
            if conn:
                conn.rollback()
                release_db_connection(conn)
            print(f"Error en delete_asesor_postventa: {e}")
            return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - CARGA DE TURNOS
# ============================================================================

@app.route('/api/postventa/exportar-turnos-no-asignados', methods=['GET'])
@login_required
def exportar_turnos_no_asignados():
    """Exportar turnos con estado='Asignado' antes de borrarlos"""
    import io
    import csv
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Obtener turnos no asignados
        cur.execute("""
            SELECT fecha_turno, hora_inicio, dominio, cliente_nombre, 
                   servicio_nombre, marca_modelo, sector
            FROM pv_turnos
            WHERE estado = 'Asignado'
            ORDER BY fecha_turno, hora_inicio
        """)
        
        turnos = cur.fetchall()
        cur.close()
        release_db_connection(conn)
        
        if not turnos:
            return jsonify({'mensaje': 'No hay turnos pendientes para exportar'}), 404
        
        # Crear CSV en memoria
        output = io.StringIO()
        writer = csv.writer(output, delimiter=';')
        
        # Escribir encabezados
        writer.writerow(['Fec.Turno', 'Inicio', 'Dominio', 'Nombre (C)', 'Servicio', 'Marca y Modelo / Versión', 'Sector'])
        
        # Escribir datos
        for row in turnos:
            writer.writerow([
                row['fecha_turno'],
                row['hora_inicio'],
                row['dominio'],
                row['cliente_nombre'],
                row['servicio_nombre'],
                row['marca_modelo'],
                row['sector']
            ])
        
        # Convertir a bytes
        output.seek(0)
        from datetime import datetime
        filename = f"turnos_no_asignados_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
        
        return send_file(
            io.BytesIO(output.getvalue().encode('utf-8-sig')),  # BOM para Excel
            mimetype='text/csv',
            as_attachment=True,
            download_name=filename
        )
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en exportar_turnos_no_asignados: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/upload-turnos', methods=['POST'])
@login_required
def upload_turnos_postventa():
    """Cargar turnos desde archivo Excel"""
    conn = None
    try:
        if 'file' not in request.files:
            return jsonify({'error': 'No se recibió archivo'}), 400
        
        file = request.files['file']
        if file.filename == '':
            return jsonify({'error': 'Archivo vacío'}), 400
        
        # Leer Excel
        df = pd.read_excel(file)
        
        # Limpiar nombres de columnas (quitar espacios y ; al final)
        df.columns = df.columns.str.strip().str.rstrip(';')
        
        print(f"Columnas encontradas: {list(df.columns)}")
        
        # Validar columnas requeridas
        required_cols = ['Fec.Turno', 'Inicio', 'Dominio', 'Nombre (C)', 'Servicio', 'Marca y Modelo / Versión']
        missing_cols = [col for col in required_cols if col not in df.columns]
        if missing_cols:
            return jsonify({'error': f'Faltan columnas: {", ".join(missing_cols)}. Columnas encontradas: {", ".join(df.columns)}'}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Limpiar turnos anteriores con estado 'Asignado' para evitar duplicados
        cur.execute("DELETE FROM pv_turnos WHERE estado = 'Asignado'")
        print(f"Turnos anteriores eliminados: {cur.rowcount}")
        
        # Obtener servicios para clasificación (incluyendo ID)
        cur.execute("SELECT id, nombre, sector FROM pv_servicios WHERE activo = TRUE")
        servicios_dict = {row['nombre']: {'id': row['id'], 'sector': row['sector']} for row in cur.fetchall()}
        
        turnos_insertados = []
        errores = []
        
        # Obtener fecha actual para asignar a los turnos
        from datetime import datetime, date
        fecha_hoy = date.today()
        
        for idx, row in df.iterrows():
            try:
                servicio_nombre = str(row['Servicio']).strip()
                
                # Obtener información del servicio (sector y ID)
                servicio_info = servicios_dict.get(servicio_nombre)
                if servicio_info:
                    sector = servicio_info['sector']
                    servicio_id = servicio_info['id']
                else:
                    sector = 'Servicios'  # Default a Servicios
                    servicio_id = None
                
                # Usar fecha actual (fecha de carga)
                fec_turno = fecha_hoy
                
                # Manejar hora_inicio que viene como número decimal (8.5 = 08:30, 16.833 = 16:50)
                hora_inicio_raw = row['Inicio']
                if pd.isna(hora_inicio_raw):
                    errores.append(f"Fila {idx+2}: Hora de inicio vacía")
                    continue
                
                if isinstance(hora_inicio_raw, str):
                    # Si viene como string, usar directamente
                    hora_inicio = hora_inicio_raw.strip()
                elif isinstance(hora_inicio_raw, (int, float)):
                    # Si viene como número decimal: 8.5 = 08:30, 16.833 = 16:50
                    horas = int(hora_inicio_raw)
                    minutos = int((hora_inicio_raw - horas) * 100)  # La parte decimal * 100
                    hora_inicio = f"{horas:02d}:{minutos:02d}"
                else:
                    # Si viene como datetime o time
                    hora_inicio = pd.to_datetime(hora_inicio_raw).strftime('%H:%M')
                
                cur.execute("""
                    INSERT INTO pv_turnos 
                    (fecha_turno, hora_inicio, dominio, cliente_nombre, servicio_nombre, servicio_id, marca_modelo, sector, estado)
                    VALUES (%s, %s, %s, %s, %s, %s, %s, %s, 'Asignado')
                    RETURNING id
                """, (fec_turno, hora_inicio, str(row['Dominio']).strip(),
                      str(row['Nombre (C)']).strip(), servicio_nombre, servicio_id,
                      str(row['Marca y Modelo / Versión']).strip(), sector))
                
                result = cur.fetchone()
                turno_id = result['id']
                turnos_insertados.append({
                    'id': turno_id,
                    'fec_turno': str(fec_turno),
                    'hora_inicio': hora_inicio,
                    'dominio': str(row['Dominio']).strip(),
                    'nombre_cliente': str(row['Nombre (C)']).strip(),
                    'servicio': servicio_nombre,
                    'marca_modelo': str(row['Marca y Modelo / Versión']).strip(),
                    'sector': sector,
                    'estado': 'Asignado'
                })
                
            except Exception as e:
                errores.append(f"Fila {idx+2}: {str(e)}")
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({
            'turnos': turnos_insertados,
            'total_insertados': len(turnos_insertados),
            'errores': errores
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en upload_turnos_postventa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/crear-turno-manual', methods=['POST'])
@login_required
def crear_turno_manual():
    """Crear un turno manualmente desde el formulario"""
    conn = None
    try:
        data = request.get_json()
        
        # Validar datos requeridos
        required_fields = ['fecha_turno', 'hora_inicio', 'sector', 'dominio', 'cliente_nombre']
        missing_fields = [field for field in required_fields if not data.get(field)]
        if missing_fields:
            return jsonify({'error': f'Faltan campos requeridos: {", ".join(missing_fields)}'}), 400
        
        # Validar formato de sector
        if data['sector'] not in ['Servicios', 'Chapería y Pintura']:
            return jsonify({'error': 'Sector inválido'}), 400
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Insertar el turno
        cur.execute("""
            INSERT INTO pv_turnos (
                fecha_turno, hora_inicio, dominio, cliente_nombre,
                servicio_nombre, marca_modelo, sector, estado
            ) VALUES (%s, %s, %s, %s, %s, %s, %s, 'Asignado')
            RETURNING id
        """, (
            data['fecha_turno'],
            data['hora_inicio'],
            data['dominio'].upper(),
            data['cliente_nombre'],
            data.get('servicio_nombre'),
            data.get('marca_modelo'),
            data['sector']
        ))
        
        result = cur.fetchone()
        turno_id = result['id']
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({
            'success': True,
            'turno_id': turno_id,
            'message': 'Turno creado exitosamente'
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en crear_turno_manual: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/turnos', methods=['GET'])
@login_required
def get_turnos_postventa():
    """Obtener turnos pendientes de asignación"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        estado = request.args.get('estado', 'Asignado')
        
        cur.execute("""
            SELECT t.id, t.fecha_turno, t.hora_inicio, t.dominio, t.cliente_nombre,
                   t.servicio_nombre, t.marca_modelo, t.sector, t.estado, t.asesor_id,
                   a.nombre as asesor_nombre
            FROM pv_turnos t
            LEFT JOIN pv_asesores a ON t.asesor_id = a.id
            WHERE t.estado = %s
            ORDER BY t.fecha_turno, t.hora_inicio
        """, (estado,))
        
        turnos = []
        for row in cur.fetchall():
            turnos.append({
                'id': row['id'],
                'fecha_turno': str(row['fecha_turno']),
                'hora_inicio': str(row['hora_inicio']) if row['hora_inicio'] else None,
                'dominio': row['dominio'],
                'cliente_nombre': row['cliente_nombre'],
                'servicio_nombre': row['servicio_nombre'],
                'marca_modelo': row['marca_modelo'],
                'sector': row['sector'],
                'estado': row['estado'],
                'asesor_id': row['asesor_id'],
                'asesor_nombre': row['asesor_nombre']
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(turnos)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_turnos_postventa: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - ASIGNACIÓN DE TURNOS
# ============================================================================

def asignar_asesor_balanceado(sector, servicio_nombre, conn):
    """
    Asigna un asesor de forma balanceada según el ranking del servicio.
    Los asesores deben recibir proporcionalmente tareas de alto ranking.
    """
    cur = conn.cursor()
    
    # Obtener ranking del servicio
    cur.execute("SELECT ranking FROM pv_servicios WHERE nombre = %s", (servicio_nombre,))
    result = cur.fetchone()
    ranking = result['ranking'] if result else 3  # Default ranking 3
    
    # Obtener asesores del sector (excluir 'Asesor de citas' de asignación automática)
    cur.execute("""
        SELECT id, nombre FROM pv_asesores 
        WHERE sector = %s AND activo = TRUE AND sector != 'Asesor de citas'
    """, (sector,))
    asesores_rows = cur.fetchall()
    
    if not asesores_rows:
        cur.close()
        return None
    
    # Contar tareas actuales por asesor desde pv_turnos (no pv_historial)
    stats = {}
    for asesor_row in asesores_rows:
        asesor_id = asesor_row['id']
        asesor_nombre = asesor_row['nombre']
        
        # Contar turnos asignados actualmente (En Atención, Hora Pactada)
        cur.execute("""
            SELECT COUNT(*) as total
            FROM pv_turnos 
            WHERE asesor_id = %s AND estado IN ('En Atención', 'Hora Pactada')
        """, (asesor_id,))
        
        count_row = cur.fetchone()
        stats[asesor_id] = {
            'nombre': asesor_nombre,
            'total': count_row['total'] if count_row else 0
        }
    
    # Asignar al asesor con menos tareas actuales
    mejor_asesor = min(stats.keys(), key=lambda a: stats[a]['total'])
    
    cur.close()
    return mejor_asesor

@app.route('/api/postventa/asignar-turno/<int:turno_id>', methods=['POST'])
@login_required
def asignar_turno_postventa(turno_id):
    """Asignar un turno a un asesor"""
    conn = None
    try:
        data = request.get_json() or {}
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Obtener información del turno
        cur.execute("""
            SELECT fecha_turno, hora_inicio, dominio, cliente_nombre, 
                   servicio_nombre, marca_modelo, sector, servicio_id
            FROM pv_turnos 
            WHERE id = %s AND estado = 'Asignado'
        """, (turno_id,))
        turno_info = cur.fetchone()
        
        if not turno_info:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'Turno no encontrado o ya fue asignado'}), 404
        
        fecha_turno = turno_info['fecha_turno']
        hora_inicio = turno_info['hora_inicio']
        dominio = turno_info['dominio']
        cliente_nombre = turno_info['cliente_nombre']
        servicio_nombre = turno_info['servicio_nombre']
        marca_modelo = turno_info['marca_modelo']
        sector = turno_info['sector']
        servicio_id = turno_info.get('servicio_id')
        
        # Primero verificar si el servicio tiene un asesor fijo asignado
        # Buscar por ID si está disponible, sino por nombre
        if servicio_id:
            cur.execute("""
                SELECT asesor_fijo_id FROM pv_servicios 
                WHERE id = %s AND activo = TRUE
            """, (servicio_id,))
        else:
            cur.execute("""
                SELECT asesor_fijo_id FROM pv_servicios 
                WHERE nombre = %s AND activo = TRUE
            """, (servicio_nombre,))
        servicio_info = cur.fetchone()
        
        # Determinar asesor
        if 'asesor_id' in data and data['asesor_id']:
            # Asignación manual desde el frontend
            asesor_id = data['asesor_id']
        elif servicio_info and servicio_info['asesor_fijo_id']:
            # El servicio tiene un asesor fijo asignado
            asesor_id = servicio_info['asesor_fijo_id']
        else:
            # Asignación automática balanceada (aleatorio inteligente)
            asesor_id = asignar_asesor_balanceado(sector, servicio_nombre, conn)
            
            if not asesor_id:
                cur.close()
                release_db_connection(conn)
                return jsonify({'error': f'No hay asesores disponibles en {sector}'}), 400
        
        # Obtener nombre del asesor para el registro
        cur.execute("SELECT nombre FROM pv_asesores WHERE id = %s", (asesor_id,))
        asesor_nombre = cur.fetchone()['nombre']
        
        # Actualizar turno a estado "En Atención" y guardar asesor original
        cur.execute("""
            UPDATE pv_turnos
            SET asesor_id = %s, 
                asesor_original_id = %s,
                estado = 'En Atención'
            WHERE id = %s
        """, (asesor_id, asesor_id, turno_id))
        
        # Registrar en el historial para reportes
        cur.execute("""
            INSERT INTO pv_historial 
            (turno_id, asesor_nombre, asesor_original_nombre, fecha_turno, dominio, cliente_nombre, 
             servicio_nombre, sector, fecha_registro, fue_reasignado)
            VALUES (%s, %s, %s, %s, %s, %s, %s, %s, NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires', FALSE)
        """, (turno_id, asesor_nombre, asesor_nombre, fecha_turno, dominio, 
              cliente_nombre, servicio_nombre, sector))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({
            'mensaje': 'Turno asignado exitosamente',
            'asesor_id': asesor_id,
            'asesor_nombre': asesor_nombre,
            'estado': 'En Atención'
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en asignar_turno_postventa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/reasignar-turno/<int:turno_id>', methods=['POST'])
@login_required
def reasignar_turno_postventa(turno_id):
    """Reasignar un turno a otro asesor (mantiene fecha_registro original)"""
    conn = None
    try:
        data = request.get_json() or {}
        
        if not data.get('asesor_id'):
            return jsonify({'error': 'Se requiere especificar el nuevo asesor'}), 400
        
        nuevo_asesor_id = data['asesor_id']
        
        conn = get_db_connection()
        cur = conn.cursor()
        
        # Verificar que el turno existe y obtener info
        cur.execute("""
            SELECT t.id, t.asesor_id, t.sector, a.nombre as asesor_actual
            FROM pv_turnos t
            LEFT JOIN pv_asesores a ON t.asesor_id = a.id
            WHERE t.id = %s
        """, (turno_id,))
        turno_info = cur.fetchone()
        
        if not turno_info:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'Turno no encontrado'}), 404
        
        if not turno_info['asesor_id']:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'El turno no tiene asesor asignado aún'}), 400
        
        # Obtener nombre del nuevo asesor y verificar que pertenece al mismo sector
        cur.execute("""
            SELECT nombre, sector FROM pv_asesores 
            WHERE id = %s AND activo = TRUE
        """, (nuevo_asesor_id,))
        nuevo_asesor = cur.fetchone()
        
        if not nuevo_asesor:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'Asesor no encontrado o inactivo'}), 404
        
        if nuevo_asesor['sector'] != turno_info['sector']:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': f'El asesor no pertenece al sector {turno_info["sector"]}'}), 400
        
        # Actualizar asesor_id y marcar como reasignado (mantiene asesor_original_id)
        cur.execute("""
            UPDATE pv_turnos
            SET asesor_id = %s,
                fue_reasignado = TRUE
            WHERE id = %s
        """, (nuevo_asesor_id, turno_id))
        
        # Actualizar el historial para reflejar el nuevo asesor y marcar reasignación
        cur.execute("""
            UPDATE pv_historial
            SET asesor_nombre = %s,
                fue_reasignado = TRUE
            WHERE turno_id = %s
        """, (nuevo_asesor['nombre'], turno_id))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({
            'mensaje': 'Turno reasignado exitosamente',
            'asesor_anterior': turno_info['asesor_actual'],
            'asesor_nuevo': nuevo_asesor['nombre'],
            'asesor_id': nuevo_asesor_id
        })
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en reasignar_turno_postventa: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - TAREAS DEL ASESOR
# ============================================================================

@app.route('/api/postventa/mis-tareas', methods=['GET'])
@login_required
def get_mis_tareas_postventa():
    """Obtener tareas asignadas al usuario actual"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Obtener asesor_id del usuario actual
        cur.execute("""
            SELECT id, sector FROM pv_asesores 
            WHERE usuario_id = %s AND activo = TRUE
        """, (current_user.id,))
        
        asesor_info = cur.fetchone()
        if not asesor_info:
            cur.close()
            release_db_connection(conn)
            return jsonify([])  # Usuario no es asesor
        
        asesor_id, sector = asesor_info
        
        # Obtener tareas asignadas
        cur.execute("""
            SELECT t.id, t.fecha_turno, t.hora_inicio, t.dominio, t.cliente_nombre,
                   t.servicio_nombre, t.marca_modelo, t.estado, t.sector,
                   t.fecha_hora_pactada
            FROM pv_turnos t
            WHERE t.asesor_id = %s AND t.estado != 'Entregado'
            ORDER BY t.fecha_turno, t.hora_inicio
        """, (asesor_id,))
        
        tareas = []
        for row in cur.fetchall():
            tareas.append({
                'id': row['id'],
                'fecha_turno': str(row['fecha_turno']) if row['fecha_turno'] else None,
                'hora_inicio': str(row['hora_inicio']) if row['hora_inicio'] else None,
                'dominio': row['dominio'],
                'cliente_nombre': row['cliente_nombre'],
                'servicio_nombre': row['servicio_nombre'],
                'marca_modelo': row['marca_modelo'],
                'estado': row['estado'],
                'sector': row['sector'],
                'fecha_hora_pactada': str(row['fecha_hora_pactada']) if row['fecha_hora_pactada'] else None
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(tareas)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_mis_tareas_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/atender/<int:turno_id>', methods=['POST'])
@login_required
def atender_turno_postventa(turno_id):
    """Registrar hora de atención de un turno"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Obtener info del turno
        cur.execute("""
            SELECT t.asesor_id, t.servicio_nombre, t.sector, t.fecha_turno, 
                   t.cliente_nombre, t.dominio, a.nombre as asesor_nombre
            FROM pv_turnos t
            LEFT JOIN pv_asesores a ON t.asesor_id = a.id
            WHERE t.id = %s
        """, (turno_id,))
        
        turno_info = cur.fetchone()
        if not turno_info:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'Turno no encontrado'}), 404
        
        # Actualizar hora_atender en pv_turnos
        cur.execute("""
            UPDATE pv_turnos
            SET hora_atender = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires'
            WHERE id = %s
        """, (turno_id,))
        
        # Actualizar/crear registro en pv_historial
        cur.execute("""
            INSERT INTO pv_historial 
            (turno_id, fecha_turno, cliente_nombre, dominio, servicio_nombre, 
             sector, asesor_nombre, hora_atender)
            VALUES (%s, %s, %s, %s, %s, %s, %s, NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires')
            ON CONFLICT (turno_id) 
            DO UPDATE SET hora_atender = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires'
        """, (turno_id, turno_info['fecha_turno'], turno_info['cliente_nombre'],
              turno_info['dominio'], turno_info['servicio_nombre'], 
              turno_info['sector'], turno_info['asesor_nombre']))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Atención registrada exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en atender_turno_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/pactar-hora/<int:turno_id>', methods=['POST'])
@login_required
def pactar_hora_postventa(turno_id):
    """Registrar hora pactada de entrega con el cliente"""
    conn = None
    try:
        data = request.get_json()
        fecha_hora_pactada = data.get('fecha_hora_pactada')
        
        if not fecha_hora_pactada:
            return jsonify({'error': 'Falta fecha_hora_pactada'}), 400
        
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar fecha_hora_pactada en pv_turnos
        cur.execute("""
            UPDATE pv_turnos
            SET fecha_hora_pactada = %s,
                estado = 'Hora Pactada'
            WHERE id = %s
        """, (fecha_hora_pactada, turno_id))
        
        # Actualizar historial con hora_pactada
        cur.execute("""
            UPDATE pv_historial
            SET hora_pactada = %s
            WHERE turno_id = %s
        """, (fecha_hora_pactada, turno_id))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Hora pactada registrada exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en pactar_hora_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/entregar/<int:turno_id>', methods=['POST'])
@login_required
def entregar_vehiculo_postventa(turno_id):
    """Marcar vehículo como entregado al cliente"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar pv_turnos
        cur.execute("""
            UPDATE pv_turnos
            SET estado = 'Entregado',
                hora_entregado = NOW()
            WHERE id = %s
        """, (turno_id,))
        
        # Actualizar historial con hora_entregado y calcular tiempo total
        cur.execute("""
            UPDATE pv_historial h
            SET hora_entregado = NOW(),
                tiempo_total_minutos = EXTRACT(EPOCH FROM (NOW() - h.hora_atender))/60
            WHERE turno_id = %s
        """, (turno_id,))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Vehículo entregado exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en entregar_vehiculo_postventa: {e}")
        return jsonify({'error': str(e)}), 500

# Versiones alternativas de los endpoints que aceptan turno_id en el body
@app.route('/api/postventa/atender', methods=['POST'])
@login_required
def atender_turno_body():
    """Atender un turno - versión que acepta turno_id en body"""
    data = request.get_json() or {}
    turno_id = data.get('turno_id')
    if not turno_id:
        return jsonify({'error': 'turno_id requerido'}), 400
    return atender_turno_postventa(turno_id)

@app.route('/api/postventa/pactar-hora', methods=['POST'])
@login_required
def pactar_hora_body():
    """Pactar hora - versión que acepta turno_id y fecha_hora_pactada en body"""
    data = request.get_json() or {}
    turno_id = data.get('turno_id')
    fecha_hora_pactada = data.get('fecha_hora_pactada')
    observaciones = data.get('observaciones')
    
    if not turno_id:
        return jsonify({'error': 'turno_id requerido'}), 400
    if not fecha_hora_pactada:
        return jsonify({'error': 'fecha_hora_pactada requerido'}), 400
    
    # Llamar directamente a la lógica sin usar el decorador de ruta
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar fecha_hora_pactada y observaciones en pv_turnos
        cur.execute("""
            UPDATE pv_turnos
            SET fecha_hora_pactada = %s,
                estado = 'Hora Pactada',
                observaciones = %s
            WHERE id = %s
        """, (fecha_hora_pactada, observaciones, turno_id))
        
        # Actualizar historial con hora_pactada y hora_pactada_registro, calcular tiempo de atención
        cur.execute("""
            UPDATE pv_historial
            SET hora_pactada = %s,
                hora_pactada_registro = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires',
                tiempo_atencion_minutos = CASE 
                    WHEN hora_atender IS NOT NULL THEN EXTRACT(EPOCH FROM (NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires' - hora_atender))/60
                    ELSE NULL
                END
            WHERE turno_id = %s
        """, (fecha_hora_pactada, turno_id))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Hora pactada registrada exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en pactar_hora_body: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/entregar', methods=['POST'])
@login_required
def entregar_vehiculo_body():
    """Entregar vehículo - versión que acepta turno_id en body"""
    data = request.get_json() or {}
    turno_id = data.get('turno_id')
    if not turno_id:
        return jsonify({'error': 'turno_id requerido'}), 400
    
    # Llamar directamente a la lógica
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar pv_turnos
        cur.execute("""
            UPDATE pv_turnos
            SET estado = 'Entregado',
                hora_entregado = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires'
            WHERE id = %s
        """, (turno_id,))
        
        # Actualizar historial con hora_entregado y calcular tiempos
        cur.execute("""
            UPDATE pv_historial h
            SET hora_entregado = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires',
                tiempo_ejecucion_minutos = CASE
                    WHEN hora_pactada_registro IS NOT NULL THEN EXTRACT(EPOCH FROM (NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires' - h.hora_pactada_registro))/60
                    ELSE NULL
                END,
                tiempo_total_minutos = CASE
                    WHEN hora_atender IS NOT NULL THEN EXTRACT(EPOCH FROM (NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires' - h.hora_atender))/60
                    ELSE NULL
                END
            WHERE turno_id = %s
        """, (turno_id,))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Vehículo entregado exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en entregar_vehiculo_body: {e}")
        return jsonify({'error': str(e)}), 500
    return entregar_vehiculo_postventa(turno_id)

@app.route('/api/postventa/retiro', methods=['POST'])
@login_required
def registrar_retiro():
    """Registrar que el cliente se retira del establecimiento (inicio de cronómetro de egreso)"""
    data = request.get_json() or {}
    turno_id = data.get('turno_id')
    if not turno_id:
        return jsonify({'error': 'turno_id requerido'}), 400
    
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar pv_turnos con hora_retiro_inicio
        cur.execute("""
            UPDATE pv_turnos
            SET hora_retiro_inicio = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires',
                estado = 'Esperando Retiro'
            WHERE id = %s
        """, (turno_id,))
        
        # Actualizar historial
        cur.execute("""
            UPDATE pv_historial
            SET hora_retiro_inicio = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires'
            WHERE turno_id = %s
        """, (turno_id,))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Retiro registrado exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en registrar_retiro: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/atender-retiro', methods=['POST'])
@login_required
def atender_retiro():
    """Atender al cliente cuando regresa a retirar el vehículo (detiene cronómetro de egreso)"""
    data = request.get_json() or {}
    turno_id = data.get('turno_id')
    if not turno_id:
        return jsonify({'error': 'turno_id requerido'}), 400
    
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Actualizar pv_turnos con hora_atender_retiro
        cur.execute("""
            UPDATE pv_turnos
            SET hora_atender_retiro = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires',
                estado = 'Listo para Entrega'
            WHERE id = %s
        """, (turno_id,))
        
        # Actualizar historial y calcular tiempo de espera de egreso
        cur.execute("""
            UPDATE pv_historial
            SET hora_atender_retiro = NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires',
                tiempo_espera_egreso_minutos = CASE 
                    WHEN hora_retiro_inicio IS NOT NULL THEN EXTRACT(EPOCH FROM (NOW() AT TIME ZONE 'America/Argentina/Buenos_Aires' - hora_retiro_inicio))/60
                    ELSE NULL
                END
            WHERE turno_id = %s
        """, (turno_id,))
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Atención de retiro registrada exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en atender_retiro: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/ordenes-trabajo', methods=['GET'])
@login_required
def get_ordenes_trabajo():
    """Obtener todas las órdenes de trabajo asignadas (todos los sectores y asesores)"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Si es usuario simple, solo ver sus propias órdenes
        # Buscar el asesor asociado a este usuario
        if current_user.role == 'user':
            cur.execute("""
                SELECT id FROM pv_asesores WHERE usuario_id = %s AND activo = TRUE
            """, (current_user.id,))
            asesor = cur.fetchone()
            
            if not asesor:
                # Usuario no tiene asesor asignado, no puede ver órdenes
                cur.close()
                release_db_connection(conn)
                return jsonify([])
            
            asesor_id = asesor['id']
            
            # Obtener solo las órdenes asignadas a este asesor
            cur.execute("""
                SELECT t.id, t.fecha_turno, t.hora_inicio, t.dominio, t.cliente_nombre,
                       t.servicio_nombre, t.marca_modelo, t.sector, t.estado,
                       t.fecha_hora_pactada AT TIME ZONE 'America/Argentina/Buenos_Aires' as fecha_hora_pactada,
                       t.hora_atender AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender,
                       t.hora_retiro_inicio AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_retiro_inicio,
                       t.hora_atender_retiro AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender_retiro,
                       t.asesor_id,
                       a.nombre as asesor_nombre,
                       h.fecha_registro as fecha_registro
                FROM pv_turnos t
                LEFT JOIN pv_asesores a ON t.asesor_id = a.id
                LEFT JOIN pv_historial h ON t.id = h.turno_id
                WHERE t.estado IN ('En Atención', 'Hora Pactada', 'Esperando Retiro', 'Listo para Entrega')
                  AND t.asesor_id = %s
                ORDER BY t.fecha_turno, t.hora_inicio
            """, (asesor_id,))
        else:
            # Admin, Usuario+, Supervisor: ver todas las órdenes
            cur.execute("""
                SELECT t.id, t.fecha_turno, t.hora_inicio, t.dominio, t.cliente_nombre,
                       t.servicio_nombre, t.marca_modelo, t.sector, t.estado,
                       t.fecha_hora_pactada AT TIME ZONE 'America/Argentina/Buenos_Aires' as fecha_hora_pactada,
                       t.hora_atender AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender,
                       t.hora_retiro_inicio AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_retiro_inicio,
                       t.hora_atender_retiro AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender_retiro,
                       t.asesor_id,
                       a.nombre as asesor_nombre,
                       h.fecha_registro as fecha_registro
                FROM pv_turnos t
                LEFT JOIN pv_asesores a ON t.asesor_id = a.id
                LEFT JOIN pv_historial h ON t.id = h.turno_id
                WHERE t.estado IN ('En Atención', 'Hora Pactada', 'Esperando Retiro', 'Listo para Entrega')
                ORDER BY t.fecha_turno, t.hora_inicio
            """)
        
        ordenes = []
        for row in cur.fetchall():
            ordenes.append({
                'id': row['id'],
                'fecha_turno': str(row['fecha_turno']) if row['fecha_turno'] else None,
                'hora_inicio': str(row['hora_inicio']) if row['hora_inicio'] else None,
                'dominio': row['dominio'],
                'cliente_nombre': row['cliente_nombre'],
                'servicio_nombre': row['servicio_nombre'],
                'marca_modelo': row['marca_modelo'],
                'sector': row['sector'],
                'estado': row['estado'],
                'fecha_hora_pactada': str(row['fecha_hora_pactada']) if row['fecha_hora_pactada'] else None,
                'hora_atender': str(row['hora_atender']) if row['hora_atender'] else None,
                'hora_retiro_inicio': str(row['hora_retiro_inicio']) if row['hora_retiro_inicio'] else None,
                'hora_atender_retiro': str(row['hora_atender_retiro']) if row['hora_atender_retiro'] else None,
                'asesor_id': row['asesor_id'],
                'asesor_nombre': row['asesor_nombre'] or 'Sin asignar',
                'fecha_registro': str(row['fecha_registro']) if row['fecha_registro'] else None
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(ordenes)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_ordenes_trabajo: {e}")
        return jsonify({'error': str(e)}), 500

# ============================================================================
# API POST VENTA - REPORTES
# ============================================================================

@app.route('/api/postventa/historial', methods=['GET'])
@login_required
def get_historial_postventa():
    """Obtener historial completo con filtros"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Filtros opcionales
        fecha_desde = request.args.get('fecha_desde')
        fecha_hasta = request.args.get('fecha_hasta')
        asesor_nombre = request.args.get('asesor_nombre')
        sector = request.args.get('sector')
        
        # Consultar directamente desde pv_turnos que tiene toda la info
        # Convertir timestamps a zona horaria de Argentina
        query = """
            SELECT t.id, t.fecha_turno, t.hora_inicio, t.dominio, t.cliente_nombre,
                   t.servicio_nombre, t.marca_modelo, t.sector, t.observaciones,
                   t.fecha_creacion,
                   a.nombre as asesor_nombre,
                   t.hora_atender AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender,
                   t.fecha_hora_pactada AT TIME ZONE 'America/Argentina/Buenos_Aires' as fecha_hora_pactada,
                   t.hora_entregado AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_entregado,
                   t.hora_retiro_inicio AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_retiro_inicio,
                   t.hora_atender_retiro AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_atender_retiro,
                   t.fue_reasignado,
                   ao.nombre as asesor_original_nombre,
                   h.fecha_registro as fecha_registro,
                   h.hora_pactada_registro AT TIME ZONE 'America/Argentina/Buenos_Aires' as hora_pactada_registro,
                   h.tiempo_atencion_minutos,
                   h.tiempo_ejecucion_minutos,
                   h.tiempo_total_minutos,
                   h.tiempo_espera_egreso_minutos,
                   t.estado
            FROM pv_turnos t
            LEFT JOIN pv_asesores a ON t.asesor_id = a.id
            LEFT JOIN pv_asesores ao ON t.asesor_original_id = ao.id
            LEFT JOIN pv_historial h ON t.id = h.turno_id
            WHERE t.estado IN ('En Atención', 'Hora Pactada', 'Entregado')
        """
        params = []
        
        if fecha_desde:
            query += " AND t.fecha_turno >= %s"
            params.append(fecha_desde)
        
        if fecha_hasta:
            query += " AND t.fecha_turno <= %s"
            params.append(fecha_hasta)
        
        if asesor_nombre:
            query += " AND a.nombre = %s"
            params.append(asesor_nombre)
        
        if sector:
            query += " AND t.sector = %s"
            params.append(sector)
        
        query += " ORDER BY t.fecha_turno DESC, t.hora_inicio DESC"
        
        cur.execute(query, params)
        
        historial = []
        for row in cur.fetchall():
            historial.append({
                'id': row['id'],
                'fecha_turno': str(row['fecha_turno']) if row['fecha_turno'] else None,
                'hora_inicio': str(row['hora_inicio']) if row.get('hora_inicio') else None,
                'fecha_creacion': str(row['fecha_creacion']) if row.get('fecha_creacion') else None,
                'fecha_registro': str(row['fecha_registro']) if row.get('fecha_registro') else None,
                'dominio': row['dominio'],
                'cliente_nombre': row['cliente_nombre'],
                'servicio_nombre': row['servicio_nombre'],
                'marca_modelo': row['marca_modelo'],
                'sector': row['sector'],
                'observaciones': row['observaciones'],
                'asesor_nombre': row['asesor_nombre'] or 'Sin asignar',
                'hora_atender': str(row['hora_atender']) if row['hora_atender'] else None,
                'fecha_hora_pactada': str(row['fecha_hora_pactada']) if row['fecha_hora_pactada'] else None,
                'hora_pactada': str(row['fecha_hora_pactada']) if row['fecha_hora_pactada'] else None,
                'hora_pactada_registro': str(row['hora_pactada_registro']) if row.get('hora_pactada_registro') else None,
                'hora_entregado': str(row['hora_entregado']) if row['hora_entregado'] else None,
                'hora_retiro_inicio': str(row['hora_retiro_inicio']) if row.get('hora_retiro_inicio') else None,
                'hora_atender_retiro': str(row['hora_atender_retiro']) if row.get('hora_atender_retiro') else None,
                'fue_reasignado': row.get('fue_reasignado', False),
                'asesor_original_nombre': row.get('asesor_original_nombre'),
                'tiempo_atencion_minutos': int(row['tiempo_atencion_minutos']) if row.get('tiempo_atencion_minutos') else None,
                'tiempo_ejecucion_minutos': int(row['tiempo_ejecucion_minutos']) if row.get('tiempo_ejecucion_minutos') else None,
                'tiempo_total_minutos': int(row['tiempo_total_minutos']) if row.get('tiempo_total_minutos') else None,
                'tiempo_espera_egreso_minutos': int(row['tiempo_espera_egreso_minutos']) if row.get('tiempo_espera_egreso_minutos') else None,
                'estado': row['estado']
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(historial)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_historial_postventa: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/horarios-pactados', methods=['GET'])
@login_required
def get_horarios_pactados():
    """Obtener todos los turnos con hora pactada, ordenados por horario"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        query = """
            SELECT t.id, t.dominio, t.cliente_nombre, t.marca_modelo,
                   t.fecha_hora_pactada AT TIME ZONE 'America/Argentina/Buenos_Aires' as fecha_hora_pactada,
                   a.nombre as asesor_nombre,
                   t.estado
            FROM pv_turnos t
            LEFT JOIN pv_asesores a ON t.asesor_id = a.id
            WHERE t.fecha_hora_pactada IS NOT NULL
              AND t.estado IN ('Hora Pactada', 'Esperando Retiro', 'Listo para Entrega', 'Entregado')
            ORDER BY t.fecha_hora_pactada ASC
        """
        
        cur.execute(query)
        
        horarios = []
        for row in cur.fetchall():
            horarios.append({
                'id': row['id'],
                'dominio': row['dominio'],
                'cliente_nombre': row['cliente_nombre'],
                'marca_modelo': row['marca_modelo'],
                'fecha_hora_pactada': str(row['fecha_hora_pactada']) if row['fecha_hora_pactada'] else None,
                'asesor_nombre': row['asesor_nombre'] or 'Sin asignar',
                'estado': row['estado']
            })
        
        cur.close()
        release_db_connection(conn)
        return jsonify(horarios)
        
    except Exception as e:
        if conn:
            release_db_connection(conn)
        print(f"Error en get_horarios_pactados: {e}")
        return jsonify({'error': str(e)}), 500

@app.route('/api/postventa/turno/<int:turno_id>', methods=['DELETE'])
@login_required
def delete_turno_postventa(turno_id):
    """Eliminar un turno del registro"""
    conn = None
    try:
        conn = get_pg_connection()
        cur = conn.cursor()
        
        # Primero eliminar de pv_historial (por foreign key)
        cur.execute("DELETE FROM pv_historial WHERE turno_id = %s", (turno_id,))
        
        # Luego eliminar de pv_turnos
        cur.execute("DELETE FROM pv_turnos WHERE id = %s", (turno_id,))
        
        if cur.rowcount == 0:
            cur.close()
            release_db_connection(conn)
            return jsonify({'error': 'Turno no encontrado'}), 404
        
        conn.commit()
        cur.close()
        release_db_connection(conn)
        
        return jsonify({'mensaje': 'Turno eliminado exitosamente'})
        
    except Exception as e:
        if conn:
            conn.rollback()
            release_db_connection(conn)
        print(f"Error en delete_turno_postventa: {e}")
        return jsonify({'error': str(e)}), 500


if __name__ == '__main__':
    # Para Hugging Face Spaces, usar puerto 7860
    port = int(os.environ.get('PORT', 7860))
    app.run(debug=False, host='0.0.0.0', port=port)
